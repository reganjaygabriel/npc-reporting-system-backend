from rest_framework import viewsets, status, mixins
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.http import FileResponse
from django.db.models import Sum, Avg, Q
from datetime import datetime
import hashlib
import os
import tempfile
import json

from .models import Plant, Unit, UploadedFile, GenerationReport, PlantCapacity, HistoricalData, WaterNomination, ActualGeneration, Testimonial, AuditLog, ESignature, ReportSignature, MonthlyTarget
from .serializers import (
    PlantSerializer, UnitSerializer, UploadedFileSerializer,
    GenerationReportSerializer, GenerationReportListSerializer,
    ExcelUploadSerializer, ReportGenerationSerializer,
    PlantCapacitySerializer, HistoricalDataSerializer, HistoricalDataUploadSerializer,
    WaterNominationSerializer, ActualGenerationSerializer, NominationVarianceSerializer,
    TestimonialSerializer, AuditLogSerializer, ESignatureSerializer, ReportSignatureSerializer,
    ESignatureCreateSerializer, MonthlyTargetSerializer
)
from .pagination import CustomPageNumberPagination
from .services.excel_importer import ExcelImporter
from .services.psr_exporter import PSRExporter
from .services.historical_data_importer import HistoricalDataImporter
from .services.template_generator import TemplateGenerator
from .services.daily_status_exporter import generate_daily_status_report
from .utils import get_location_from_ip, get_client_ip
from .audit_utils import (
    AuditLogger, audit_action, AuditContext, audit_file_upload, 
    audit_report_generation, audit_signature_creation, audit_data_export
)


class PlantViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Plant.objects.filter(is_active=True)
    serializer_class = PlantSerializer
    permission_classes = [AllowAny]  # Allow unauthenticated access for internal system
    pagination_class = None  # Disable pagination for plants


class UnitViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Unit.objects.filter(is_active=True).select_related('plant')
    serializer_class = UnitSerializer
    permission_classes = [AllowAny]  # Allow unauthenticated access for internal system
    
    def get_queryset(self):
        queryset = super().get_queryset()
        plant_code = self.request.query_params.get('plant_code')
        if plant_code:
            queryset = queryset.filter(plant__code=plant_code)
        return queryset


class UploadedFileViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = UploadedFile.objects.filter(is_archived=False).select_related('plant', 'uploaded_by')
    serializer_class = UploadedFileSerializer
    permission_classes = [AllowAny]  # Allow unauthenticated access for internal system
    authentication_classes = []  # Disable authentication to avoid CSRF issues
    
    @action(detail=False, methods=['get'], url_path='download-template/daily-generation')
    def download_daily_generation_template(self, request):
        """Download Daily Generation Report template"""
        wb = TemplateGenerator.generate_daily_generation_template()
        return TemplateGenerator.create_http_response(wb, 'Daily_Generation_Template.xlsx')
    
    @action(detail=False, methods=['get'], url_path='download-template/water-nomination')
    def download_water_nomination_template(self, request):
        """Download Water Nomination template"""
        wb = TemplateGenerator.generate_water_nomination_template()
        return TemplateGenerator.create_http_response(wb, 'Water_Nomination_Template.xlsx')
    
    @action(detail=False, methods=['get'], url_path='download-template/historical-data')
    def download_historical_data_template(self, request):
        """Download Historical Data Import template"""
        wb = TemplateGenerator.generate_historical_data_template()
        return TemplateGenerator.create_http_response(wb, 'Historical_Data_Template.xlsx')
    
    @action(detail=False, methods=['get'], url_path='download-template/plant-capacity')
    def download_plant_capacity_template(self, request):
        """Download Plant Capacity template"""
        wb = TemplateGenerator.generate_plant_capacity_template()
        return TemplateGenerator.create_http_response(wb, 'Plant_Capacity_Template.xlsx')
    
    @action(detail=False, methods=['get'], url_path='download-template/plant-status')
    def download_plant_status_template(self, request):
        """Download Plant Status template"""
        wb = TemplateGenerator.generate_plant_status_template()
        return TemplateGenerator.create_http_response(wb, 'Plant_Status_Template.xlsx')
    
    @action(detail=False, methods=['get'], url_path='download-template/psr')
    def download_psr_template(self, request):
        """Download PSR (Plant Status Report) template with right side section"""
        wb = TemplateGenerator.generate_psr_template()
        return TemplateGenerator.create_http_response(wb, 'PSR_Template.xlsx')
    
    @action(detail=True, methods=['delete'])
    def delete_upload(self, request, pk=None):
        """Delete an uploaded file and all its associated generation reports"""
        try:
            # Get the file without the is_archived filter
            uploaded_file = UploadedFile.objects.get(pk=pk)
            filename = uploaded_file.original_filename
            
            # Delete associated generation reports first
            deleted_reports = GenerationReport.objects.filter(uploaded_file=uploaded_file).delete()
            
            # Delete the uploaded file record and physical file
            file_path = uploaded_file.file.path if uploaded_file.file else None
            uploaded_file.delete()
            
            # Try to delete physical file
            if file_path:
                try:
                    import os
                    if os.path.exists(file_path):
                        os.remove(file_path)
                except Exception as e:
                    print(f"Warning: Could not delete physical file: {e}")
            
            # Log file deletion
            AuditLogger.log_user_action(
                user=request.user,
                action='FILE_DELETE',
                description=f'Deleted uploaded file: {filename} and {deleted_reports[0] if deleted_reports else 0} associated reports',
                model_name='UploadedFile',
                object_id=pk,
                category='file_management',
                severity='HIGH',
                request=request
            )
            
            return Response({
                'message': 'File and associated records deleted successfully',
                'reports_deleted': deleted_reports[0] if deleted_reports else 0
            }, status=status.HTTP_200_OK)
            
        except UploadedFile.DoesNotExist:
            AuditLogger.log_user_action(
                user=request.user,
                action='FILE_DELETE',
                description=f'Failed to delete file: File not found (ID: {pk})',
                category='file_management',
                severity='LOW',
                success=False,
                error_message='File not found',
                request=request
            )
            return Response({'error': 'File not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            AuditLogger.log_user_action(
                user=request.user,
                action='FILE_DELETE',
                description=f'Failed to delete file (ID: {pk}): {str(e)}',
                category='file_management',
                severity='MEDIUM',
                success=False,
                error_message=str(e),
                request=request
            )
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)    
    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        """Archive an uploaded file"""
        try:
            from django.utils import timezone
            # Get the file without the is_archived filter
            uploaded_file = UploadedFile.objects.get(pk=pk)
            
            if uploaded_file.is_archived:
                return Response({'error': 'File is already archived'}, status=status.HTTP_400_BAD_REQUEST)
            
            uploaded_file.is_archived = True
            uploaded_file.archived_at = timezone.now()
            uploaded_file.archived_by = request.user if request.user.is_authenticated else None
            uploaded_file.save()
            
            # Log the action
            if request.user.is_authenticated:
                AuditLogger.log_user_action(
                    user=request.user,
                    action='FILE_ARCHIVE',
                    model_name='UploadedFile',
                    object_id=uploaded_file.id,
                    description=f'Archived file: {uploaded_file.original_filename}',
                    category='file_management',
                    severity='MEDIUM',
                    request=request
                )
            
            return Response({
                'message': 'File archived successfully',
                'file_id': uploaded_file.id
            }, status=status.HTTP_200_OK)
            
        except UploadedFile.DoesNotExist:
            return Response({'error': 'File not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        """Restore an archived file"""
        try:
            # Get the file without the is_archived filter
            uploaded_file = UploadedFile.objects.get(pk=pk)
            
            if not uploaded_file.is_archived:
                return Response({'error': 'File is not archived'}, status=status.HTTP_400_BAD_REQUEST)
            
            uploaded_file.is_archived = False
            uploaded_file.archived_at = None
            uploaded_file.archived_by = None
            uploaded_file.save()
            
            # Log the action
            if request.user.is_authenticated:
                AuditLogger.log_user_action(
                    user=request.user,
                    action='FILE_RESTORE',
                    model_name='UploadedFile',
                    object_id=uploaded_file.id,
                    description=f'Restored file: {uploaded_file.original_filename}',
                    category='file_management',
                    severity='MEDIUM',
                    request=request
                )
            
            return Response({
                'message': 'File restored successfully',
                'file_id': uploaded_file.id
            }, status=status.HTTP_200_OK)
            
        except UploadedFile.DoesNotExist:
            return Response({'error': 'File not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def archived(self, request):
        """Get all archived files"""
        archived_files = UploadedFile.objects.filter(is_archived=True).select_related('plant', 'uploaded_by', 'archived_by')
        page = self.paginate_queryset(archived_files)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(archived_files, many=True)
        return Response(serializer.data)

    
    @action(detail=False, methods=['post'])
    @audit_action('FILE_UPLOAD', 'File upload and processing', category='file_management', severity='MEDIUM')
    def upload(self, request):
        serializer = ExcelUploadSerializer(data=request.data)
        if not serializer.is_valid():
            AuditLogger.log_user_action(
                user=request.user,
                action='FILE_UPLOAD',
                description='File upload failed: Invalid data',
                category='file_management',
                severity='LOW',
                success=False,
                error_message=str(serializer.errors),
                request=request
            )
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        file = serializer.validated_data['file']
        plant_code = serializer.validated_data['plant_code']
        
        try:
            plant = Plant.objects.get(code=plant_code)
        except Plant.DoesNotExist:
            AuditLogger.log_user_action(
                user=request.user,
                action='FILE_UPLOAD',
                description=f'File upload failed: Plant {plant_code} not found',
                category='file_management',
                severity='LOW',
                success=False,
                error_message='Plant not found',
                request=request
            )
            return Response({'error': 'Plant not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Calculate checksum
        file.seek(0)
        checksum = hashlib.sha256(file.read()).hexdigest()
        file.seek(0)
        
        # Check for duplicate
        if UploadedFile.objects.filter(checksum=checksum, plant=plant).exists():
            AuditLogger.log_user_action(
                user=request.user,
                action='FILE_UPLOAD',
                description=f'File upload rejected: Duplicate file for plant {plant_code}',
                category='file_management',
                severity='LOW',
                success=False,
                error_message='Duplicate file',
                request=request
            )
            return Response({'error': 'This file has already been uploaded'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Create uploaded file record
        uploaded_file = UploadedFile.objects.create(
            file=file,
            original_filename=file.name,
            plant=plant,
            uploaded_by=request.user if request.user.is_authenticated else None,
            file_size=file.size,
            checksum=checksum,
            status='PROCESSING'
        )
        
        # Log successful file upload
        audit_file_upload(
            user=request.user,
            filename=file.name,
            file_size=file.size,
            request=request
        )
        
        # Process the file
        try:
            with AuditContext(request.user, 'FILE_PROCESSING', f'Processing uploaded file: {file.name}'):
                importer = ExcelImporter(uploaded_file)
                records_imported = importer.process()
                
                uploaded_file.status = 'COMPLETED'
                uploaded_file.records_imported = records_imported
                uploaded_file.save()
                
                # Log successful processing
                AuditLogger.log_user_action(
                    user=request.user,
                    action='FILE_UPLOAD',
                    description=f'File processed successfully: {records_imported} records imported from {file.name}',
                    model_name='UploadedFile',
                    category='data_processing',
                    severity='MEDIUM',
                    request=request
                )
            
            return Response({
                'message': 'File uploaded and processed successfully',
                'records_imported': records_imported,
                'file_id': uploaded_file.id
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            uploaded_file.status = 'FAILED'
            uploaded_file.error_message = str(e)
            uploaded_file.save()
            
            # Log processing failure
            AuditLogger.log_user_action(
                user=request.user,
                action='FILE_UPLOAD',
                description=f'File processing failed for {file.name}: {str(e)}',
                category='file_management',
                severity='HIGH',
                success=False,
                error_message=str(e),
                request=request
            )
            
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class GenerationReportViewSet(mixins.ListModelMixin,
                              mixins.RetrieveModelMixin,
                              viewsets.GenericViewSet):
    queryset = GenerationReport.objects.all().select_related('plant', 'unit', 'uploaded_file')
    permission_classes = [AllowAny]  # Allow unauthenticated access for internal system
    authentication_classes = []  # Disable authentication to avoid CSRF issues
    
    def get_serializer_class(self):
        if self.action == 'list':
            return GenerationReportListSerializer
        return GenerationReportSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by plant - handle both plant_code and plant_code[] formats
        plant_codes = self.request.query_params.getlist('plant_code[]') or self.request.query_params.getlist('plant_code')
        print(f"DEBUG: plant_codes from getlist = {plant_codes}")
        print(f"DEBUG: query_params = {dict(self.request.query_params)}")
        
        if plant_codes:
            queryset = queryset.filter(plant__code__in=plant_codes)
            print(f"DEBUG: Filtered queryset count = {queryset.count()}")
        else:
            print(f"DEBUG: NO plant_codes filter, returning all {queryset.count()} records")
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(report_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(report_date__lte=end_date)
        
        # Filter by unit
        unit_id = self.request.query_params.get('unit_id')
        if unit_id:
            queryset = queryset.filter(unit_id=unit_id)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get aggregated summary statistics"""
        queryset = self.get_queryset()
        
        summary = queryset.aggregate(
            total_generation=Sum('generation_kwh'),
            avg_capacity_factor=Avg('capacity_factor'),
            avg_availability_factor=Avg('availability_factor'),
            total_operating_hours=Sum('operating_hours'),
            total_forced_outage_hours=Sum('forced_outage_hours')
        )
        
        return Response(summary)
    
    @action(detail=False, methods=['post'], url_path='generate-report')
    @audit_action('REPORT_GENERATE', 'Excel report generation', category='reporting', severity='MEDIUM')
    def generate_report(self, request):
        """Generate Excel report based on filters"""
        serializer = ReportGenerationSerializer(data=request.data)
        if not serializer.is_valid():
            AuditLogger.log_user_action(
                user=request.user,
                action='REPORT_GENERATE',
                description='Report generation failed: Invalid parameters',
                category='reporting',
                severity='LOW',
                success=False,
                error_message=str(serializer.errors),
                request=request
            )
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        report_type = data.get('report_type', 'psr')
        
        # Get filtered data
        reports = GenerationReport.objects.filter(
            plant__code__in=data['plant_codes'],
            report_date__gte=data['start_date'],
            report_date__lte=data['end_date']
        ).select_related('plant', 'unit').order_by('report_date', 'plant', 'unit')
        
        if not reports.exists():
            AuditLogger.log_user_action(
                user=request.user,
                action='REPORT_GENERATE',
                description=f'Report generation failed: No data found for criteria - Plants: {data["plant_codes"]}, Date range: {data["start_date"]} to {data["end_date"]}',
                category='reporting',
                severity='LOW',
                success=False,
                error_message='No data found',
                request=request
            )
            return Response({'error': 'No data found for the specified criteria'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        # Generate Excel file
        try:
            report_date = data['start_date']
            plant_names = ', '.join([code for code in data['plant_codes']])
            
            with AuditContext(request.user, 'REPORT_GENERATION', f'Generating {report_type} report for {plant_names}'):
                # Both daily_status and PSR now use the same PSR template
                # Pass report_type to customize header styling
                exporter = PSRExporter(reports, report_date, report_type=report_type)
                file_path = exporter.generate()
                
                if report_type == 'daily_status':
                    filename = f"DAILY_PLANT_STATUS_{report_date.strftime('%Y%m%d')}.xlsx"
                    report_name = "Daily Plant Status Report"
                else:
                    filename = f"PSR_REPORT_{report_date.strftime('%Y%m%d')}.xlsx"
                    report_name = "PSR Report"
                
                # Comprehensive audit logging for report generation
                audit_report_generation(
                    user=request.user,
                    report_date=report_date.strftime('%Y-%m-%d'),
                    report_type=report_type,
                    request=request
                )
                
                # Additional detailed logging
                AuditLogger.log_user_action(
                    user=request.user,
                    action='REPORT_GENERATE',
                    description=f'Successfully generated {report_name} - Plants: {plant_names}, Date: {report_date.strftime("%Y-%m-%d")}, Records: {reports.count()}',
                    model_name='GenerationReport',
                    category='reporting',
                    severity='MEDIUM',
                    request=request
                )
                
                # Log data export
                audit_data_export(
                    user=request.user,
                    data_type=report_name,
                    record_count=reports.count(),
                    request=request
                )
            
            response = FileResponse(
                open(file_path, 'rb'),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            # Log generation failure
            AuditLogger.log_user_action(
                user=request.user,
                action='REPORT_GENERATE',
                description=f'Report generation failed for {report_type}: {str(e)}',
                category='reporting',
                severity='HIGH',
                success=False,
                error_message=str(e),
                request=request
            )
            
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], url_path='preview-report')
    def preview_report(self, request):
        """Preview report data in exact Excel format structure"""
        serializer = ReportGenerationSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        
        # Get filtered data
        reports = GenerationReport.objects.filter(
            plant__code__in=data['plant_codes'],
            report_date__gte=data['start_date'],
            report_date__lte=data['end_date']
        ).select_related('plant', 'unit').order_by('report_date', 'plant', 'unit')
        
        if not reports.exists():
            return Response({'error': 'No data found for the specified criteria'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        # Use PSRExporter to get the exact Excel structure
        try:
            from .services.psr_exporter import PSRExporter
            
            exporter = PSRExporter(reports, data['start_date'], data.get('report_type', 'psr'))
            
            # Get organized data by plant
            organized_data = exporter._organize_data()
            
            # Build Excel-like structure for preview with ALL PSR sections
            preview_structure = {
                'header': {
                    'title': 'PLANT STATUS REPORT' if data.get('report_type', 'psr') == 'psr' else 'DAILY PLANT STATUS',
                    'subtitle': 'MINDANAO GENERATION',
                    'portfolio': '(PSALM PORTFOLIO)',
                    'date_text': f'as of 0800H {data["start_date"].strftime("%A, %d %B %Y")}' if data.get('report_type', 'psr') == 'psr' else f'as of 12:00 NN    {data["start_date"].strftime("%A, %B %d, %Y")}',
                    'report_date': data['start_date'].strftime('%Y-%m-%d')
                },
                'plants_data': [],
                'totals': {
                    'total_capacity': 0,
                    'total_nominated': 0,
                    'total_actual': 0,
                    'total_variance': 0
                },
                'forecasted_load': {
                    'date': 'Jan 02, 2026',
                    'agus_load': 500.8,
                    'pulangi_load': 150,
                    'total_load': 650.8,
                    'mindanao_load': 1850.0,
                    'luzon_visayas': 8500.0,
                    'total_philippines': 10350.0
                },
                'ipp_data': [
                    {'name': 'STEAG', 'capacity': 232, 'nominated': 200, 'actual': 195, 'variance': -5},
                    {'name': 'THERMA SOUTH', 'capacity': 300, 'nominated': 280, 'actual': 275, 'variance': -5},
                    {'name': 'THERMA MOBILE', 'capacity': 100, 'nominated': 90, 'actual': 85, 'variance': -5},
                    {'name': 'FDC MISAMIS', 'capacity': 405, 'nominated': 380, 'actual': 370, 'variance': -10},
                    {'name': 'PALM CONCEPCION', 'capacity': 135, 'nominated': 120, 'actual': 115, 'variance': -5}
                ],
                'notes': [
                    'Dependable Capacity (DC) is the maximum capacity, modified for ambient limitations for a specific period of time, such as month or a season.',
                    'Available Capacity (AC) is the dependable capacity, modified for equipment limitations for any time.',
                    'The usual occurrence of Peak is at 1800H.',
                    'AGUS 5 HEP gate no. 2 dogged at 0.10m for Newtech Pulp Inc. (NPI) plant water use.',
                    'AGUS 6 HEP unit 5 is derated at 46 MW due generator turbine control system problem.',
                    'AGUS 6 HEP unit 3 is derated at 40 MW due to take-off transformer cooling problem.'
                ],
                # NEW: Right-side sections from PSR Excel
                'storage_data': [
                    {'lake': 'Lake Lanao', 'level': '701.20', 'remarks': 'Normal'},
                    {'lake': 'Agus 2 Forebay', 'level': '637.30', 'remarks': ''},
                    {'lake': 'Agus 4 Forebay', 'level': '358.50', 'remarks': ''},
                    {'lake': 'Agus 5 Forebay', 'level': '242.80', 'remarks': ''},
                    {'lake': 'Agus 6 Forebay', 'level': '199.80', 'remarks': ''},
                    {'lake': 'Agus 7 Forebay', 'level': '34.60', 'remarks': ''},
                    {'lake': 'Pulangi IV Reservoir', 'level': '283.50', 'remarks': ''},
                ],
                'inflow_outflow_data': [
                    {'plant': 'Lake Lanao', 'inflow': '0.00', 'outflow': '0.00', 'remarks': ''},
                    {'plant': 'Agus 1', 'inflow': '0.00', 'outflow': '0.00', 'remarks': ''},
                    {'plant': 'Agus 2', 'inflow': '0.00', 'outflow': '0.00', 'remarks': ''},
                    {'plant': 'Agus 4', 'inflow': '0.00', 'outflow': '0.00', 'remarks': ''},
                    {'plant': 'Agus 5', 'inflow': '0.00', 'outflow': '0.00', 'remarks': ''},
                    {'plant': 'Agus 6', 'inflow': '0.00', 'outflow': '0.00', 'remarks': ''},
                    {'plant': 'Agus 7', 'inflow': '0.00', 'outflow': '0.00', 'remarks': ''},
                    {'plant': 'Pulangi IV', 'inflow': '0.00', 'outflow': '0.00', 'remarks': ''},
                ],
                'generation_data': exporter._calculate_generation_data(),
                'capacity_factor_data': exporter._calculate_capacity_factor(),
                'gate_elevation_data': [
                    {'plant': 'Lake Lanao', 'gates': ['0.100', '0.100'], 'elevation': '701.190'},
                    {'plant': 'Agus 2', 'gates': ['0.000', '0.000'], 'elevation': '637.800', 'note': 'Mr. Dennis'},
                    {'plant': 'Agus 4', 'gates': ['0.500', '0.000'], 'elevation': '358.800'},
                    {'plant': 'Agus 5', 'gates': ['0.550', '0.000', '0.100'], 'elevation': '243.300'},
                    {'plant': 'Agus 6', 'gates': ['0.200', '0.200', '0.200', '0.000'], 'elevation': '199.800'},
                    {'plant': 'Agus 7', 'gates': ['0.000', '0.000', '0.000'], 'elevation': '34.100'},
                    {'plant': 'Pulangi IV', 'gates': ['0.000', '0.000', '0.000', '0.000', '0.000', '0.000'], 'elevation': '285.450'},
                ],
                'operational_reference': {
                    'riparian_flow': [
                        {'plant': 'AGUS 1', 'cms_mv': '0.85', 'cms': '85.0'},
                        {'plant': 'AGUS 2', 'cms_mv': '0.90', 'cms': '90.0'},
                        {'plant': 'AGUS 3', 'cms_mv': '0.88', 'cms': '88.0'},
                        {'plant': 'AGUS 4', 'cms_mv': '0.92', 'cms': '92.0'},
                        {'plant': 'AGUS 5', 'cms_mv': '0.87', 'cms': '87.0'},
                        {'plant': 'AGUS 6', 'cms_mv': '0.89', 'cms': '89.0'},
                        {'plant': 'AGUS 7', 'cms_mv': '0.91', 'cms': '91.0'},
                    ],
                    'spillage_data': [
                        {'location': 'Lake Lanao', 'value': '0.00'},
                        {'location': 'Agus 2', 'value': '0.00'},
                        {'location': 'Agus 4', 'value': '0.00'},
                        {'location': 'Agus 5', 'value': '0.00'},
                        {'location': 'Agus 6', 'value': '0.00'},
                        {'location': 'Agus 7', 'value': '0.00'},
                        {'location': 'Pulangi IV', 'value': '0.00'},
                    ]
                },
                'input_workflow': {
                    'sections': ['INPUT 2', 'PROCESS', 'OUTPUT'],
                    'notes': ['Dependable Capacity = Pmax', 'MLRD GATES OPENING: Status: Operational']
                },
                'signatures': {
                    'first_row': [
                        {
                            'role': 'Prepared by:',
                            'name': 'O.M. LAVA',
                            'title': 'Prin. Engr. A, GPD'
                        },
                        {
                            'role': 'Checked and Reviewed by:',
                            'name': 'JMM MATA',
                            'title': 'Manager, GPD'
                        },
                        {
                            'role': 'Checked and Reviewed by:',
                            'name': 'EL ADIONG',
                            'title': 'Acting Manager, GPD'
                        },
                        {
                            'role': 'Approved by:',
                            'name': 'C.C. AMIGABLE JR.',
                            'title': 'Dept. Manager, GPD'
                        }
                    ],
                    'second_row': [
                        {
                            'role': 'Prepared by:',
                            'name': 'D.R.B. CAIRO',
                            'title': 'Prin. Engr. B, GPD'
                        },
                        {
                            'role': 'Checked and Reviewed by:',
                            'name': 'JMM MATA',
                            'title': 'Manager, GPD'
                        },
                        {
                            'role': 'Checked and Reviewed by:',
                            'name': 'EL ADIONG',
                            'title': 'OIC-Dept Manager, GPD'
                        },
                        {
                            'role': 'Approved by:',
                            'name': 'DB ESMADE JR.',
                            'title': 'Acting Dept. Manager, GPD'
                        }
                    ]
                },
                'footer_note': 'Agus 2 HEP is limited to 40 MW/per unit due to water constraints as per Environmental Compliance Certificate "that Agus 2 shall not be operated at full capacity..." This is to prevent risk of flooding at lakeshore areas and Baloi plains.',
                'additional_notes': [
                    'The Available Capacity in this report includes equipment limitation and water outflow consideration based on the 2020 Lake Lanao Operating Guide Curve.',
                    'AGUS 6 HEP units 1 & 2 up-rated from 25MW to 34.5MW. Turned-over to NPC last 14 February 2020.',
                    'AGUS 2 HEP is limited to 120 MW total load to prevent risk of flooding at lakeshore areas and Baloi plains as per Environmental Compliance Certificate dated January 14, 1992.',
                    'AGUS 5 HEP gate no. 2 dogged at 0.10m for Newtech Pulp Inc. (NPI) plant water use.',
                    'The usual occurrence of Peak is at 1800H.',
                    'Forecast inflow of Lake Lanao is stable and operating at Normal Stage.',
                    'Agus 6 HEP: 18cms of water spilled due to partially opened spillway gate no. 1.'
                ]
            }
            
            # Process each plant according to PSR structure
            for plant_code in exporter.PLANTS_CONFIG:
                if plant_code in data['plant_codes']:
                    plant_config = exporter.PLANTS_CONFIG[plant_code]
                    plant_data = organized_data.get(plant_code, {})
                    
                    plant_info = {
                        'code': plant_code,
                        'name': plant_config['name'],
                        'units': [],
                        'plant_totals': {
                            'capacity': 0,
                            'nominated': 0,
                            'actual': 0,
                            'variance': 0
                        }
                    }
                    
                    # Process each unit
                    for unit_config in plant_config['units']:
                        unit_num = unit_config['num']
                        unit_report_data = plant_data.get(unit_num, {})
                        
                        # Calculate actual generation (convert from kWh to MW if needed)
                        actual_generation = unit_report_data.get('generation', 0) / 1000  # Convert kWh to MWh, then approximate MW
                        if actual_generation > 0:
                            actual_generation = min(actual_generation / 24, unit_config['capacity'])  # Rough MW calculation
                        
                        # For the new PSR format, we need different data mapping
                        # Available Capacity = Nominated capacity
                        # Lake Lanao Projected Ave. Outflow = Actual generation
                        # Load at 0800H = Variance (or could be a different calculation)
                        
                        unit_info = {
                            'number': unit_num,
                            'label': unit_config['label'],
                            'capacity': unit_config['capacity'],  # Rated Capacity
                            'nominated': unit_config['nominated'],  # Available Capacity
                            'actual': round(actual_generation, 1),  # Lake Lanao Projected Ave. Outflow
                            'variance': round(unit_config['nominated'] - actual_generation, 1),  # Load at 0800H (difference)
                            'operating_hours': unit_report_data.get('operating_hours', 0),
                            'forced_outage': unit_report_data.get('forced_outage', 0),
                            'scheduled_outage': unit_report_data.get('scheduled_outage', 0),
                            'remarks': unit_report_data.get('remarks', '') or (
                                'Lake Lanao Elevation is 701.19 m.a.s.l. (G1- 0.10 m, G2- 0.10 m)' 
                                if plant_code == 'AGUS1' else ''
                            )
                        }
                        
                        plant_info['units'].append(unit_info)
                        
                        # Add to plant totals
                        plant_info['plant_totals']['capacity'] += unit_config['capacity']
                        plant_info['plant_totals']['nominated'] += unit_config['nominated']
                        plant_info['plant_totals']['actual'] += actual_generation
                        plant_info['plant_totals']['variance'] += (unit_config['nominated'] - actual_generation)
                    
                    # Round plant totals
                    for key in plant_info['plant_totals']:
                        if key != 'capacity':  # capacity is already integer
                            plant_info['plant_totals'][key] = round(plant_info['plant_totals'][key], 1)
                    
                    preview_structure['plants_data'].append(plant_info)
                    
                    # Add to grand totals
                    preview_structure['totals']['total_capacity'] += plant_info['plant_totals']['capacity']
                    preview_structure['totals']['total_nominated'] += plant_info['plant_totals']['nominated']
                    preview_structure['totals']['total_actual'] += plant_info['plant_totals']['actual']
                    preview_structure['totals']['total_variance'] += plant_info['plant_totals']['variance']
            
            # Round grand totals
            for key in preview_structure['totals']:
                if key != 'total_capacity':
                    preview_structure['totals'][key] = round(preview_structure['totals'][key], 1)
            
            return Response(preview_structure)
            
        except Exception as e:
            return Response({'error': f'Failed to generate preview: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HistoricalDataViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for historical data"""
    queryset = HistoricalData.objects.all().select_related('plant')
    serializer_class = HistoricalDataSerializer
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by plant
        plant_codes = self.request.query_params.getlist('plant_code[]') or self.request.query_params.getlist('plant_code')
        if plant_codes:
            queryset = queryset.filter(plant__code__in=plant_codes)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        return queryset
    
    @action(detail=False, methods=['post'], url_path='import')
    @audit_action('DATA_IMPORT', 'Historical data import', category='data_processing', severity='MEDIUM')
    def import_historical(self, request):
        """Import historical data from Excel files"""
        serializer = HistoricalDataUploadSerializer(data=request.data)
        if not serializer.is_valid():
            AuditLogger.log_user_action(
                user=request.user,
                action='DATA_IMPORT',
                description='Historical data import failed: Invalid data',
                category='data_processing',
                severity='LOW',
                success=False,
                error_message=str(serializer.errors),
                request=request
            )
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        capacity_file = serializer.validated_data.get('capacity_file')
        historical_file = serializer.validated_data.get('historical_file')
        
        try:
            with AuditContext(request.user, 'HISTORICAL_DATA_IMPORT', 'Processing historical data files'):
                importer = HistoricalDataImporter()
                results = {}
                
                # Save files temporarily
                temp_files = []
                
                if capacity_file:
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        for chunk in capacity_file.chunks():
                            tmp.write(chunk)
                        capacity_path = tmp.name
                        temp_files.append(capacity_path)
                    
                    results['capacity'] = importer.import_plant_capacity(capacity_path)
                    
                    # Log capacity import
                    AuditLogger.log_user_action(
                        user=request.user,
                        action='DATA_IMPORT',
                        description=f'Plant capacity data imported: {results["capacity"].get("imported", 0)} records',
                        model_name='PlantCapacity',
                        category='data_processing',
                        severity='MEDIUM',
                        request=request
                    )
                
                if historical_file:
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        for chunk in historical_file.chunks():
                            tmp.write(chunk)
                        historical_path = tmp.name
                        temp_files.append(historical_path)
                    
                    results['historical'] = importer.import_historical_data(historical_path)
                    
                    # Log historical import
                    AuditLogger.log_user_action(
                        user=request.user,
                        action='DATA_IMPORT',
                        description=f'Historical data imported: {results["historical"].get("imported", 0)} records',
                        model_name='HistoricalData',
                        category='data_processing',
                        severity='MEDIUM',
                        request=request
                    )
                
                # Clean up temp files
                for temp_file in temp_files:
                    try:
                        os.unlink(temp_file)
                    except:
                        pass
                
                # Calculate totals
                total_imported = sum(r.get('imported', 0) for r in results.values())
                all_errors = []
                all_warnings = []
                
                for key, result in results.items():
                    all_errors.extend(result.get('errors', []))
                    all_warnings.extend(result.get('warnings', []))
                
                # Log overall import result
                AuditLogger.log_user_action(
                    user=request.user,
                    action='DATA_IMPORT',
                    description=f'Historical data import completed: {total_imported} total records imported',
                    category='data_processing',
                    severity='MEDIUM',
                    request=request
                )
                
                return Response({
                    'success': all(r.get('success', False) for r in results.values()),
                    'total_imported': total_imported,
                    'errors': all_errors,
                    'warnings': all_warnings,
                    'details': results
                }, status=status.HTTP_201_CREATED if total_imported > 0 else status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            AuditLogger.log_user_action(
                user=request.user,
                action='DATA_IMPORT',
                description=f'Historical data import failed: {str(e)}',
                category='data_processing',
                severity='HIGH',
                success=False,
                error_message=str(e),
                request=request
            )
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PlantCapacityViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for plant capacity records"""
    queryset = PlantCapacity.objects.all().select_related('plant')
    serializer_class = PlantCapacitySerializer
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by plant
        plant_codes = self.request.query_params.getlist('plant_code[]') or self.request.query_params.getlist('plant_code')
        if plant_codes:
            queryset = queryset.filter(plant__code__in=plant_codes)
        
        # Filter by date
        effective_date = self.request.query_params.get('effective_date')
        if effective_date:
            queryset = queryset.filter(effective_date=effective_date)
        
        return queryset


class WaterNominationViewSet(viewsets.ModelViewSet):
    """ViewSet for water nominations"""
    queryset = WaterNomination.objects.all().select_related('plant', 'submitted_by', 'approved_by')
    serializer_class = WaterNominationSerializer
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by plant
        plant_codes = self.request.query_params.getlist('plant_code[]') or self.request.query_params.getlist('plant_code')
        if plant_codes:
            queryset = queryset.filter(plant__code__in=plant_codes)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(nomination_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(nomination_date__lte=end_date)
        
        # Filter by status
        status = self.request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by nomination type
        nomination_type = self.request.query_params.get('nomination_type')
        if nomination_type:
            queryset = queryset.filter(nomination_type=nomination_type)
        
        return queryset
    
    def perform_create(self, serializer):
        nomination = serializer.save(submitted_by=self.request.user if self.request.user.is_authenticated else None)
        
        # Log water nomination creation
        AuditLogger.log_user_action(
            user=self.request.user,
            action='DATA_CREATE',
            description=f'Created water nomination for {nomination.plant.name} on {nomination.nomination_date}',
            model_name='WaterNomination',
            object_id=nomination.id,
            category='data_management',
            severity='MEDIUM',
            request=self.request
        )
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit a nomination for approval"""
        nomination = self.get_object()
        
        if nomination.status != 'DRAFT':
            return Response({'error': 'Only draft nominations can be submitted'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        nomination.status = 'SUBMITTED'
        nomination.submitted_at = datetime.now()
        nomination.submitted_by = request.user if request.user.is_authenticated else None
        nomination.save()
        
        # Log nomination submission
        AuditLogger.log_user_action(
            user=request.user,
            action='DATA_UPDATE',
            description=f'Submitted water nomination for {nomination.plant.name} on {nomination.nomination_date}',
            model_name='WaterNomination',
            object_id=nomination.id,
            category='workflow',
            severity='MEDIUM',
            request=request
        )
        
        return Response({'message': 'Nomination submitted successfully'}, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve a submitted nomination"""
        nomination = self.get_object()
        
        if nomination.status != 'SUBMITTED':
            return Response({'error': 'Only submitted nominations can be approved'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        nomination.status = 'APPROVED'
        nomination.approved_at = datetime.now()
        nomination.approved_by = request.user if request.user.is_authenticated else None
        nomination.save()
        
        # Log nomination approval
        AuditLogger.log_user_action(
            user=request.user,
            action='DATA_UPDATE',
            description=f'Approved water nomination for {nomination.plant.name} on {nomination.nomination_date}',
            model_name='WaterNomination',
            object_id=nomination.id,
            category='workflow',
            severity='MEDIUM',
            request=request
        )
        
        return Response({'message': 'Nomination approved successfully'}, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject a submitted nomination"""
        nomination = self.get_object()
        
        if nomination.status != 'SUBMITTED':
            return Response({'error': 'Only submitted nominations can be rejected'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        remarks = request.data.get('remarks', '')
        nomination.status = 'REJECTED'
        nomination.remarks = remarks
        nomination.save()
        
        # Log nomination rejection
        AuditLogger.log_user_action(
            user=request.user,
            action='DATA_UPDATE',
            description=f'Rejected water nomination for {nomination.plant.name} on {nomination.nomination_date}. Reason: {remarks}',
            model_name='WaterNomination',
            object_id=nomination.id,
            category='workflow',
            severity='MEDIUM',
            request=request
        )
        
        return Response({'message': 'Nomination rejected'}, status=status.HTTP_200_OK)


class ActualGenerationViewSet(viewsets.ModelViewSet):
    """ViewSet for actual generation data"""
    queryset = ActualGeneration.objects.all().select_related('plant')
    serializer_class = ActualGenerationSerializer
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by plant
        plant_codes = self.request.query_params.getlist('plant_code[]') or self.request.query_params.getlist('plant_code')
        if plant_codes:
            queryset = queryset.filter(plant__code__in=plant_codes)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(generation_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(generation_date__lte=end_date)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def variance_analysis(self, request):
        """Compare nominations with actual generation"""
        plant_code = request.query_params.get('plant_code')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if not all([plant_code, start_date, end_date]):
            return Response({'error': 'plant_code, start_date, and end_date are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            plant = Plant.objects.get(code=plant_code)
        except Plant.DoesNotExist:
            return Response({'error': 'Plant not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get nominations and actuals
        nominations = WaterNomination.objects.filter(
            plant=plant,
            nomination_date__gte=start_date,
            nomination_date__lte=end_date,
            status='APPROVED'
        )
        
        actuals = ActualGeneration.objects.filter(
            plant=plant,
            generation_date__gte=start_date,
            generation_date__lte=end_date
        )
        
        # Build comparison data
        results = []
        for nomination in nominations:
            try:
                actual = actuals.get(generation_date=nomination.nomination_date)
                
                # Calculate variance
                variance_mwh = float(actual.total_actual_mwh) - float(nomination.total_nominated_mwh)
                variance_percent = (variance_mwh / float(nomination.total_nominated_mwh) * 100) if nomination.total_nominated_mwh > 0 else 0
                
                # Hourly comparison
                hourly_comparison = []
                for i in range(24):
                    hour_field = f'hour_{str(i).zfill(2)}'
                    nominated = float(getattr(nomination, hour_field, 0) or 0)
                    actual_val = float(getattr(actual, hour_field, 0) or 0)
                    hourly_comparison.append({
                        'hour': i,
                        'time': f"{str(i).zfill(2)}:00-{str(i+1).zfill(2)}:00",
                        'nominated_mw': nominated,
                        'actual_mw': actual_val,
                        'variance_mw': actual_val - nominated,
                        'variance_percent': ((actual_val - nominated) / nominated * 100) if nominated > 0 else 0
                    })
                
                results.append({
                    'date': nomination.nomination_date,
                    'plant_code': plant.code,
                    'plant_name': plant.name,
                    'nomination_type': nomination.nomination_type,
                    'total_nominated_mwh': nomination.total_nominated_mwh,
                    'total_actual_mwh': actual.total_actual_mwh,
                    'variance_mwh': variance_mwh,
                    'variance_percent': round(variance_percent, 2),
                    'hourly_comparison': hourly_comparison
                })
            except ActualGeneration.DoesNotExist:
                # No actual data for this nomination
                pass
        
        return Response(results, status=status.HTTP_200_OK)


class TestimonialViewSet(viewsets.ModelViewSet):
    """
    API endpoint for testimonials
    - GET: Public access to view active testimonials
    - POST: Authenticated users can submit testimonials (pending approval)
    """
    serializer_class = TestimonialSerializer
    permission_classes = [AllowAny]  # Allow public read and authenticated write
    
    def get_queryset(self):
        # Only show active testimonials for list/retrieve
        if self.action in ['list', 'retrieve']:
            return Testimonial.objects.filter(is_active=True).order_by('order', '-created_at')
        # For admin actions, show all
        return Testimonial.objects.all()
    
    def perform_create(self, serializer):
        # New testimonials default to inactive (pending admin approval)
        testimonial = serializer.save(
            submitted_by=self.request.user if self.request.user.is_authenticated else None,
            is_active=False
        )
        
        # Log testimonial submission
        AuditLogger.log_user_action(
            user=self.request.user,
            action='DATA_CREATE',
            description=f'Submitted testimonial: "{testimonial.content[:100]}..." (pending approval)',
            model_name='Testimonial',
            object_id=testimonial.id,
            category='content_management',
            severity='LOW',
            request=self.request
        )


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for audit logs
    - GET: List all audit logs with filtering
    - Requires authentication
    """
    queryset = AuditLog.objects.all().select_related('user')
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by action type
        action = self.request.query_params.get('action')
        if action:
            queryset = queryset.filter(action__icontains=action)
        
        # Filter by username
        username = self.request.query_params.get('username')
        if username:
            queryset = queryset.filter(user__username__icontains=username)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        if start_date:
            queryset = queryset.filter(timestamp__gte=start_date)
        
        end_date = self.request.query_params.get('end_date')
        if end_date:
            queryset = queryset.filter(timestamp__lte=end_date)
        
        return queryset.order_by('-timestamp')
    
    @action(detail=False, methods=['get'], url_path='export')
    def export_logs(self, request):
        """Export audit logs to Excel"""
        queryset = self.get_queryset()
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        
        # Headers
        headers = ['Timestamp', 'User', 'User Role', 'Action', 'Model', 'Description', 'IP Address', 'Location']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Model mapping for export
        model_map = {
            'Plant': 'Power Plant',
            'Unit': 'Generation Unit',
            'UploadedFile': 'File Upload',
            'GenerationReport': 'Generation Report',
            'PlantCapacity': 'Plant Capacity',
            'HistoricalData': 'Historical Data',
            'WaterNomination': 'Water Nomination',
            'ActualGeneration': 'Actual Generation',
            'Testimonial': 'User Testimonial',
            'UserProfile': 'User Profile',
            'User': 'System User',
            'AuditLog': 'Audit Log',
            'PasswordResetRequest': 'Password Reset',
            'ESignature': 'Electronic Signature',
            'ReportSignature': 'Report Signature',
            'SignatoryAuthorization': 'Signatory Authorization',
            'SignatureVerificationToken': 'Security Token',
            'SignatureSecuritySettings': 'Security Settings',
            'Document': 'System Document',
            'AuthRequest': 'Authorization Request',
            'SignatoryAuthorizationRequest': 'Authorization Request',
        }
        
        # Data rows
        for log in queryset:
            # Get user full name
            user_display = 'System'
            if log.user:
                full_name = log.user.get_full_name()
                user_display = full_name if full_name else log.user.username
            
            # Get model display name
            model_display = log.model_name
            if log.model_name in model_map:
                model_display = model_map[log.model_name]
            elif log.model_name:
                import re
                model_display = re.sub(r'(?<!^)(?=[A-Z])', ' ', log.model_name)
            else:
                # Handle empty model_name
                if log.category == 'authentication':
                    model_display = "Authentication"
                elif log.category == 'security':
                    model_display = "Security"
                elif log.category == 'system':
                    model_display = "System"
                elif 'view' in log.description.lower() or 'access' in log.description.lower():
                    model_display = "Page Access"
                else:
                    model_display = "General System"
            
            # Get location display
            location_display = log.location
            if not log.location or log.location.lower() == 'unknown':
                if log.ip_address in ['127.0.0.1', '::1']:
                    location_display = "Local System (Internal)"
                else:
                    location_display = "Internal Network"
            
            # Get user role
            user_role = "System Role"
            if log.user:
                if hasattr(log.user, 'profile'):
                    role_map = {
                        'VIEWER': 'Viewer',
                        'OPERATOR': 'Data Encoder / Operator',
                        'MANAGER': 'Data Manager',
                        'ADMIN': 'Administrator'
                    }
                    user_role = role_map.get(log.user.profile.role, log.user.profile.get_role_display())
                elif log.user.is_superuser:
                    user_role = "Administrator"

            ws.append([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                user_display,
                user_role,
                log.get_action_display(),
                model_display,
                log.description,
                log.ip_address or 'N/A',
                location_display
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from django.http import HttpResponse
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Audit_Logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response

class ESignatureViewSet(viewsets.ModelViewSet):
    """ViewSet for managing e-signatures with enhanced security"""
    queryset = ESignature.objects.filter(is_active=True)
    serializer_class = ESignatureSerializer
    permission_classes = [IsAuthenticated]  # SECURITY: Require authentication
    
    def get_queryset(self):
        """Filter signatures based on user permissions"""
        queryset = super().get_queryset()
        
        # Superusers see all signatures
        if self.request.user.is_superuser:
            queryset_filtered = queryset
        else:
            # Regular users see only their own signatures
            queryset_filtered = queryset.filter(created_by=self.request.user)
        
        # Apply additional filters
        signatory_name = self.request.query_params.get('signatory_name')
        if signatory_name:
            queryset_filtered = queryset_filtered.filter(signatory_name__icontains=signatory_name)
        
        return queryset_filtered.order_by('-created_at')
    
    @action(detail=False, methods=['post'], url_path='create-from-data')
    @audit_action('SIGNATURE_CREATE', 'E-signature creation from data', category='e_signature', severity='MEDIUM')
    def create_from_data(self, request):
        """Create e-signature from base64 data"""
        import base64
        from django.core.files.base import ContentFile
        
        try:
            signatory_name = request.data.get('signatory_name')
            signatory_title = request.data.get('signatory_title', '')
            signatory_role = request.data.get('signatory_role', '')
            signature_type = request.data.get('signature_type', 'DRAW')
            signature_data = request.data.get('signature_data')
            is_default = request.data.get('is_default', True)
            
            if not signatory_name or not signature_data:
                AuditLogger.log_user_action(
                    user=request.user,
                    action='SIGNATURE_CREATE',
                    description='E-signature creation failed: Missing required data',
                    category='e_signature',
                    severity='LOW',
                    success=False,
                    error_message='signatory_name and signature_data are required',
                    request=request
                )
                return Response(
                    {'error': 'signatory_name and signature_data are required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Remove data URL prefix if present
            if signature_data.startswith('data:image'):
                signature_data = signature_data.split(',')[1]
            
            # Decode base64 data
            try:
                image_data = base64.b64decode(signature_data)
            except Exception as e:
                AuditLogger.log_user_action(
                    user=request.user,
                    action='SIGNATURE_CREATE',
                    description=f'E-signature creation failed for {signatory_name}: Invalid base64 data',
                    category='e_signature',
                    severity='LOW',
                    success=False,
                    error_message=str(e),
                    request=request
                )
                return Response(
                    {'error': f'Invalid base64 data: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create filename
            filename = f"{signatory_name.replace(' ', '_').replace('.', '').lower()}_signature.png"
            
            # Create signature instance
            signature = ESignature.objects.create(
                signatory_name=signatory_name,
                signatory_title=signatory_title,
                signatory_role=signatory_role,
                signature_type=signature_type,
                signature_image=ContentFile(image_data, filename),
                signature_data=signature_data,
                is_default=is_default,
                created_by=request.user
            )
            
            # Log successful signature creation
            audit_signature_creation(
                user=request.user,
                signatory_name=signatory_name,
                request=request
            )
            
            AuditLogger.log_user_action(
                user=request.user,
                action='SIGNATURE_CREATE',
                description=f'Created e-signature for {signatory_name} (type: {signature_type})',
                model_name='ESignature',
                object_id=signature.id,
                category='e_signature',
                severity='MEDIUM',
                request=request
            )
            
            serializer = self.get_serializer(signature)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            AuditLogger.log_user_action(
                user=request.user,
                action='SIGNATURE_CREATE',
                description=f'E-signature creation failed for {signatory_name}: {str(e)}',
                category='e_signature',
                severity='HIGH',
                success=False,
                error_message=str(e),
                request=request
            )
            return Response(
                {'error': f'Failed to create signature: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='by-signatory')
    def by_signatory(self, request):
        """Get signatures for a specific signatory"""
        signatory_name = request.query_params.get('name')
        
        if not signatory_name:
            return Response(
                {'error': 'name parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        signatures = self.get_queryset().filter(signatory_name=signatory_name)
        serializer = self.get_serializer(signatures, many=True)
        return Response(serializer.data)


class ReportSignatureViewSet(viewsets.ModelViewSet):
    """ViewSet for managing report signatures with enhanced security"""
    queryset = ReportSignature.objects.all()
    serializer_class = ReportSignatureSerializer
    permission_classes = [IsAuthenticated]  # SECURITY: Require authentication
    
    def get_queryset(self):
        queryset = super().get_queryset()
        report_date = self.request.query_params.get('report_date')
        report_type = self.request.query_params.get('report_type', 'PSR')
        
        if report_date:
            queryset = queryset.filter(report_date=report_date)
        if report_type:
            queryset = queryset.filter(report_type=report_type)
            
        return queryset.order_by('-signed_at')
    
    @action(detail=False, methods=['post'], url_path='sign-report')
    @audit_action('REPORT_SIGN', 'Report signing with e-signature', category='e_signature', severity='HIGH')
    def sign_report(self, request):
        """Sign a report with an e-signature"""
        data = request.data.copy()
        signatory_name = data.get('signatory_name')
        report_date = data.get('report_date')
        report_type = data.get('report_type', 'PSR')
        
        # Generate verification hash
        import hashlib
        import json
        hash_data = {
            'report_date': report_date,
            'report_type': report_type,
            'signatory_name': signatory_name,
            'signatory_role': data.get('signatory_role'),
            'timestamp': datetime.now().isoformat()
        }
        verification_hash = hashlib.sha256(json.dumps(hash_data, sort_keys=True).encode()).hexdigest()
        data['verification_hash'] = verification_hash
        
        serializer = self.get_serializer(data=data, context={'request': request})
        if serializer.is_valid():
            signature = serializer.save()
            
            # Log report signing
            AuditLogger.log_user_action(
                user=request.user,
                action='REPORT_SIGN',
                description=f'Signed {report_type} report for {report_date} as {signatory_name}',
                model_name='ReportSignature',
                object_id=signature.id,
                category='e_signature',
                severity='HIGH',
                request=request
            )
            
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            # Log signing failure
            AuditLogger.log_user_action(
                user=request.user,
                action='REPORT_SIGN',
                description=f'Failed to sign {report_type} report for {report_date} as {signatory_name}: {str(serializer.errors)}',
                category='e_signature',
                severity='MEDIUM',
                success=False,
                error_message=str(serializer.errors),
                request=request
            )
            
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'], url_path='for-report')
    def for_report(self, request):
        """Get all signatures for a specific report"""
        report_date = request.query_params.get('report_date')
        report_type = request.query_params.get('report_type', 'PSR')
        
        if not report_date:
            return Response(
                {'error': 'report_date parameter is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        signatures = self.get_queryset().filter(
            report_date=report_date,
            report_type=report_type
        )
        serializer = self.get_serializer(signatures, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'], url_path='request-2fa')
    def request_2fa(self, request):
        """Request 2FA code for signature operation"""
        from .signature_utils.signature_2fa import Signature2FA
        from .models import SignatureVerificationToken, SignatureSecuritySettings
        from datetime import timedelta
        
        signatory_name = request.data.get('signatory_name')
        signature_intent = request.data.get('signature_intent', {})
        
        if not signatory_name:
            return Response(
                {'error': 'signatory_name is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if user is authorized for this signatory
        from .models import SignatoryAuthorization
        try:
            auth = SignatoryAuthorization.objects.get(
                user=request.user,
                signatory_name=signatory_name,
                is_active=True
            )
            if not auth.is_valid():
                return Response(
                    {'error': 'Your authorization for this signatory has expired'},
                    status=status.HTTP_403_FORBIDDEN
                )
        except SignatoryAuthorization.DoesNotExist:
            if not request.user.is_superuser:
                return Response(
                    {'error': 'You are not authorized to sign as this signatory'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        # Generate OTP
        otp_code, secret = Signature2FA.generate_simple_otp()
        
        # Get settings
        settings = SignatureSecuritySettings.get_settings()
        
        # Create verification token
        token = SignatureVerificationToken.objects.create(
            user=request.user,
            token=otp_code,
            secret=secret,
            signature_intent=signature_intent,
            expires_at=timezone.now() + timedelta(minutes=settings.otp_validity_minutes),
            ip_address=self._get_client_ip(request)
        )
        
        # Send OTP via email
        email_sent = Signature2FA.send_otp_email(request.user, otp_code, signatory_name)
        
        # Log 2FA request
        from .models import SignatureAuditLog
        SignatureAuditLog.objects.create(
            user=request.user,
            action='2FA_REQUEST',
            ip_address=self._get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            success=True,
            additional_data={'signatory_name': signatory_name}
        )
        
        return Response({
            'token_id': token.id,
            'message': 'Verification code sent to your email' if email_sent else 'Verification code generated',
            'expires_at': token.expires_at,
            'otp_code': otp_code if not email_sent else None  # Only return if email failed
        })
    
    @action(detail=False, methods=['post'], url_path='verify-2fa')
    def verify_2fa(self, request):
        """Verify 2FA code and return authorization"""
        from .signature_utils.signature_2fa import Signature2FA
        from .models import SignatureVerificationToken, SignatureAuditLog
        
        token_id = request.data.get('token_id')
        otp_code = request.data.get('otp_code')
        
        if not token_id or not otp_code:
            return Response(
                {'error': 'token_id and otp_code are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            token = SignatureVerificationToken.objects.get(id=token_id, user=request.user)
        except SignatureVerificationToken.DoesNotExist:
            return Response(
                {'error': 'Invalid token'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if token is valid
        if not token.is_valid():
            SignatureAuditLog.objects.create(
                user=request.user,
                action='2FA_FAILURE',
                ip_address=self._get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                success=False,
                failure_reason='Token expired or max attempts reached'
            )
            return Response(
                {'error': 'Token has expired or maximum attempts reached'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verify OTP
        is_valid = Signature2FA.verify_simple_otp(
            token.token,
            otp_code,
            token.created_at
        )
        
        token.increment_attempts()
        
        if is_valid:
            token.is_used = True
            token.verified_at = timezone.now()
            token.save()
            
            # Log successful 2FA
            SignatureAuditLog.objects.create(
                user=request.user,
                action='2FA_SUCCESS',
                ip_address=self._get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                success=True
            )
            
            return Response({
                'verified': True,
                'message': 'Verification successful',
                'signature_intent': token.signature_intent
            })
        else:
            # Log failed 2FA
            SignatureAuditLog.objects.create(
                user=request.user,
                action='2FA_FAILURE',
                ip_address=self._get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                success=False,
                failure_reason='Invalid OTP code'
            )
            
            return Response(
                {'error': 'Invalid verification code', 'attempts_remaining': token.max_attempts - token.attempts},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post'], url_path='sign-with-2fa')
    def sign_with_2fa(self, request):
        """Sign report with 2FA verification"""
        from .permissions import CanSignAsSignatory
        
        # Verify 2FA first
        token_id = request.data.get('token_id')
        otp_code = request.data.get('otp_code')
        
        if not token_id or not otp_code:
            return Response(
                {'error': 'token_id and otp_code are required for 2FA'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verify the 2FA code
        verify_response = self.verify_2fa(request)
        if verify_response.status_code != 200:
            return verify_response
        
        # If 2FA verified, proceed with signing
        return self.sign_report(request)
    
    def _get_client_ip(self, request):
        """Get client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            return x_forwarded_for.split(',')[0].strip()
        return request.META.get('REMOTE_ADDR', '127.0.0.1')


class MonthlyTargetViewSet(viewsets.ModelViewSet):
    queryset = MonthlyTarget.objects.all().select_related('plant')
    serializer_class = MonthlyTargetSerializer
    permission_classes = [AllowAny]
    authentication_classes = []  # Disable authentication to avoid CSRF issues
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        queryset = super().get_queryset()
        plant_code = self.request.query_params.get('plant_code')
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')

        if plant_code:
            queryset = queryset.filter(plant__code=plant_code)
        if year:
            queryset = queryset.filter(year=year)
        if month:
            queryset = queryset.filter(month=month)
            
        return queryset

    @action(detail=False, methods=['get'], url_path='current')
    def get_current_target(self, request):
        """Get the current target for a specific plant, month, and year"""
        plant_code = request.query_params.get('plant_code')
        month = request.query_params.get('month')
        year = request.query_params.get('year')

        if not plant_code or not month or not year:
            return Response(
                {'error': 'plant_code, month, and year are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            target = MonthlyTarget.objects.get(
                plant__code=plant_code,
                month=month,
                year=year
            )
            return Response(self.get_serializer(target).data)
        except MonthlyTarget.DoesNotExist:
            return Response(
                {'error': 'Target not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=False, methods=['post'], url_path='set-current')
    def set_current_target(self, request):
        """Set or update the target for a plant, month, and year"""
        plant_code = request.data.get('plant_code')
        month = request.data.get('month')
        year = request.data.get('year')
        target_percentage = request.data.get('target_percentage')

        if not plant_code or not month or not year or target_percentage is None:
            return Response(
                {'error': 'plant_code, month, year, and target_percentage are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            plant = Plant.objects.get(code=plant_code)
            
            target, created = MonthlyTarget.objects.update_or_create(
                plant=plant,
                month=month,
                year=year,
                defaults={'target_percentage': target_percentage}
            )
            
            # Log target set/update
            AuditLogger.log_user_action(
                user=request.user if request.user.is_authenticated else None,
                action='DATA_UPDATE' if not created else 'DATA_CREATE',
                description=f'{"Updated" if not created else "Set"} monthly target for {plant.name} ({year}-{str(month).zfill(2)}): {target_percentage}%',
                model_name='MonthlyTarget',
                object_id=target.id,
                category='target_management',
                severity='MEDIUM',
                request=request
            )
            
            return Response({'success': True, 'target': self.get_serializer(target).data})
        except Plant.DoesNotExist:
            return Response({'error': f'Plant {plant_code} not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='bulk-set')
    def bulk_set_targets(self, request):
        """Bulk set or update monthly targets for multiple plants"""
        targets_data = request.data.get('targets')
        if not targets_data or not isinstance(targets_data, list):
            return Response({'error': 'A list of targets is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        from datetime import datetime
        now = datetime.now()
        
        results = []
        errors = []
        
        from django.db import transaction
        
        try:
            with transaction.atomic():
                for target_item in targets_data:
                    plant_code = target_item.get('plant_code')
                    target_percentage = target_item.get('target_percentage')
                    month = int(target_item.get('month', now.month))
                    year = int(target_item.get('year', now.year))
                    
                    if not plant_code or target_percentage is None:
                        errors.append({'plant_code': plant_code, 'error': 'Missing required fields'})
                        continue
                    
                    try:
                        plant = Plant.objects.get(code=plant_code)
                        target, created = MonthlyTarget.objects.update_or_create(
                            plant=plant,
                            month=month,
                            year=year,
                            defaults={
                                'target_percentage': target_percentage,
                                'created_by': request.user if request.user.is_authenticated else None
                            }
                        )
                        results.append(self.get_serializer(target).data)
                    except Plant.DoesNotExist:
                        errors.append({'plant_code': plant_code, 'error': 'Plant not found'})
            
            # Log bulk action
            AuditLogger.log_user_action(
                user=request.user if request.user.is_authenticated else None,
                action='DATA_UPDATE',
                description=f'Bulk updated monthly targets for {len(results)} plants',
                model_name='MonthlyTarget',
                category='target_management',
                severity='MEDIUM',
                request=request
            )
            
            return Response({
                'success': True, 
                'updated_count': len(results),
                'results': results,
                'errors': errors
            })
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)