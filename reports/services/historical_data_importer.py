"""
Historical Data Importer Service
Handles importing plant capacity and historical data from Excel files
Supports PSR REPORT format and other NPC data formats
"""

import openpyxl
from datetime import datetime, date
from decimal import Decimal
from django.db import transaction
from reports.models import Plant, PlantCapacity, HistoricalData


class HistoricalDataImporter:
    """Service for importing historical plant data from Excel files"""
    
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.imported_count = 0
    
    def import_plant_capacity(self, file_path):
        """
        Import plant capacity data from 0PLANT DEPCAP.xlsx format
        
        Expected columns:
        - Plant Code
        - Plant Name
        - Dependable Capacity (MW)
        - Installed Capacity (MW)
        """
        self.errors = []
        self.warnings = []
        self.imported_count = 0
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Find header row
            header_row = None
            for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10), start=1):
                cell_values = [str(cell.value).strip().upper() if cell.value else '' for cell in row]
                if any('PLANT' in val and 'CODE' in val for val in cell_values):
                    header_row = idx
                    break
            
            if not header_row:
                return {
                    'success': False,
                    'error': 'Could not find header row with plant data',
                    'imported': 0
                }
            
            # Parse headers
            headers = {}
            for idx, cell in enumerate(sheet[header_row]):
                if cell.value:
                    header_text = str(cell.value).strip().upper()
                    if 'PLANT' in header_text and 'CODE' in header_text:
                        headers['code'] = idx
                    elif 'PLANT' in header_text and 'NAME' in header_text:
                        headers['name'] = idx
                    elif 'DEPENDABLE' in header_text or 'DEP' in header_text:
                        headers['dependable'] = idx
                    elif 'INSTALLED' in header_text or 'INST' in header_text:
                        headers['installed'] = idx
            
            # Import data
            with transaction.atomic():
                for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
                    try:
                        # Get plant code
                        code_cell = row[headers.get('code', 0)]
                        if not code_cell.value:
                            continue
                        
                        plant_code = str(code_cell.value).strip().upper()
                        if not plant_code or plant_code in ['TOTAL', 'SUBTOTAL']:
                            continue
                        
                        # Get or create plant
                        plant, created = Plant.objects.get_or_create(
                            code=plant_code,
                            defaults={'name': plant_code}
                        )
                        
                        # Update plant name if provided
                        if 'name' in headers:
                            name_cell = row[headers['name']]
                            if name_cell.value:
                                plant.name = str(name_cell.value).strip()
                                plant.save()
                        
                        # Get capacity values
                        dependable = self._parse_decimal(row[headers.get('dependable', 1)].value)
                        installed = self._parse_decimal(row[headers.get('installed', 2)].value)
                        
                        if dependable or installed:
                            PlantCapacity.objects.update_or_create(
                                plant=plant,
                                defaults={
                                    'dependable_capacity': dependable or Decimal('0'),
                                    'installed_capacity': installed or Decimal('0')
                                }
                            )
                            self.imported_count += 1
                    
                    except Exception as e:
                        self.errors.append(f"Row {row_idx}: {str(e)}")
            
            return {
                'success': True,
                'imported': self.imported_count,
                'errors': self.errors,
                'warnings': self.warnings
            }
        
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'imported': 0
            }
    
    def import_historical_data(self, file_path):
        """
        Import historical data from 1DATA APAO.xlsx or PSR REPORT format
        
        Supports multiple formats:
        - Daily data with columns for each day
        - PSR Report format with plant status data
        - Transposed PSR format with dates in rows
        """
        self.errors = []
        self.warnings = []
        self.imported_count = 0
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Try to find a sheet with importable data
            best_sheet = None
            best_format = None
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                format_type = self._detect_format(sheet)
                
                # Prefer TRANSPOSED_PSR or PSR formats over DAILY
                if format_type == 'TRANSPOSED_PSR':
                    best_sheet = sheet
                    best_format = format_type
                    break  # Found the best format
                elif format_type == 'PSR' and best_format != 'TRANSPOSED_PSR':
                    best_sheet = sheet
                    best_format = format_type
                elif format_type == 'DAILY' and not best_sheet:
                    best_sheet = sheet
                    best_format = format_type
            
            if not best_sheet:
                return {
                    'success': False,
                    'error': 'No importable data found in any sheet',
                    'imported': 0
                }
            
            # Import using detected format
            if best_format == 'TRANSPOSED_PSR':
                return self._import_transposed_psr_format(best_sheet)
            elif best_format == 'PSR':
                return self._import_psr_format(best_sheet)
            elif best_format == 'DAILY':
                return self._import_daily_format(best_sheet)
            else:
                return {
                    'success': False,
                    'error': 'Unknown file format. Expected PSR Report or Daily Data format.',
                    'imported': 0
                }
        
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'imported': 0,
                'errors': self.errors
            }
    
    def _detect_format(self, sheet):
        """Detect the format of the Excel file"""
        # Check for transposed format (dates in first column, plants as headers)
        # Look at first few rows
        first_col_dates = 0
        for row_idx in range(2, min(10, sheet.max_row + 1)):
            cell = sheet.cell(row_idx, 1)
            if isinstance(cell.value, (datetime, date)):
                first_col_dates += 1
        
        if first_col_dates >= 3:
            # Check if row 2 has plant names
            row2 = list(sheet[2])
            plant_names = 0
            for cell in row2[1:8]:  # Check first few columns after date column
                if cell.value and isinstance(cell.value, str):
                    text = str(cell.value).strip().upper()
                    if any(keyword in text for keyword in ['AGUS', 'PULANGI', 'PLANT', 'UNIT']):
                        plant_names += 1
            
            if plant_names >= 2:
                return 'TRANSPOSED_PSR'
        
        # Check first 20 rows for format indicators
        for row in sheet.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value:
                    text = str(cell.value).upper()
                    if 'PSR' in text or 'PLANT STATUS REPORT' in text:
                        return 'PSR'
                    if 'DAILY' in text or any(str(i) in text for i in range(1, 32)):
                        return 'DAILY'
        
        # Default to daily format
        return 'DAILY'
    
    def _import_psr_format(self, sheet):
        """
        Import PSR (Plant Status Report) format
        
        Expected structure:
        - Plant codes in first column (starting around row 10-15)
        - Multiple data columns with various metrics
        - May have merged cells and complex headers
        """
        # Find the data start row by looking for plant codes
        data_start_row = None
        plant_code_col = 0
        
        # Scan first 30 rows to find where plant data starts
        for idx in range(1, 31):
            try:
                row = list(sheet[idx])
                first_cell = row[0].value
                
                if first_cell:
                    text = str(first_cell).strip().upper()
                    # Look for plant code pattern (letters + numbers, short length)
                    if (len(text) <= 10 and 
                        any(c.isalpha() for c in text) and 
                        any(c.isdigit() for c in text) and
                        text not in ['PLANT', 'CODE', 'TOTAL', 'SUBTOTAL']):
                        data_start_row = idx
                        break
            except:
                continue
        
        if not data_start_row:
            return {
                'success': False,
                'error': 'Could not find plant data rows',
                'imported': 0
            }
        
        # Try to extract date from filename or sheet
        report_date = self._extract_report_date(sheet)
        if not report_date:
            # Use today's date as fallback
            report_date = date.today()
            self.warnings.append(f"Could not determine report date, using {report_date}")
        
        # Find data columns - look for numeric data
        # Typically columns B onwards contain generation/capacity data
        data_columns = []
        sample_row = list(sheet[data_start_row])
        
        for col_idx in range(1, min(len(sample_row), 20)):  # Check first 20 columns
            cell_value = sample_row[col_idx].value
            if cell_value is not None and isinstance(cell_value, (int, float)):
                data_columns.append(col_idx)
        
        if not data_columns:
            self.warnings.append("No numeric data columns found, will import plant codes only")
        
        # Import data
        with transaction.atomic():
            current_row = data_start_row
            
            while current_row <= sheet.max_row:
                try:
                    row = list(sheet[current_row])
                    
                    # Get plant code from first column
                    plant_code_cell = row[0]
                    if not plant_code_cell.value:
                        current_row += 1
                        continue
                    
                    plant_code = str(plant_code_cell.value).strip().upper()
                    
                    # Stop if we hit totals or empty rows
                    if not plant_code or plant_code in ['TOTAL', 'SUBTOTAL', 'GRAND TOTAL', 'SUMMARY']:
                        break
                    
                    # Skip if not a valid plant code pattern
                    if len(plant_code) > 15 or not any(c.isalpha() for c in plant_code):
                        current_row += 1
                        continue
                    
                    # Get or create plant
                    plant, created = Plant.objects.get_or_create(
                        code=plant_code,
                        defaults={'name': plant_code}
                    )
                    
                    # Import generation data from first numeric column
                    if data_columns:
                        first_data_col = data_columns[0]
                        value = row[first_data_col].value
                        
                        if value is not None:
                            generation = self._parse_decimal(value)
                            
                            if generation is not None and generation > 0:
                                HistoricalData.objects.update_or_create(
                                    plant=plant,
                                    date=report_date,
                                    defaults={
                                        'generation_mwh': generation,
                                        'remarks': 'Imported from PSR Report'
                                    }
                                )
                                self.imported_count += 1
                    else:
                        # Just ensure plant exists
                        self.imported_count += 1
                
                except Exception as e:
                    self.errors.append(f"Row {current_row}: {str(e)}")
                
                current_row += 1
        
        return {
            'success': True,
            'imported': self.imported_count,
            'errors': self.errors,
            'warnings': self.warnings,
            'report_date': str(report_date)
        }
    
    def _extract_report_date(self, sheet):
        """Try to extract report date from sheet headers"""
        # Check first 15 rows for date information
        for row_idx in range(1, 16):
            try:
                row = list(sheet[row_idx])
                for cell in row[:10]:  # Check first 10 columns
                    if cell.value:
                        # Check if it's a date object
                        if isinstance(cell.value, (datetime, date)):
                            return cell.value if isinstance(cell.value, date) else cell.value.date()
                        
                        # Try to parse date from string
                        text = str(cell.value).strip()
                        parsed_date = self._parse_date_from_text(text)
                        if parsed_date:
                            return parsed_date
            except:
                continue
        
        return None
    
    def _parse_date_from_text(self, text):
        """Parse date from various text formats"""
        import re
        
        # Try to find date patterns
        # Format: MM/DD/YYYY or DD/MM/YYYY
        date_pattern = r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})'
        match = re.search(date_pattern, text)
        if match:
            try:
                # Try MM/DD/YYYY first
                month, day, year = map(int, match.groups())
                if 1 <= month <= 12 and 1 <= day <= 31:
                    return date(year, month, day)
            except:
                pass
            
            try:
                # Try DD/MM/YYYY
                day, month, year = map(int, match.groups())
                if 1 <= month <= 12 and 1 <= day <= 31:
                    return date(year, month, day)
            except:
                pass
        
        # Look for month names
        months = {
            'JANUARY': 1, 'JAN': 1, 'FEBRUARY': 2, 'FEB': 2,
            'MARCH': 3, 'MAR': 3, 'APRIL': 4, 'APR': 4,
            'MAY': 5, 'JUNE': 6, 'JUN': 6, 'JULY': 7, 'JUL': 7,
            'AUGUST': 8, 'AUG': 8, 'SEPTEMBER': 9, 'SEP': 9, 'SEPT': 9,
            'OCTOBER': 10, 'OCT': 10, 'NOVEMBER': 11, 'NOV': 11,
            'DECEMBER': 12, 'DEC': 12
        }
        
        text_upper = text.upper()
        for month_name, month_num in months.items():
            if month_name in text_upper:
                # Find year
                year_match = re.search(r'\b(19|20)\d{2}\b', text)
                if year_match:
                    year = int(year_match.group())
                    # Find day if present
                    day_match = re.search(r'\b(\d{1,2})\b', text)
                    day = int(day_match.group()) if day_match else 1
                    
                    try:
                        return date(year, month_num, day)
                    except:
                        pass
        
        return None
        
        return {
            'success': True,
            'imported': self.imported_count,
            'errors': self.errors,
            'warnings': self.warnings
        }
    
    def _import_transposed_psr_format(self, sheet):
        """
        Import Transposed PSR format (dates in rows, plants in columns)
        
        Expected structure:
        - Row 1: Header with date
        - Row 2: Plant names as column headers (Agus 1, Agus 2, etc.)
        - Row 3+: Dates in column 1, generation values in subsequent columns
        
        Example:
        |  Date      | Agus 1 | Agus 2 | Agus 4 | ...
        |------------|--------|--------|--------|----
        | 2021-01-01 |   20   |   71   |   83   | ...
        | 2021-01-02 |   20   |   71   |   83   | ...
        """
        self.errors = []
        self.warnings = []
        self.imported_count = 0
        
        # Find header row with plant names (usually row 2)
        header_row = None
        plant_columns = {}
        
        for row_idx in range(1, 5):
            row = list(sheet[row_idx])
            plant_count = 0
            temp_plant_cols = {}
            
            for col_idx, cell in enumerate(row[1:], start=1):  # Skip first column (dates)
                if cell.value:
                    text = str(cell.value).strip()
                    # Check if it looks like a plant name
                    if any(keyword in text.upper() for keyword in ['AGUS', 'PULANGI', 'PLANT']):
                        # Extract plant code (e.g., "Agus 1" -> "AGUS1")
                        plant_code = text.upper().replace(' ', '').replace('-', '')
                        temp_plant_cols[col_idx] = plant_code
                        plant_count += 1
            
            if plant_count >= 2:
                header_row = row_idx
                plant_columns = temp_plant_cols
                break
        
        if not header_row or not plant_columns:
            return {
                'success': False,
                'error': 'Could not find header row with plant names',
                'imported': 0
            }
        
        # Import data starting from row after header
        with transaction.atomic():
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                try:
                    row = list(sheet[row_idx])
                    
                    # Get date from first column
                    date_cell = row[0]
                    if not date_cell.value:
                        continue
                    
                    # Parse date
                    if isinstance(date_cell.value, datetime):
                        report_date = date_cell.value.date()
                    elif isinstance(date_cell.value, date):
                        report_date = date_cell.value
                    else:
                        # Try to parse as string
                        continue
                    
                    # Import data for each plant column
                    for col_idx, plant_code in plant_columns.items():
                        try:
                            if col_idx < len(row):
                                value = row[col_idx].value
                                
                                if value is not None:
                                    generation = self._parse_decimal(value)
                                    
                                    if generation is not None and generation >= 0:
                                        # Get or create plant
                                        plant, created = Plant.objects.get_or_create(
                                            code=plant_code,
                                            defaults={'name': plant_code}
                                        )
                                        
                                        # Create or update historical data
                                        HistoricalData.objects.update_or_create(
                                            plant=plant,
                                            date=report_date,
                                            defaults={
                                                'generation_mwh': generation,
                                                'remarks': 'Imported from PSR Report (Transposed format)'
                                            }
                                        )
                                        self.imported_count += 1
                        
                        except Exception as e:
                            self.warnings.append(
                                f"Row {row_idx}, Col {col_idx} ({plant_code}): {str(e)}"
                            )
                
                except Exception as e:
                    self.errors.append(f"Row {row_idx}: {str(e)}")
        
        return {
            'success': True,
            'imported': self.imported_count,
            'errors': self.errors,
            'warnings': self.warnings
        }
    
    def _import_daily_format(self, sheet):
        """
        Import daily format (1DATA APAO.xlsx style)
        
        Expected structure:
        - Plant codes in first column
        - Day numbers (1-31) as column headers
        - Month/Year information in header
        """
        # Find header row and extract month/year
        header_row = None
        month = None
        year = None
        day_columns = {}
        
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15), start=1):
            for cell in row:
                if cell.value:
                    text = str(cell.value).upper()
                    # Look for month/year
                    if any(m in text for m in ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
                                                'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']):
                        month_year = self._parse_month_year(text)
                        if month_year:
                            month, year = month_year
            
            # Check if this is the header row with day numbers
            cell_values = [cell.value for cell in row]
            if any(isinstance(val, (int, float)) and 1 <= val <= 31 for val in cell_values):
                header_row = idx
                
                # Map day numbers to columns
                for col_idx, cell in enumerate(row):
                    if isinstance(cell.value, (int, float)):
                        day = int(cell.value)
                        if 1 <= day <= 31:
                            day_columns[col_idx] = day
                
                break
        
        if not header_row or not day_columns or not month or not year:
            return {
                'success': False,
                'error': 'Could not parse daily format. Missing header, month, or year information.',
                'imported': 0
            }
        
        # Import data
        with transaction.atomic():
            for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
                try:
                    # Get plant code
                    plant_code_cell = row[0]
                    if not plant_code_cell.value:
                        continue
                    
                    plant_code = str(plant_code_cell.value).strip().upper()
                    
                    # Skip totals
                    if not plant_code or plant_code in ['TOTAL', 'SUBTOTAL']:
                        continue
                    
                    # Get or create plant
                    plant, created = Plant.objects.get_or_create(
                        code=plant_code,
                        defaults={'name': plant_code}
                    )
                    
                    # Import data for each day
                    for col_idx, day in day_columns.items():
                        try:
                            # Create date
                            report_date = date(year, month, day)
                            
                            value = row[col_idx].value
                            if value is not None:
                                generation = self._parse_decimal(value)
                                
                                if generation is not None:
                                    HistoricalData.objects.update_or_create(
                                        plant=plant,
                                        date=report_date,
                                        defaults={
                                            'generation_mwh': generation,
                                            'remarks': f'Imported from daily data'
                                        }
                                    )
                                    self.imported_count += 1
                        
                        except ValueError:
                            # Invalid date (e.g., Feb 30)
                            continue
                        except Exception as e:
                            self.warnings.append(
                                f"Row {row_idx}, Day {day}: {str(e)}"
                            )
                
                except Exception as e:
                    self.errors.append(f"Row {row_idx}: {str(e)}")
        
        return {
            'success': True,
            'imported': self.imported_count,
            'errors': self.errors,
            'warnings': self.warnings
        }
    
    def _parse_decimal(self, value):
        """Parse a value as Decimal, handling various formats"""
        if value is None or value == '':
            return None
        
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            
            # Clean string value
            str_value = str(value).strip().replace(',', '')
            if not str_value or str_value == '-':
                return None
            
            return Decimal(str_value)
        
        except:
            return None
    
    def _parse_date_header(self, value):
        """Parse date from header cell"""
        if isinstance(value, datetime):
            return value.date()
        
        if isinstance(value, date):
            return value
        
        if isinstance(value, (int, float)):
            # Excel date serial number
            try:
                excel_date = datetime(1899, 12, 30) + timedelta(days=int(value))
                return excel_date.date()
            except:
                pass
        
        # Try parsing string
        if isinstance(value, str):
            value = value.strip()
            
            # Try common date formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%d-%m-%Y', '%m-%d-%Y']:
                try:
                    return datetime.strptime(value, fmt).date()
                except:
                    continue
        
        return None
    
    def _parse_month_year(self, text):
        """Extract month and year from text"""
        months = {
            'JANUARY': 1, 'FEBRUARY': 2, 'MARCH': 3, 'APRIL': 4,
            'MAY': 5, 'JUNE': 6, 'JULY': 7, 'AUGUST': 8,
            'SEPTEMBER': 9, 'OCTOBER': 10, 'NOVEMBER': 11, 'DECEMBER': 12
        }
        
        month = None
        year = None
        
        # Find month
        for month_name, month_num in months.items():
            if month_name in text:
                month = month_num
                break
        
        # Find year (4-digit number)
        import re
        year_match = re.search(r'\b(19|20)\d{2}\b', text)
        if year_match:
            year = int(year_match.group())
        
        if month and year:
            return (month, year)
        
        return None


# Add missing import
from datetime import timedelta
