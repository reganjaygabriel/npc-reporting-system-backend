"""
Excel Template Generator Service
Generates downloadable Excel templates for data import
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from django.http import HttpResponse
import io


class TemplateGenerator:
    """Generate Excel templates for data import"""
    
    # Color scheme
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    SAMPLE_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    INSTRUCTION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    INSTRUCTION_FONT = Font(italic=True, size=10)
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def _style_header_row(ws, row_num, columns):
        """Apply styling to header row"""
        for col_num, _ in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = TemplateGenerator.HEADER_FILL
            cell.font = TemplateGenerator.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = TemplateGenerator.THIN_BORDER
    
    @staticmethod
    def _style_sample_row(ws, row_num, num_columns):
        """Apply styling to sample data row"""
        for col_num in range(1, num_columns + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = TemplateGenerator.SAMPLE_FILL
            cell.border = TemplateGenerator.THIN_BORDER
    
    @staticmethod
    def _auto_size_columns(ws):
        """Auto-size columns based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def generate_daily_generation_template():
        """Generate Daily Generation Report template"""
        wb = Workbook()
        
        # Data sheet
        ws_data = wb.active
        ws_data.title = "Daily Generation Data"
        
        # Headers
        headers = [
            'Date (YYYY-MM-DD)',
            'Unit Number',
            'Generation (kWh)',
            'Operating Hours',
            'Capacity Factor (%)',
            'Availability Factor (%)',
            'Forced Outage Hours',
            'Scheduled Outage Hours',
            'Remarks'
        ]
        
        ws_data.append(headers)
        TemplateGenerator._style_header_row(ws_data, 1, headers)
        
        # Sample data
        today = datetime.now()
        sample_data = [
            today.strftime('%Y-%m-%d'),
            '1',
            '50000',
            '24',
            '85.5',
            '95.0',
            '0',
            '0',
            'Normal operation'
        ]
        ws_data.append(sample_data)
        TemplateGenerator._style_sample_row(ws_data, 2, len(headers))
        
        # Add a few empty rows for data entry
        for i in range(5):
            ws_data.append([''] * len(headers))
        
        TemplateGenerator._auto_size_columns(ws_data)
        
        # Instructions sheet
        ws_inst = wb.create_sheet("Instructions")
        instructions = [
            ["Daily Generation Report Template - Instructions"],
            [""],
            ["Column Descriptions:"],
            ["Date", "Format: YYYY-MM-DD (e.g., 2026-02-19)"],
            ["Unit Number", "Enter the unit number (1, 2, 3, etc.)"],
            ["Generation (kWh)", "Total energy generated in kilowatt-hours"],
            ["Operating Hours", "Number of hours the unit operated (0-24)"],
            ["Capacity Factor (%)", "Percentage of maximum possible generation (0-100)"],
            ["Availability Factor (%)", "Percentage of time unit was available (0-100)"],
            ["Forced Outage Hours", "Hours of unplanned outages"],
            ["Scheduled Outage Hours", "Hours of planned maintenance"],
            ["Remarks", "Any additional notes or comments"],
            [""],
            ["Important Notes:"],
            ["• All fields are required except Remarks"],
            ["• Date must be in YYYY-MM-DD format"],
            ["• Numeric values should not contain commas or currency symbols"],
            ["• Percentages should be entered as numbers (e.g., 85.5 not 85.5%)"],
            ["• Operating Hours cannot exceed 24"],
            ["• Delete the sample data row before uploading"],
            [""],
            ["For support, contact your system administrator."]
        ]
        
        for row in instructions:
            ws_inst.append(row)
        
        # Style instructions
        ws_inst['A1'].font = Font(bold=True, size=14)
        ws_inst['A3'].font = Font(bold=True, size=12)
        ws_inst['A14'].font = Font(bold=True, size=12)
        
        TemplateGenerator._auto_size_columns(ws_inst)
        
        return wb
    
    @staticmethod
    def generate_water_nomination_template():
        """Generate Water Nomination template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Water Nomination"
        
        # Headers
        headers = ['Hour'] + [f'Hour {i:02d}' for i in range(24)]
        ws.append(['Date:', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        ws.append(['Plant:', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        ws.append(['Nomination Type:', 'DAY_AHEAD or REAL_TIME', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        ws.append([])
        
        # Hourly data headers
        hourly_headers = ['Hour', 'Nominated MW']
        ws.append(hourly_headers)
        TemplateGenerator._style_header_row(ws, 5, hourly_headers)
        
        # Sample hourly data
        for hour in range(24):
            ws.append([f'{hour:02d}:00', '50.0'])
        
        TemplateGenerator._auto_size_columns(ws)
        
        # Instructions sheet
        ws_inst = wb.create_sheet("Instructions")
        instructions = [
            ["Water Nomination Template - Instructions"],
            [""],
            ["How to Fill:"],
            ["1. Enter the date in YYYY-MM-DD format in cell B1"],
            ["2. Enter the plant code in cell B2 (e.g., AGUS1, AGUS2)"],
            ["3. Enter nomination type in cell B3 (DAY_AHEAD or REAL_TIME)"],
            ["4. Fill in the nominated MW for each hour (00:00 to 23:00)"],
            [""],
            ["Important Notes:"],
            ["• All 24 hours must have values"],
            ["• MW values should be realistic for your plant capacity"],
            ["• Use decimal points for fractional MW (e.g., 50.5)"],
            ["• Do not leave any hour blank"],
            [""],
            ["For support, contact your system administrator."]
        ]
        
        for row in instructions:
            ws_inst.append(row)
        
        ws_inst['A1'].font = Font(bold=True, size=14)
        TemplateGenerator._auto_size_columns(ws_inst)
        
        return wb
    
    @staticmethod
    def generate_historical_data_template():
        """Generate Historical Data Import template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Historical Data"
        
        headers = [
            'Date (YYYY-MM-DD)',
            'Plant Code',
            'Generation (MWh)',
            'Capacity Factor (%)',
            'Availability Factor (%)',
            'Operating Hours',
            'Peak Load (MW)',
            'Average Load (MW)',
            'Remarks'
        ]
        
        ws.append(headers)
        TemplateGenerator._style_header_row(ws, 1, headers)
        
        # Sample data for last 7 days
        today = datetime.now()
        for i in range(7):
            date = today - timedelta(days=i)
            sample = [
                date.strftime('%Y-%m-%d'),
                'AGUS1',
                '1200',
                '85.5',
                '95.0',
                '24',
                '55.0',
                '50.0',
                'Normal operation'
            ]
            ws.append(sample)
            if i == 0:
                TemplateGenerator._style_sample_row(ws, 2, len(headers))
        
        TemplateGenerator._auto_size_columns(ws)
        
        # Instructions sheet
        ws_inst = wb.create_sheet("Instructions")
        instructions = [
            ["Historical Data Import Template - Instructions"],
            [""],
            ["Column Descriptions:"],
            ["Date", "Format: YYYY-MM-DD"],
            ["Plant Code", "Use official plant codes (AGUS1, AGUS2, etc.)"],
            ["Generation (MWh)", "Total daily generation in megawatt-hours"],
            ["Capacity Factor (%)", "Daily capacity factor percentage"],
            ["Availability Factor (%)", "Daily availability percentage"],
            ["Operating Hours", "Total operating hours for the day"],
            ["Peak Load (MW)", "Maximum load during the day"],
            ["Average Load (MW)", "Average load for the day"],
            ["Remarks", "Optional notes"],
            [""],
            ["Important Notes:"],
            ["• You can import multiple days at once"],
            ["• Dates should be in chronological order"],
            ["• All numeric fields are required"],
            ["• First row with gray background is sample data - delete before upload"],
            [""],
            ["For support, contact your system administrator."]
        ]
        
        for row in instructions:
            ws_inst.append(row)
        
        ws_inst['A1'].font = Font(bold=True, size=14)
        TemplateGenerator._auto_size_columns(ws_inst)
        
        return wb
    
    @staticmethod
    def generate_plant_capacity_template():
        """Generate Plant Capacity template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Plant Capacity"
        
        headers = [
            'Plant Code',
            'Plant Name',
            'Installed Capacity (MW)',
            'Dependable Capacity (MW)',
            'Number of Units',
            'Effective Date (YYYY-MM-DD)',
            'Remarks'
        ]
        
        ws.append(headers)
        TemplateGenerator._style_header_row(ws, 1, headers)
        
        # Sample data
        samples = [
            ['AGUS1', 'Agus 1 Hydroelectric Power Plant', '50.0', '48.0', '1', datetime.now().strftime('%Y-%m-%d'), 'Active'],
            ['AGUS2', 'Agus 2 Hydroelectric Power Plant', '100.0', '95.0', '2', datetime.now().strftime('%Y-%m-%d'), 'Active'],
        ]
        
        for sample in samples:
            ws.append(sample)
        
        TemplateGenerator._style_sample_row(ws, 2, len(headers))
        TemplateGenerator._auto_size_columns(ws)
        
        # Instructions sheet
        ws_inst = wb.create_sheet("Instructions")
        instructions = [
            ["Plant Capacity Template - Instructions"],
            [""],
            ["Column Descriptions:"],
            ["Plant Code", "Unique plant identifier (e.g., AGUS1)"],
            ["Plant Name", "Full name of the power plant"],
            ["Installed Capacity (MW)", "Total installed capacity"],
            ["Dependable Capacity (MW)", "Reliable capacity under normal conditions"],
            ["Number of Units", "Total number of generating units"],
            ["Effective Date", "Date when this capacity became effective (YYYY-MM-DD)"],
            ["Remarks", "Additional notes"],
            [""],
            ["Important Notes:"],
            ["• Plant Code must be unique"],
            ["• Capacities should be in MW (megawatts)"],
            ["• Delete sample data rows before uploading"],
            [""],
            ["For support, contact your system administrator."]
        ]
        
        for row in instructions:
            ws_inst.append(row)
        
        ws_inst['A1'].font = Font(bold=True, size=14)
        TemplateGenerator._auto_size_columns(ws_inst)
        
        return wb
    
    @staticmethod
    def generate_plant_status_template():
        """Generate Plant Status template for daily plant operational status"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Plant Status"
        
        headers = [
            'Date (YYYY-MM-DD)',
            'Plant Code',
            'Plant Name',
            'Status',
            'Installed Capacity (MW)',
            'Available Capacity (MW)',
            'Generation (MWh)',
            'Capacity Factor (%)',
            'Availability (%)',
            'Operating Units',
            'Total Units',
            'Remarks'
        ]
        
        ws.append(headers)
        TemplateGenerator._style_header_row(ws, 1, headers)
        
        # Sample data for multiple plants
        today = datetime.now()
        samples = [
            [
                today.strftime('%Y-%m-%d'),
                'AGUS1',
                'Agus 1 Hydroelectric Power Plant',
                'Operating',
                '50.0',
                '50.0',
                '1200.0',
                '100.0',
                '100.0',
                '1',
                '1',
                'Normal operation'
            ],
            [
                today.strftime('%Y-%m-%d'),
                'AGUS2',
                'Agus 2 Hydroelectric Power Plant',
                'Operating',
                '100.0',
                '95.0',
                '2200.0',
                '91.7',
                '95.0',
                '2',
                '2',
                'Unit 1 under maintenance'
            ],
            [
                today.strftime('%Y-%m-%d'),
                'PULANGI4',
                'Pulangi 4 Hydroelectric Power Plant',
                'Operating',
                '255.0',
                '255.0',
                '6000.0',
                '98.0',
                '100.0',
                '4',
                '4',
                'All units operational'
            ],
        ]
        
        for i, sample in enumerate(samples):
            ws.append(sample)
            if i == 0:
                TemplateGenerator._style_sample_row(ws, 2, len(headers))
        
        # Add empty rows for data entry
        for _ in range(5):
            ws.append([''] * len(headers))
        
        TemplateGenerator._auto_size_columns(ws)
        
        # Instructions sheet
        ws_inst = wb.create_sheet("Instructions")
        instructions = [
            ["Plant Status Template - Instructions"],
            [""],
            ["Purpose:"],
            ["This template is used to upload daily operational status for all power plants."],
            ["It provides a snapshot of plant performance and availability."],
            [""],
            ["Column Descriptions:"],
            ["Date", "Report date in YYYY-MM-DD format (e.g., 2026-02-23)"],
            ["Plant Code", "Official plant code (AGUS1, AGUS2, AGUS4, AGUS5, AGUS6, AGUS7, PULANGI4)"],
            ["Plant Name", "Full name of the power plant"],
            ["Status", "Current operational status (Operating, Maintenance, Shutdown, Standby)"],
            ["Installed Capacity (MW)", "Total installed capacity in megawatts"],
            ["Available Capacity (MW)", "Currently available capacity (may be less due to maintenance)"],
            ["Generation (MWh)", "Total energy generated for the day in megawatt-hours"],
            ["Capacity Factor (%)", "Percentage of maximum possible generation (0-100)"],
            ["Availability (%)", "Percentage of installed capacity available (0-100)"],
            ["Operating Units", "Number of units currently operating"],
            ["Total Units", "Total number of units in the plant"],
            ["Remarks", "Additional notes, maintenance info, or operational comments"],
            [""],
            ["Valid Status Values:"],
            ["• Operating - Plant is generating power"],
            ["• Maintenance - Plant is under scheduled maintenance"],
            ["• Shutdown - Plant is temporarily shut down"],
            ["• Standby - Plant is ready but not generating"],
            [""],
            ["Important Notes:"],
            ["• All fields are required except Remarks"],
            ["• Date must be in YYYY-MM-DD format"],
            ["• Plant Code must match existing plants in the system"],
            ["• Available Capacity cannot exceed Installed Capacity"],
            ["• Operating Units cannot exceed Total Units"],
            ["• Percentages should be between 0 and 100"],
            ["• Delete the sample data rows (gray background) before uploading"],
            ["• You can upload status for multiple plants in one file"],
            [""],
            ["Example Use Cases:"],
            ["• Daily plant status reporting"],
            ["• Bulk import of historical plant status data"],
            ["• Monthly or annual status data compilation"],
            [""],
            ["For support or questions, contact your system administrator."]
        ]
        
        for row in instructions:
            ws_inst.append(row)
        
        # Style instructions
        ws_inst['A1'].font = Font(bold=True, size=14, color="4472C4")
        ws_inst['A3'].font = Font(bold=True, size=12)
        ws_inst['A7'].font = Font(bold=True, size=12)
        ws_inst['A21'].font = Font(bold=True, size=12)
        ws_inst['A24'].font = Font(bold=True, size=12)
        ws_inst['A33'].font = Font(bold=True, size=12)
        
        # Add some color to important sections
        ws_inst['A3'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws_inst['A21'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws_inst['A24'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        TemplateGenerator._auto_size_columns(ws_inst)
        
        return wb
    
    @staticmethod
    def generate_psr_template():
        """Generate PSR (Plant Status Report) template with right side section"""
        wb = Workbook()
        ws = wb.active
        ws.title = "PSR Template"
        
        # Define styles
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_blue_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        header_font = Font(bold=True, color="FFFFFF", size=10)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=9)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 8
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col].width = 10
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 3
        # Right side columns
        for col in ['N', 'O', 'P', 'Q', 'R', 'S', 'T']:
            ws.column_dimensions[col].width = 12
        
        # Header section (rows 1-5)
        ws.merge_cells('A1:L1')
        ws['A1'] = 'NATIONAL POWER CORPORATION'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:L2')
        ws['A2'] = 'MINDANAO GENERATION'
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A4:L4')
        ws['A4'] = 'PLANT STATUS REPORT (PSR)'
        ws['A4'].font = Font(bold=True, size=12)
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A5:L5')
        ws['A5'] = 'Date: [Enter Date Here]'
        ws['A5'].font = bold_font
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Main table headers (row 7)
        main_headers = ['Unit', 'Rated\nCap\n(MW)', 'Dep\nCap\n(MW)', 'GENFCST', 'GENFCST', 'GENFCST', 
                       'GENFCST', 'GENFCST', 'GENFCST', 'GENFCST', 'GENFCST', 'Remarks']
        
        for col_num, header in enumerate(main_headers, 1):
            cell = ws.cell(row=7, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Sample plant data (rows 8-15)
        sample_plants = [
            ['AGUS1', '', '', '', '', '', '', '', '', '', '', ''],
            ['  Unit 1', '50.0', '48.0', '0.00', '0.00', '0.00', '0.00', '0.00', '0.00', '0.00', '50.00', 'Sample data'],
            ['AGUS2', '', '', '', '', '', '', '', '', '', '', ''],
            ['  Unit 1', '50.0', '48.0', '0.00', '0.00', '0.00', '0.00', '0.00', '0.00', '0.00', '50.00', ''],
            ['  Unit 2', '50.0', '48.0', '0.00', '0.00', '0.00', '0.00', '0.00', '0.00', '0.00', '50.00', ''],
        ]
        
        current_row = 8
        for plant_row in sample_plants:
            for col_num, value in enumerate(plant_row, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center' if col_num > 1 else 'left', vertical='center')
                if not plant_row[0].startswith('  '):
                    cell.fill = light_gray_fill
                    cell.font = bold_font
            current_row += 1
        
        # RIGHT SIDE SECTION
        # Primary Storage of Hydro HEPs (N4:T10)
        ws.merge_cells('N4:T4')
        ws['N4'] = 'PRIMARY STORAGE OF HYDRO HEPs'
        ws['N4'].font = Font(bold=True, size=10, color="FFFFFF")
        ws['N4'].fill = header_fill
        ws['N4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['N4'].border = thin_border
        
        # Storage table headers
        storage_headers = ['Lake/Dam', 'Level (m)', 'Remarks']
        ws.merge_cells('N5:P5')
        ws['N5'] = storage_headers[0]
        ws['N5'].fill = yellow_fill
        ws['N5'].font = bold_font
        ws['N5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['N5'].border = thin_border
        
        ws.merge_cells('Q5:R5')
        ws['Q5'] = storage_headers[1]
        ws['Q5'].fill = yellow_fill
        ws['Q5'].font = bold_font
        ws['Q5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['Q5'].border = thin_border
        
        ws.merge_cells('S5:T5')
        ws['S5'] = storage_headers[2]
        ws['S5'].fill = yellow_fill
        ws['S5'].font = bold_font
        ws['S5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['S5'].border = thin_border
        
        # Storage data
        storage_data = [
            ['Lake Lanao', '701.20', 'Normal'],
            ['Agus 2 Forebay', '637.30', ''],
            ['Agus 4 Forebay', '358.50', ''],
            ['Agus 5 Forebay', '242.80', ''],
            ['Agus 6 Forebay', '199.80', ''],
        ]
        
        for i, row_data in enumerate(storage_data, 6):
            ws.merge_cells(f'N{i}:P{i}')
            ws[f'N{i}'] = row_data[0]
            ws[f'N{i}'].border = thin_border
            ws[f'N{i}'].alignment = Alignment(horizontal='left', vertical='center')
            
            ws.merge_cells(f'Q{i}:R{i}')
            ws[f'Q{i}'] = row_data[1]
            ws[f'Q{i}'].border = thin_border
            ws[f'Q{i}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells(f'S{i}:T{i}')
            ws[f'S{i}'] = row_data[2]
            ws[f'S{i}'].border = thin_border
            ws[f'S{i}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Hydro Inflow/Outflow (N12:T18)
        ws.merge_cells('N12:T12')
        ws['N12'] = 'HYDRO INFLOW/OUTFLOW (cms)'
        ws['N12'].font = Font(bold=True, size=10, color="FFFFFF")
        ws['N12'].fill = header_fill
        ws['N12'].alignment = Alignment(horizontal='center', vertical='center')
        ws['N12'].border = thin_border
        
        # Inflow/Outflow headers
        ws.merge_cells('N13:P13')
        ws['N13'] = 'Plant'
        ws['N13'].fill = yellow_fill
        ws['N13'].font = bold_font
        ws['N13'].alignment = Alignment(horizontal='center', vertical='center')
        ws['N13'].border = thin_border
        
        ws['Q13'] = 'Inflow'
        ws['Q13'].fill = yellow_fill
        ws['Q13'].font = bold_font
        ws['Q13'].alignment = Alignment(horizontal='center', vertical='center')
        ws['Q13'].border = thin_border
        
        ws['R13'] = 'Outflow'
        ws['R13'].fill = yellow_fill
        ws['R13'].font = bold_font
        ws['R13'].alignment = Alignment(horizontal='center', vertical='center')
        ws['R13'].border = thin_border
        
        ws.merge_cells('S13:T13')
        ws['S13'] = 'Remarks'
        ws['S13'].fill = yellow_fill
        ws['S13'].font = bold_font
        ws['S13'].alignment = Alignment(horizontal='center', vertical='center')
        ws['S13'].border = thin_border
        
        # Inflow/Outflow data
        flow_data = [
            ['Lake Lanao', '0.00', '0.00', ''],
            ['Agus 1', '0.00', '0.00', ''],
            ['Agus 2', '0.00', '0.00', ''],
            ['Agus 4', '0.00', '0.00', ''],
            ['Agus 5', '0.00', '0.00', ''],
        ]
        
        for i, row_data in enumerate(flow_data, 14):
            ws.merge_cells(f'N{i}:P{i}')
            ws[f'N{i}'] = row_data[0]
            ws[f'N{i}'].border = thin_border
            ws[f'N{i}'].alignment = Alignment(horizontal='left', vertical='center')
            
            ws[f'Q{i}'] = row_data[1]
            ws[f'Q{i}'].border = thin_border
            ws[f'Q{i}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'R{i}'] = row_data[2]
            ws[f'R{i}'].border = thin_border
            ws[f'R{i}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells(f'S{i}:T{i}')
            ws[f'S{i}'] = row_data[3]
            ws[f'S{i}'].border = thin_border
            ws[f'S{i}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Generation Data (N20:T26)
        ws.merge_cells('N20:T20')
        ws['N20'] = 'GENERATION DATA (MWh)'
        ws['N20'].font = Font(bold=True, size=10, color="FFFFFF")
        ws['N20'].fill = header_fill
        ws['N20'].alignment = Alignment(horizontal='center', vertical='center')
        ws['N20'].border = thin_border
        
        # Generation headers
        ws.merge_cells('N21:P21')
        ws['N21'] = 'Plant'
        ws['N21'].fill = yellow_fill
        ws['N21'].font = bold_font
        ws['N21'].alignment = Alignment(horizontal='center', vertical='center')
        ws['N21'].border = thin_border
        
        ws['Q21'] = 'Today'
        ws['Q21'].fill = yellow_fill
        ws['Q21'].font = bold_font
        ws['Q21'].alignment = Alignment(horizontal='center', vertical='center')
        ws['Q21'].border = thin_border
        
        ws['R21'] = 'MTD'
        ws['R21'].fill = yellow_fill
        ws['R21'].font = bold_font
        ws['R21'].alignment = Alignment(horizontal='center', vertical='center')
        ws['R21'].border = thin_border
        
        ws.merge_cells('S21:T21')
        ws['S21'] = 'YTD'
        ws['S21'].fill = yellow_fill
        ws['S21'].font = bold_font
        ws['S21'].alignment = Alignment(horizontal='center', vertical='center')
        ws['S21'].border = thin_border
        
        # Generation data
        gen_data = [
            ['AGUS1', '1200', '36000', '72000'],
            ['AGUS2', '2400', '72000', '144000'],
            ['PULANGI4', '6000', '180000', '360000'],
            ['Total NPC', '9600', '288000', '576000'],
        ]
        
        for i, row_data in enumerate(gen_data, 22):
            ws.merge_cells(f'N{i}:P{i}')
            ws[f'N{i}'] = row_data[0]
            ws[f'N{i}'].border = thin_border
            ws[f'N{i}'].alignment = Alignment(horizontal='left', vertical='center')
            if 'Total' in row_data[0]:
                ws[f'N{i}'].font = bold_font
                ws[f'N{i}'].fill = light_gray_fill
            
            ws[f'Q{i}'] = row_data[1]
            ws[f'Q{i}'].border = thin_border
            ws[f'Q{i}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Total' in row_data[0]:
                ws[f'Q{i}'].font = bold_font
                ws[f'Q{i}'].fill = light_gray_fill
            
            ws[f'R{i}'] = row_data[2]
            ws[f'R{i}'].border = thin_border
            ws[f'R{i}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Total' in row_data[0]:
                ws[f'R{i}'].font = bold_font
                ws[f'R{i}'].fill = light_gray_fill
            
            ws.merge_cells(f'S{i}:T{i}')
            ws[f'S{i}'] = row_data[3]
            ws[f'S{i}'].border = thin_border
            ws[f'S{i}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Total' in row_data[0]:
                ws[f'S{i}'].font = bold_font
                ws[f'S{i}'].fill = light_gray_fill
        
        # Capacity Factor (N28:T34)
        ws.merge_cells('N28:T28')
        ws['N28'] = 'CAPACITY FACTOR (%)'
        ws['N28'].font = Font(bold=True, size=10, color="FFFFFF")
        ws['N28'].fill = header_fill
        ws['N28'].alignment = Alignment(horizontal='center', vertical='center')
        ws['N28'].border = thin_border
        
        # Capacity Factor headers
        ws.merge_cells('N29:P29')
        ws['N29'] = 'Plant'
        ws['N29'].fill = yellow_fill
        ws['N29'].font = bold_font
        ws['N29'].alignment = Alignment(horizontal='center', vertical='center')
        ws['N29'].border = thin_border
        
        ws['Q29'] = 'Today'
        ws['Q29'].fill = yellow_fill
        ws['Q29'].font = bold_font
        ws['Q29'].alignment = Alignment(horizontal='center', vertical='center')
        ws['Q29'].border = thin_border
        
        ws['R29'] = 'MTD'
        ws['R29'].fill = yellow_fill
        ws['R29'].font = bold_font
        ws['R29'].alignment = Alignment(horizontal='center', vertical='center')
        ws['R29'].border = thin_border
        
        ws.merge_cells('S29:T29')
        ws['S29'] = 'YTD'
        ws['S29'].fill = yellow_fill
        ws['S29'].font = bold_font
        ws['S29'].alignment = Alignment(horizontal='center', vertical='center')
        ws['S29'].border = thin_border
        
        # Capacity Factor data
        cf_data = [
            ['AGUS1', '100.0', '100.0', '100.0'],
            ['AGUS2', '100.0', '100.0', '100.0'],
            ['PULANGI4', '98.0', '98.5', '98.2'],
            ['Average', '99.3', '99.5', '99.4'],
        ]
        
        for i, row_data in enumerate(cf_data, 30):
            ws.merge_cells(f'N{i}:P{i}')
            ws[f'N{i}'] = row_data[0]
            ws[f'N{i}'].border = thin_border
            ws[f'N{i}'].alignment = Alignment(horizontal='left', vertical='center')
            if 'Average' in row_data[0]:
                ws[f'N{i}'].font = bold_font
                ws[f'N{i}'].fill = light_gray_fill
            
            ws[f'Q{i}'] = row_data[1]
            ws[f'Q{i}'].border = thin_border
            ws[f'Q{i}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Average' in row_data[0]:
                ws[f'Q{i}'].font = bold_font
                ws[f'Q{i}'].fill = light_gray_fill
            
            ws[f'R{i}'] = row_data[2]
            ws[f'R{i}'].border = thin_border
            ws[f'R{i}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Average' in row_data[0]:
                ws[f'R{i}'].font = bold_font
                ws[f'R{i}'].fill = light_gray_fill
            
            ws.merge_cells(f'S{i}:T{i}')
            ws[f'S{i}'] = row_data[3]
            ws[f'S{i}'].border = thin_border
            ws[f'S{i}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Average' in row_data[0]:
                ws[f'S{i}'].font = bold_font
                ws[f'S{i}'].fill = light_gray_fill
        
        # Instructions sheet
        ws_inst = wb.create_sheet("Instructions")
        instructions = [
            ["PSR Template - Instructions"],
            [""],
            ["Purpose:"],
            ["This template is used to create Plant Status Reports (PSR) with comprehensive"],
            ["operational data including generation, storage levels, and flow rates."],
            [""],
            ["Main Table (Left Side):"],
            ["• Unit - Plant and unit identification"],
            ["• Rated Cap (MW) - Installed capacity"],
            ["• Dep Cap (MW) - Dependable capacity"],
            ["• GENFCST columns - Generation forecast data"],
            ["• Remarks - Operational notes"],
            [""],
            ["Right Side Sections:"],
            [""],
            ["1. PRIMARY STORAGE OF HYDRO HEPs"],
            ["   • Lake/Dam levels in meters above sea level"],
            ["   • Current status and remarks"],
            [""],
            ["2. HYDRO INFLOW/OUTFLOW"],
            ["   • Inflow and outflow rates in cubic meters per second (cms)"],
            ["   • Data for each plant and reservoir"],
            [""],
            ["3. GENERATION DATA"],
            ["   • Today - Current day generation in MWh"],
            ["   • MTD - Month-to-date generation"],
            ["   • YTD - Year-to-date generation"],
            [""],
            ["4. CAPACITY FACTOR"],
            ["   • Percentage of maximum possible generation"],
            ["   • Today, MTD, and YTD values"],
            [""],
            ["Important Notes:"],
            ["• All numeric values should be entered without commas"],
            ["• Dates should be in YYYY-MM-DD format"],
            ["• Delete sample data before entering actual data"],
            ["• Ensure all calculations are accurate"],
            ["• Review all sections before finalizing"],
            [""],
            ["For support, contact your system administrator."]
        ]
        
        for row in instructions:
            ws_inst.append(row)
        
        ws_inst['A1'].font = Font(bold=True, size=14, color="2F5496")
        ws_inst['A3'].font = Font(bold=True, size=12)
        ws_inst['A7'].font = Font(bold=True, size=12)
        ws_inst['A13'].font = Font(bold=True, size=12)
        ws_inst['A35'].font = Font(bold=True, size=12)
        
        TemplateGenerator._auto_size_columns(ws_inst)
        
        return wb
    
    @staticmethod
    def create_http_response(workbook, filename):
        """Create HTTP response with Excel file"""
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
