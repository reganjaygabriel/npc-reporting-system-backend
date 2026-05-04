from datetime import datetime, timedelta
from django.utils import timezone
from django.core.mail import EmailMessage
from django.conf import settings
import os
from ..models import GenerationReport, Plant
from ..models_scheduled import ScheduledReport, ReportExecution
import logging

logger = logging.getLogger(__name__)


class AutomatedReportService:
    """Service for generating and sending automated reports"""
    
    def __init__(self):
        pass
    
    def get_due_reports(self):
        """Get all reports that are due to run"""
        now = timezone.now()
        return ScheduledReport.objects.filter(
            status='ACTIVE',
            next_run__lte=now
        )
    
    def execute_report(self, scheduled_report):
        """Execute a scheduled report"""
        execution = ReportExecution.objects.create(
            scheduled_report=scheduled_report,
            status='RUNNING'
        )
        
        try:
            logger.info(f"Starting report execution: {scheduled_report.name}")
            
            # Generate report data
            data = self._generate_report_data(scheduled_report)
            logger.info(f"Generated {len(data)} records")
            
            # Create report file even if no data
            file_path = self._create_report_file(scheduled_report, data)
            logger.info(f"Created report file: {file_path}")
            
            # Send to recipients (skip if email not configured)
            sent, failed = self._send_report(scheduled_report, file_path)
            logger.info(f"Email sent: {sent}, failed: {failed}")
            
            # Update execution
            execution.status = 'COMPLETED'
            execution.completed_at = timezone.now()
            execution.duration_seconds = (execution.completed_at - execution.started_at).seconds
            execution.file_path = file_path
            execution.file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            execution.records_processed = len(data)
            execution.recipients_sent = sent
            execution.recipients_failed = failed
            execution.save()
            
            # Update scheduled report
            scheduled_report.last_run = timezone.now()
            scheduled_report.next_run = self._calculate_next_run(scheduled_report)
            scheduled_report.run_count += 1
            scheduled_report.save()
            
            logger.info(f"Report executed successfully: {scheduled_report.name}")
            return True
            
        except Exception as e:
            execution.status = 'FAILED'
            execution.completed_at = timezone.now()
            execution.error_message = str(e)
            execution.duration_seconds = (execution.completed_at - execution.started_at).seconds if execution.completed_at else 0
            execution.save()
            
            logger.error(f"Report execution failed: {scheduled_report.name} - {str(e)}", exc_info=True)
            raise  # Re-raise to let the view handle it
    
    def _generate_report_data(self, scheduled_report):
        """Generate report data based on type"""
        try:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=scheduled_report.date_range_days)
            
            # Get plants
            plants = scheduled_report.plants.all()
            if not plants.exists():
                plants = Plant.objects.filter(is_active=True)
            
            if not plants.exists():
                logger.warning(f"No plants found for report: {scheduled_report.name}")
                return []
            
            # Generate PSR report data
            return self._get_psr_data(plants, start_date, end_date)
        except Exception as e:
            logger.error(f"Error generating report data: {str(e)}")
            raise
    
    def _get_psr_data(self, plants, start_date, end_date):
        """Get Plant Status Report data"""
        reports = GenerationReport.objects.filter(
            plant__in=plants,
            report_date__range=[start_date, end_date]
        ).select_related('plant', 'unit').order_by('report_date', 'plant', 'unit')
        
        return [{
            'date': r.report_date,
            'plant': r.plant.name,
            'plant_code': r.plant.code,
            'unit': r.unit.unit_number,
            'generation_kwh': float(r.generation_kwh),
            'operating_hours': float(r.operating_hours),
            'capacity_factor': float(r.capacity_factor or 0),
            'availability_factor': float(r.availability_factor or 0),
            'forced_outage_hours': float(r.forced_outage_hours or 0),
            'scheduled_outage_hours': float(r.scheduled_outage_hours or 0)
        } for r in reports]
    
    def _create_report_file(self, scheduled_report, data):
        """Create PSR report file"""
        from ..services.psr_exporter import PSRExporter
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"PSR_AUTOMATED_{timestamp}.xlsx"
        
        # Create exports directory if not exists
        export_dir = os.path.join(settings.MEDIA_ROOT, 'automated_reports')
        os.makedirs(export_dir, exist_ok=True)
        file_path = os.path.join(export_dir, filename)
        
        # Get report date (use most recent date from data or today)
        report_date = timezone.now().date()
        if data:
            report_date = max(item['date'] for item in data)
        
        # Get GenerationReport queryset for PSR exporter
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=scheduled_report.date_range_days)
        
        plants = scheduled_report.plants.all()
        if not plants.exists():
            plants = Plant.objects.filter(is_active=True)
        
        reports = GenerationReport.objects.filter(
            plant__in=plants,
            report_date__range=[start_date, end_date]
        ).select_related('plant', 'unit').order_by('report_date', 'plant', 'unit')
        
        # Generate PSR using PSRExporter
        if reports.exists():
            exporter = PSRExporter(reports, report_date)
            file_path = exporter.generate()
        else:
            # Create empty report if no data
            self._create_empty_psr_report(file_path, report_date)
        
        return file_path
    
    def _create_empty_psr_report(self, file_path, report_date):
        """Create an empty PSR report when no data is available"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Plant Status Report"
        
        # Header
        ws['A1'] = 'PLANT STATUS REPORT (PSR)'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Date: {report_date.strftime('%Y-%m-%d')}"
        ws['A3'] = 'No data available for the selected period'
        ws['A3'].font = Font(color='FF0000')
        
        wb.save(file_path)
    
    def _send_report(self, scheduled_report, file_path):
        """Send report to recipients"""
        try:
            recipients = list(scheduled_report.recipients.values_list('email', flat=True))
            
            # Add additional emails
            if scheduled_report.additional_emails:
                additional = [e.strip() for e in scheduled_report.additional_emails.split('\n') if e.strip()]
                recipients.extend(additional)
            
            if not recipients:
                logger.info(f"No recipients for report: {scheduled_report.name}")
                return 0, 0
            
            subject = f"Automated Report: {scheduled_report.name}"
            body = f"""
Dear Recipient,

Please find attached the automated report: {scheduled_report.name}

Report Type: {scheduled_report.get_report_type_display()}
Frequency: {scheduled_report.get_frequency_display()}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}

This is an automated message from GPD Reporting System (Generation and Performance Division).
            """
            
            email = EmailMessage(
                subject=subject,
                body=body,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@gpd.gov.ph'),
                to=recipients
            )
            
            if os.path.exists(file_path):
                email.attach_file(file_path)
            
            email.send()
            logger.info(f"Report sent to {len(recipients)} recipients")
            return len(recipients), 0
            
        except Exception as e:
            logger.error(f"Failed to send report: {str(e)}")
            return 0, len(recipients) if 'recipients' in locals() else 0
    
    def _calculate_next_run(self, scheduled_report):
        """Calculate next run time"""
        now = timezone.now()
        schedule_time = scheduled_report.schedule_time
        
        if scheduled_report.frequency == 'DAILY':
            next_run = now.replace(hour=schedule_time.hour, minute=schedule_time.minute, second=0, microsecond=0)
            if next_run <= now:
                next_run += timedelta(days=1)
        
        elif scheduled_report.frequency == 'WEEKLY':
            next_run = now.replace(hour=schedule_time.hour, minute=schedule_time.minute, second=0, microsecond=0)
            days_ahead = (scheduled_report.schedule_day - now.weekday()) % 7
            if days_ahead == 0 and next_run <= now:
                days_ahead = 7
            next_run += timedelta(days=days_ahead)
        
        elif scheduled_report.frequency == 'MONTHLY':
            next_run = now.replace(day=scheduled_report.schedule_day, hour=schedule_time.hour, 
                                  minute=schedule_time.minute, second=0, microsecond=0)
            if next_run <= now:
                # Move to next month
                if now.month == 12:
                    next_run = next_run.replace(year=now.year + 1, month=1)
                else:
                    next_run = next_run.replace(month=now.month + 1)
        
        else:  # QUARTERLY
            next_run = now.replace(hour=schedule_time.hour, minute=schedule_time.minute, second=0, microsecond=0)
            next_run += timedelta(days=90)
        
        return next_run
