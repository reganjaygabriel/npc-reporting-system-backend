"""
Plant Status Report (PSR) Excel Exporter
Generates Excel reports matching 100% the exact PSR format from PSR REPORT-8AM.xlsx
INCLUDING: Forecasted Load, IPP sections, Charts, and Notes
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from datetime import datetime
import os
from django.conf import settings


class PSRExporter:
    """Service class for generating Plant Status Report (PSR) Excel files"""
    
    # Plant configuration matching the template EXACTLY
    PLANTS_CONFIG = {
        'AGUS1': {
            'name': 'AGUS 1',
            'units': [
                {'num': 1, 'label': 'unit 1', 'capacity': 40, 'nominated': 0},
                {'num': 2, 'label': 'unit 2', 'capacity': 40, 'nominated': 0}
            ]
        },
        'AGUS2': {
            'name': 'AGUS 2',
            'units': [
                {'num': 1, 'label': 'unit  1', 'capacity': 60, 'nominated': 60},
                {'num': 2, 'label': 'unit  2', 'capacity': 60, 'nominated': 60},
                {'num': 3, 'label': 'unit  3', 'capacity': 60, 'nominated': 60}
            ]
        },
        'AGUS4': {
            'name': 'AGUS 4',
            'units': [
                {'num': 1, 'label': 'unit  1', 'capacity': 52.7, 'nominated': 0},
                {'num': 2, 'label': 'unit  2', 'capacity': 52.7, 'nominated': 52.7},
                {'num': 3, 'label': 'unit  3', 'capacity': 52.7, 'nominated': 52.7}
            ]
        },
        'AGUS5': {
            'name': 'AGUS 5',
            'units': [
                {'num': 1, 'label': 'unit 1', 'capacity': 27.5, 'nominated': 27.5},
                {'num': 2, 'label': 'unit 2', 'capacity': 27.5, 'nominated': 27.5}
            ]
        },
        'AGUS6': {
            'name': 'AGUS 6',
            'units': [
                {'num': 1, 'label': '  unit  1', 'capacity': 34.5, 'nominated': 20},
                {'num': 2, 'label': '  unit  2', 'capacity': 34.5, 'nominated': 21},
                {'num': 3, 'label': 'unit  3', 'capacity': 50, 'nominated': 42},
                {'num': 4, 'label': 'unit  4', 'capacity': 50, 'nominated': 38},
                {'num': 5, 'label': 'unit  5', 'capacity': 50, 'nominated': 44}
            ]
        },
        'AGUS7': {
            'name': 'AGUS 7',
            'units': [
                {'num': 1, 'label': 'unit 1', 'capacity': 27, 'nominated': 27},
                {'num': 2, 'label': 'unit 2', 'capacity': 27, 'nominated': 27}
            ]
        },
        'PULANGI4': {
            'name': 'PULANGI IV',
            'units': [
                {'num': 1, 'label': 'unit  1', 'capacity': 85, 'nominated': 75},
                {'num': 2, 'label': 'unit  2', 'capacity': 85, 'nominated': 70},
                {'num': 3, 'label': 'unit  3', 'capacity': 85, 'nominated': 70}
            ]
        }
    }
    
    def __init__(self, queryset, report_date, report_type='psr'):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        self.queryset = queryset
        self.report_date = report_date
        self.report_type = report_type  # 'psr' or 'daily_status'
        self.data_by_plant = self._organize_data()
    
    def _organize_data(self):
        """Organize queryset data by plant and unit"""
        data = {}
        for report in self.queryset:
            plant_code = report.plant.code.upper()
            unit_num = report.unit.unit_number
            
            if plant_code not in data:
                data[plant_code] = {}
            
            data[plant_code][unit_num] = {
                'generation': float(report.generation_kwh),
                'operating_hours': float(report.operating_hours),
                'forced_outage': float(report.forced_outage_hours),
                'scheduled_outage': float(report.scheduled_outage_hours),
                'remarks': report.remarks or ''
            }
        
        return data
    
    def generate(self):
        """Generate PSR Excel file matching 100% the exact format"""
        wb = Workbook()
        ws = wb.active
        ws.title = "PSR PSALM Edit (2)"
        
        # Set column widths EXACTLY as template
        ws.column_dimensions['A'].width = 25.86
        ws.column_dimensions['B'].width = 15.14
        ws.column_dimensions['C'].width = 13.71
        ws.column_dimensions['D'].width = 15.71
        ws.column_dimensions['E'].width = 13.71
        ws.column_dimensions['F'].width = 13.57
        ws.column_dimensions['G'].width = 13.29
        ws.column_dimensions['H'].width = 14.71
        ws.column_dimensions['I'].width = 13.0
        ws.column_dimensions['J'].width = 13.0
        ws.column_dimensions['K'].width = 13.0
        ws.column_dimensions['L'].width = 13.0
        ws.column_dimensions['M'].width = 15.71
        ws.column_dimensions['N'].width = 14.71
        # Right side columns
        ws.column_dimensions['O'].width = 3.0
        ws.column_dimensions['P'].width = 15.0
        ws.column_dimensions['Q'].width = 12.0
        ws.column_dimensions['R'].width = 12.0
        ws.column_dimensions['S'].width = 15.0
        ws.column_dimensions['T'].width = 3.0
        ws.column_dimensions['U'].width = 10.0
        ws.column_dimensions['V'].width = 10.0
        ws.column_dimensions['W'].width = 10.0
        ws.column_dimensions['X'].width = 10.0
        ws.column_dimensions['Y'].width = 10.0
        ws.column_dimensions['Z'].width = 10.0
        ws.column_dimensions['AA'].width = 12.0
        
        # Add all sections
        self._add_header(ws)
        self._add_column_headers(ws)
        current_row = self._add_plant_data(ws)
        current_row = self._add_forecasted_load(ws, current_row)
        current_row = self._add_ipp_section(ws, current_row)
        current_row = self._add_notes_section(ws, current_row)
        current_row = self._add_footer(ws, current_row)
        
        # Add chart and Agus 2 note at the very bottom left
        self._add_chart_and_agus2_note(ws, current_row)
        
        # Add right side sections
        self._add_right_side_sections(ws)
        
        # Save file
        file_path = self._get_file_path()
        wb.save(file_path)
        return file_path
    
    def _add_header(self, ws):
        """Add header section with dark gray/teal styling for both report types"""
        # Both reports now use dark gray/teal header
        header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        header_font_color = "FFFFFF"
        
        # Determine title and time text based on report type
        if self.report_type == 'daily_status':
            title_text = 'DAILY PLANT STATUS'
            time_text = f'as of 12:00 NN    {self.report_date.strftime("%A, %B %d, %Y")}'
        else:
            title_text = ' PLANT STATUS REPORT'
            time_text = f'as of 0800H {self.report_date.strftime("%A, %d %B %Y")}'
        
        # Row 1: Empty with specific formatting
        ws['A1'] = ' '
        ws['A1'].font = Font(size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 29.45
        
        # Row 2: MINDANAO GENERATION
        ws['A2'] = 'MINDANAO GENERATION'
        ws['A2'].font = Font(size=20, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:N2')
        ws.row_dimensions[2].height = 25.5
        
        # Row 3: (PSALM PORTFOLIO)
        ws['A3'] = '(PSALM PORTFOLIO)'
        ws['A3'].font = Font(size=14, bold=True)
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A3:N3')
        ws.row_dimensions[3].height = 14.25
        
        # Row 4-5: Empty
        ws.row_dimensions[4].height = 15.95
        ws.row_dimensions[5].height = 8.25
        
        # Row 6: FOR section
        ws['A6'] = 'FOR     :'
        ws['A6'].font = Font(size=14, bold=True)
        ws['A6'].alignment = Alignment(horizontal='right')
        
        ws['B6'] = 'MR. LARRY I. SABELLINA'
        ws['B6'].font = Font(size=14, bold=True)
        
        ws['G6'] = 'MR. DENNIS EDWARD A. DELA SERNA'
        ws['G6'].font = Font(size=14, bold=True)
        
        ws['K6'] = 'MR. ARNOLD C. FRANCISCO'
        ws['K6'].font = Font(size=14, bold=True)
        
        ws.row_dimensions[6].height = 15.95
        
        # Row 7: Titles
        ws['B7'] = 'VP, Mindanao Generation'
        ws['B7'].font = Font(size=12)
        
        ws['G7'] = 'President and CEO, PSALM'
        ws['G7'].font = Font(size=12)
        
        ws['K7'] = 'VP - PAMG, PSALM'
        ws['K7'].font = Font(size=12)
        ws['K7'].alignment = Alignment(horizontal='left')
        
        ws.row_dimensions[7].height = 15.95
        
        # Row 8: Empty
        ws.row_dimensions[8].height = 8.25
        
        # Row 9: PLANT STATUS REPORT or DAILY PLANT STATUS (with dark gray/teal background)
        ws['A9'] = title_text
        ws['A9'].font = Font(size=18, bold=True, italic=True, color=header_font_color)
        ws['A9'].fill = header_fill
        ws['A9'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A9:N9')
        ws.row_dimensions[9].height = 24.95
        
        # Row 10: Date (with dark gray/teal background)
        ws['A10'] = time_text
        ws['A10'].font = Font(size=14, bold=True, color=header_font_color)
        ws['A10'].fill = header_fill
        ws['A10'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A10:N10')
        ws.row_dimensions[10].height = 22.5
        
        # Row 11: Empty
        ws.row_dimensions[11].height = 7.5
        
        # Row 12: Separator
        ws['A12'] = ' '
        ws['A12'].font = Font(size=9, bold=True)
        ws['A12'].alignment = Alignment(horizontal='left')
        ws.row_dimensions[12].height = 3.75
    
    def _add_column_headers(self, ws):
        """Add column headers EXACTLY as template (rows 13-16)"""
        # Row 13-16: Headers with exact formatting
        ws['A13'] = 'PLANT NAME'
        ws['A13'].font = Font(size=14, bold=True)
        ws['A13'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('A13:A16')
        
        ws['B13'] = 'Rated Capacity (MW)'
        ws['B13'].font = Font(size=12, bold=True)
        ws['B13'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('B13:B16')
        
        ws['C13'] = 'Nominated'
        ws['C13'].font = Font(size=12, bold=True)
        ws['C13'].alignment = Alignment(horizontal='center')
        
        ws['C16'] = 'Capability'
        ws['C16'].font = Font(size=12, bold=True)
        ws['C16'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['D13'] = 'Available Capacity (MW)'
        ws['D13'].font = Font(size=12, bold=True)
        ws['D13'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('D13:D16')
        
        ws['F13'] = 'Lake Lanao Projected Ave. Outflow'
        ws['F13'].font = Font(size=12, bold=True)
        ws['F13'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('F13:F16')
        
        ws['G13'] = 'Load at 0800H '
        ws['G13'].font = Font(size=12, bold=True)
        ws['G13'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('G13:G16')
        
        ws['H13'] = 'REMARKS'
        ws['H13'].font = Font(size=14, bold=True)
        ws['H13'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('H13:N16')
        
        # Set row heights
        ws.row_dimensions[13].height = 15.75
        ws.row_dimensions[14].height = 15.75
        ws.row_dimensions[15].height = 7.5
        ws.row_dimensions[16].height = 15.75
    
    def _add_plant_data(self, ws):
        """Add plant data starting from row 17 EXACTLY as template"""
        current_row = 17
        
        # Process each plant in order
        for plant_code in ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7']:
            current_row = self._add_plant_section(ws, plant_code, current_row)
        
        # Add TOTAL AGUS row
        ws[f'A{current_row}'] = ' TOTAL AGUS'
        ws[f'A{current_row}'].font = Font(size=14, bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws.row_dimensions[current_row].height = 23.45
        current_row += 1
        
        # Add PULANGI IV
        current_row = self._add_plant_section(ws, 'PULANGI4', current_row)
        
        # Add TOTAL HYDRO row
        ws[f'A{current_row}'] = 'TOTAL HYDRO'
        ws[f'A{current_row}'].font = Font(size=14, bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws.row_dimensions[current_row].height = 23.45
        current_row += 1
        
        return current_row
    
    def _add_plant_section(self, ws, plant_code, start_row):
        """Add a plant section with all its units EXACTLY as template"""
        if plant_code not in self.PLANTS_CONFIG:
            return start_row
        
        config = self.PLANTS_CONFIG[plant_code]
        plant_data = self.data_by_plant.get(plant_code, {})
        
        # Plant header row
        ws[f'A{start_row}'] = config['name']
        ws[f'A{start_row}'].font = Font(size=14, bold=True, italic=True)
        ws[f'A{start_row}'].alignment = Alignment(vertical='center')
        
        # Calculate totals
        total_capacity = sum(u['capacity'] for u in config['units'])
        total_nominated = sum(u['nominated'] for u in config['units'])
        total_available = 0
        total_load = 0
        
        for unit in config['units']:
            if unit['num'] in plant_data:
                unit_data = plant_data[unit['num']]
                total_load += unit_data['generation']
                total_available += unit['capacity']
        
        ws[f'B{start_row}'] = total_capacity
        ws[f'B{start_row}'].font = Font(size=14, bold=True)
        ws[f'B{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'C{start_row}'] = total_nominated
        ws[f'C{start_row}'].font = Font(size=14, bold=True)
        ws[f'C{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'D{start_row}'] = total_available
        ws[f'D{start_row}'].font = Font(size=14, bold=True)
        ws[f'D{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'E{start_row}'] = 0
        ws[f'E{start_row}'].font = Font(size=12, bold=True)
        ws[f'E{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'G{start_row}'] = total_load
        ws[f'G{start_row}'].font = Font(size=14, bold=True)
        ws[f'G{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add remarks for plant
        if plant_code.startswith('AGUS'):
            remarks = 'Lake Lanao Elevation is 701.50 m.a.s.l.'
        elif plant_code == 'PULANGI4':
            remarks = 'Reservoir Elevation is 290.00 m.a.s.l.'
        else:
            remarks = ''
        
        ws[f'H{start_row}'] = remarks
        ws[f'H{start_row}'].font = Font(size=12, bold=True)
        ws[f'H{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.row_dimensions[start_row].height = 23.45
        current_row = start_row + 1
        
        # Add unit rows
        for unit in config['units']:
            ws[f'A{current_row}'] = unit['label']
            ws[f'A{current_row}'].font = Font(size=12)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'B{current_row}'] = unit['capacity']
            ws[f'B{current_row}'].font = Font(size=12)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            ws[f'C{current_row}'] = unit['nominated']
            ws[f'C{current_row}'].font = Font(size=12)
            ws[f'C{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            if unit['num'] in plant_data:
                unit_data = plant_data[unit['num']]
                
                ws[f'D{current_row}'] = unit['capacity']
                ws[f'D{current_row}'].font = Font(size=12)
                ws[f'D{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws[f'G{current_row}'] = unit_data['generation']
                ws[f'G{current_row}'].font = Font(size=12)
                ws[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Add remarks
                unit_remarks = self._get_unit_remarks(unit_data)
                ws[f'H{current_row}'] = unit_remarks
                ws[f'H{current_row}'].font = Font(size=12)
                ws[f'H{current_row}'].alignment = Alignment(vertical='center', wrap_text=True)
            else:
                ws[f'D{current_row}'] = 0
                ws[f'D{current_row}'].font = Font(size=12)
                ws[f'D{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws[f'G{current_row}'] = 0
                ws[f'G{current_row}'].font = Font(size=12)
                ws[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                ws[f'H{current_row}'] = 'No data'
                ws[f'H{current_row}'].font = Font(size=12)
                ws[f'H{current_row}'].alignment = Alignment(vertical='center', wrap_text=True)
            
            ws.row_dimensions[current_row].height = 21.75
            current_row += 1
        
        return current_row
    
    def _get_unit_remarks(self, unit_data):
        """Generate remarks for a unit based on its data"""
        remarks = []
        
        if unit_data['forced_outage'] > 0:
            remarks.append(f"Forced outage: {unit_data['forced_outage']:.1f}h")
        
        if unit_data['scheduled_outage'] > 0:
            remarks.append(f"Scheduled outage: {unit_data['scheduled_outage']:.1f}h")
        
        if unit_data['remarks']:
            remarks.append(unit_data['remarks'])
        
        if not remarks:
            if unit_data['generation'] > 0:
                return 'OPERATIONAL'
            else:
                return 'STANDBY'
        
        return '. '.join(remarks)
    
    def _add_forecasted_load(self, ws, current_row):
        """Add Agus-Pulangi Forecasted Load section (yellow highlighted row) based on actual data"""
        # Calculate forecasted loads from actual uploaded data
        agus_forecast = 0.0
        pulangi_forecast = 0.0
        
        # Sum up loads from AGUS plants
        for plant_code in ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7']:
            if plant_code in self.data_by_plant:
                for unit_num, unit_data in self.data_by_plant[plant_code].items():
                    agus_forecast += unit_data['generation']
        
        # Sum up loads from PULANGI4
        if 'PULANGI4' in self.data_by_plant:
            for unit_num, unit_data in self.data_by_plant['PULANGI4'].items():
                pulangi_forecast += unit_data['generation']
        
        total_forecast = agus_forecast + pulangi_forecast
        
        # Forecasted load row with yellow background
        date_str = self.report_date.strftime('%b %d, %Y')
        forecast_text = f'Agus-Pulangi Forecasted Load @ 6pm, {date_str} : Agus = {agus_forecast:.1f} MW & Pulangui IV = {pulangi_forecast:.1f} MW, Total Load: {total_forecast:.1f} MW'
        
        ws[f'A{current_row}'] = forecast_text
        ws[f'A{current_row}'].font = Font(size=11, bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        ws[f'A{current_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws.merge_cells(f'A{current_row}:N{current_row}')
        
        ws.row_dimensions[current_row].height = 18.0
        current_row += 1
        
        return current_row
    
    def _add_ipp_section(self, ws, current_row):
        """Add IPP (Independent Power Producer) section EXACTLY as template"""
        # MCFPP (STEAG), unit 1
        ws[f'A{current_row}'] = 'MCFPP (STEAG), unit 1'
        ws[f'A{current_row}'].font = Font(size=11)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'B{current_row}'] = 116.0
        ws[f'B{current_row}'].font = Font(size=11)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'C{current_row}'] = 105.0
        ws[f'C{current_row}'].font = Font(size=11)
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'D{current_row}'] = 105.00
        ws[f'D{current_row}'].font = Font(size=11)
        ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'D{current_row}'].number_format = '0.00'
        
        ws[f'E{current_row}'] = 61.50
        ws[f'E{current_row}'].font = Font(size=11)
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'E{current_row}'].number_format = '0.00'
        
        ws[f'H{current_row}'] = 'Normal Operation'
        ws[f'H{current_row}'].font = Font(size=11)
        ws[f'H{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.row_dimensions[current_row].height = 18.0
        current_row += 1
        
        # MCFPP (STEAG), unit 2
        ws[f'A{current_row}'] = 'MCFPP (STEAG), unit 2'
        ws[f'A{current_row}'].font = Font(size=11)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'B{current_row}'] = 116.0
        ws[f'B{current_row}'].font = Font(size=11)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'C{current_row}'] = 105.0
        ws[f'C{current_row}'].font = Font(size=11)
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'D{current_row}'] = 105.00
        ws[f'D{current_row}'].font = Font(size=11)
        ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'D{current_row}'].number_format = '0.00'
        
        ws[f'E{current_row}'] = 62.60
        ws[f'E{current_row}'].font = Font(size=11)
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'E{current_row}'].number_format = '0.00'
        
        ws[f'H{current_row}'] = 'Normal Operation'
        ws[f'H{current_row}'].font = Font(size=11)
        ws[f'H{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.row_dimensions[current_row].height = 18.0
        current_row += 1
        
        # TOTAL IPP
        ws[f'A{current_row}'] = 'TOTAL IPP'
        ws[f'A{current_row}'].font = Font(size=11, bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'A{current_row}'].fill = PatternFill(start_color='CCECFF', end_color='CCECFF', fill_type='solid')
        
        ws[f'B{current_row}'] = 232.00
        ws[f'B{current_row}'].font = Font(size=11, bold=True)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{current_row}'].fill = PatternFill(start_color='CCECFF', end_color='CCECFF', fill_type='solid')
        ws[f'B{current_row}'].number_format = '0.00'
        
        ws[f'C{current_row}'] = 210.00
        ws[f'C{current_row}'].font = Font(size=11, bold=True)
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{current_row}'].fill = PatternFill(start_color='CCECFF', end_color='CCECFF', fill_type='solid')
        ws[f'C{current_row}'].number_format = '0.00'
        
        ws[f'D{current_row}'] = 210.00
        ws[f'D{current_row}'].font = Font(size=11, bold=True)
        ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'D{current_row}'].fill = PatternFill(start_color='CCECFF', end_color='CCECFF', fill_type='solid')
        ws[f'D{current_row}'].number_format = '0.00'
        
        ws[f'E{current_row}'] = 124.10
        ws[f'E{current_row}'].font = Font(size=11, bold=True)
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'E{current_row}'].fill = PatternFill(start_color='CCECFF', end_color='CCECFF', fill_type='solid')
        ws[f'E{current_row}'].number_format = '0.00'
        
        ws.row_dimensions[current_row].height = 18.0
        current_row += 1
        
        # TOTAL NPC-PSALM
        ws[f'A{current_row}'] = 'TOTAL NPC-PSALM'
        ws[f'A{current_row}'].font = Font(size=11, bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'A{current_row}'].fill = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
        
        ws[f'B{current_row}'] = '1,233.10'
        ws[f'B{current_row}'].font = Font(size=11, bold=True)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{current_row}'].fill = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
        
        ws[f'C{current_row}'] = '1,021.3'
        ws[f'C{current_row}'].font = Font(size=11, bold=True)
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{current_row}'].fill = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
        
        ws[f'D{current_row}'] = '857.00'
        ws[f'D{current_row}'].font = Font(size=11, bold=True)
        ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'D{current_row}'].fill = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
        
        ws[f'E{current_row}'] = '759.59'
        ws[f'E{current_row}'].font = Font(size=11, bold=True)
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'E{current_row}'].fill = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
        
        ws.row_dimensions[current_row].height = 18.0
        current_row += 1
        
        # Add empty row
        ws.row_dimensions[current_row].height = 6.0
        current_row += 1
        
        return current_row
    
    def _add_input_workflow_right_side(self, ws):
        """Add INPUT workflow section on the right side of the report"""
        # Starting position - far right columns (AC onwards)
        start_col = 'AC'  # Start after ELEVATION column (AA)
        start_row = 13
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths for the input section
        ws.column_dimensions['AC'].width = 3.0  # Spacer
        ws.column_dimensions['AD'].width = 15.0
        ws.column_dimensions['AE'].width = 12.0
        ws.column_dimensions['AF'].width = 12.0
        ws.column_dimensions['AG'].width = 12.0
        ws.column_dimensions['AH'].width = 12.0
        
        # Header with blue background
        ws[f'AD{start_row}'] = 'INPUT From JMM-DCM/CPN'
        ws[f'AD{start_row}'].font = Font(size=10, bold=True, color='FFFFFF')
        ws[f'AD{start_row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws[f'AD{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'AD{start_row}'].border = thin_border
        ws.merge_cells(f'AD{start_row}:AH{start_row}')
        ws.row_dimensions[start_row].height = 18.0
        start_row += 1
        
        # Date info
        date_str = self.report_date.strftime('%A, %d %B %Y')
        ws[f'AD{start_row}'] = f'today {date_str}'
        ws[f'AD{start_row}'].font = Font(size=9)
        ws[f'AD{start_row}'].fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
        ws[f'AD{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'AD{start_row}'].border = thin_border
        ws.merge_cells(f'AD{start_row}:AH{start_row}')
        ws.row_dimensions[start_row].height = 15.0
        start_row += 1
        
        # INPUT 1 label
        ws[f'AD{start_row}'] = 'INPUT 1'
        ws[f'AD{start_row}'].font = Font(size=10, bold=True, color='FF0000')
        ws.row_dimensions[start_row].height = 15.0
        start_row += 1
        
        # Total load display - calculate from actual data
        total_load_all = 0.0
        for plant_code in self.data_by_plant:
            for unit_num, unit_data in self.data_by_plant[plant_code].items():
                total_load_all += unit_data['generation']
        
        ws[f'AD{start_row}'] = 'Total load @ 0800H.xls'
        ws[f'AD{start_row}'].font = Font(size=9, bold=True)
        ws[f'AD{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'AF{start_row}'] = f'{total_load_all:.2f} MW'
        ws[f'AF{start_row}'].font = Font(size=10, bold=True)
        ws[f'AF{start_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws[f'AF{start_row}'].border = thin_border
        ws[f'AF{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[start_row].height = 18.0
        start_row += 1
        
        # Plant data table header
        start_row += 1
        headers = ['Plant', 'Rated', 'Available', 'Load']
        cols = ['AD', 'AE', 'AF', 'AG']
        
        for col, header in zip(cols, headers):
            ws[f'{col}{start_row}'] = header
            ws[f'{col}{start_row}'].font = Font(size=9, bold=True)
            ws[f'{col}{start_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            ws[f'{col}{start_row}'].border = thin_border
            ws[f'{col}{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[start_row].height = 15.0
        start_row += 1
        
        # Plant data - use actual data from uploaded files
        plants_list = ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7', 'PULANGI4']
        
        for plant_code in plants_list:
            if plant_code in self.PLANTS_CONFIG:
                config = self.PLANTS_CONFIG[plant_code]
                plant_data = self.data_by_plant.get(plant_code, {})
                
                # Calculate totals from actual data
                total_capacity = sum(u['capacity'] for u in config['units'])
                total_available = 0
                total_load = 0
                
                for unit in config['units']:
                    if unit['num'] in plant_data:
                        unit_data = plant_data[unit['num']]
                        total_load += unit_data['generation']
                        total_available += unit['capacity']
                
                ws[f'AD{start_row}'] = plant_code
                ws[f'AD{start_row}'].font = Font(size=8)
                ws[f'AD{start_row}'].border = thin_border
                ws[f'AD{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                ws[f'AE{start_row}'] = total_capacity
                ws[f'AE{start_row}'].font = Font(size=8)
                ws[f'AE{start_row}'].border = thin_border
                ws[f'AE{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws[f'AF{start_row}'] = total_available
                ws[f'AF{start_row}'].font = Font(size=8)
                ws[f'AF{start_row}'].border = thin_border
                ws[f'AF{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws[f'AG{start_row}'] = total_load
                ws[f'AG{start_row}'].font = Font(size=8)
                ws[f'AG{start_row}'].border = thin_border
                ws[f'AG{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws.row_dimensions[start_row].height = 14.0
                start_row += 1
        
        # IPP data
        ipp_data = [
            ('STEAG1', 105.00, 105.00, 0.00),
            ('PMGPP1', 0.00, 0.00, 0.00),
            ('PMGPP2', 0.00, 0.00, 0.00),
        ]
        
        for plant, rated, available, load in ipp_data:
            ws[f'AD{start_row}'] = plant
            ws[f'AD{start_row}'].font = Font(size=8)
            ws[f'AD{start_row}'].border = thin_border
            ws[f'AD{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            ws[f'AE{start_row}'] = rated
            ws[f'AE{start_row}'].font = Font(size=8)
            ws[f'AE{start_row}'].border = thin_border
            ws[f'AE{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            ws[f'AF{start_row}'] = available
            ws[f'AF{start_row}'].font = Font(size=8)
            ws[f'AF{start_row}'].border = thin_border
            ws[f'AF{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            ws[f'AG{start_row}'] = load
            ws[f'AG{start_row}'].font = Font(size=8)
            ws[f'AG{start_row}'].border = thin_border
            ws[f'AG{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            ws.row_dimensions[start_row].height = 14.0
            start_row += 1
        
        # Additional summary tables
        start_row += 2
        
        # Rated/Available/C @0800H Load table
        ws[f'AD{start_row}'] = 'Rated'
        ws[f'AD{start_row}'].font = Font(size=8, bold=True)
        ws[f'AD{start_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws[f'AD{start_row}'].border = thin_border
        
        ws[f'AE{start_row}'] = 'Available'
        ws[f'AE{start_row}'].font = Font(size=8, bold=True)
        ws[f'AE{start_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws[f'AE{start_row}'].border = thin_border
        
        ws[f'AF{start_row}'] = 'C @0800H Load'
        ws[f'AF{start_row}'].font = Font(size=8, bold=True)
        ws[f'AF{start_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws[f'AF{start_row}'].border = thin_border
        start_row += 1
        
        # Plant summary rows - use actual data
        summary_plants = []
        total_rated = 0.0
        total_available = 0.0
        total_load = 0.0
        
        for plant_code in ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7', 'PULANGI4']:
            if plant_code in self.PLANTS_CONFIG:
                config = self.PLANTS_CONFIG[plant_code]
                plant_data = self.data_by_plant.get(plant_code, {})
                
                rated = sum(u['capacity'] for u in config['units'])
                available = 0.0
                load = 0.0
                
                for unit in config['units']:
                    if unit['num'] in plant_data:
                        unit_data = plant_data[unit['num']]
                        load += unit_data['generation']
                        available += unit['capacity']
                
                summary_plants.append((plant_code, rated, available, load))
                total_rated += rated
                total_available += available
                total_load += load
        
        # Add Total row
        summary_plants.append(('Total', total_rated, total_available, total_load))
        
        for plant, rated, available, load in summary_plants:
            ws[f'AD{start_row}'] = plant
            ws[f'AD{start_row}'].font = Font(size=8, bold=('Total' in plant))
            ws[f'AD{start_row}'].border = thin_border
            
            ws[f'AE{start_row}'] = rated
            ws[f'AE{start_row}'].font = Font(size=8)
            ws[f'AE{start_row}'].border = thin_border
            ws[f'AE{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            ws[f'AF{start_row}'] = available
            ws[f'AF{start_row}'].font = Font(size=8)
            ws[f'AF{start_row}'].border = thin_border
            ws[f'AF{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            ws[f'AG{start_row}'] = load
            ws[f'AG{start_row}'].font = Font(size=8)
            ws[f'AG{start_row}'].border = thin_border
            ws[f'AG{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            ws.row_dimensions[start_row].height = 13.0
            start_row += 1
        
        # Capacity factor table
        # Capacity factor table
        start_row += 2
        ws[f'AD{start_row}'] = 'Rated'
        ws[f'AD{start_row}'].font = Font(size=8, bold=True)
        ws[f'AD{start_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws[f'AD{start_row}'].border = thin_border
        
        ws[f'AE{start_row}'] = 'Available'
        ws[f'AE{start_row}'].font = Font(size=8, bold=True)
        ws[f'AE{start_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws[f'AE{start_row}'].border = thin_border
        
        ws[f'AF{start_row}'] = 'C for day % Pmax'
        ws[f'AF{start_row}'].font = Font(size=8, bold=True)
        ws[f'AF{start_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws[f'AF{start_row}'].border = thin_border
        start_row += 1
        
        # Capacity factor data - use actual data
        for plant_code in ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7', 'PULANGI4']:
            if plant_code in self.PLANTS_CONFIG:
                config = self.PLANTS_CONFIG[plant_code]
                plant_data = self.data_by_plant.get(plant_code, {})
                
                rated = sum(u['capacity'] for u in config['units'])
                available = 0.0
                load = 0.0
                
                for unit in config['units']:
                    if unit['num'] in plant_data:
                        unit_data = plant_data[unit['num']]
                        load += unit_data['generation']
                        available += unit['capacity']
                
                # Calculate capacity factor as percentage
                cf = (load / rated * 100) if rated > 0 else 0.0
                
                ws[f'AD{start_row}'] = plant_code
                ws[f'AD{start_row}'].font = Font(size=8)
                ws[f'AD{start_row}'].border = thin_border
                
                ws[f'AE{start_row}'] = rated
                ws[f'AE{start_row}'].font = Font(size=8)
                ws[f'AE{start_row}'].border = thin_border
                ws[f'AE{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws[f'AF{start_row}'] = available
                ws[f'AF{start_row}'].font = Font(size=8)
                ws[f'AF{start_row}'].border = thin_border
                ws[f'AF{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws[f'AG{start_row}'] = cf
                ws[f'AG{start_row}'].font = Font(size=8)
                ws[f'AG{start_row}'].border = thin_border
                ws[f'AG{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws.row_dimensions[start_row].height = 13.0
                start_row += 1
            
            ws[f'AG{start_row}'] = cf
            ws[f'AG{start_row}'].font = Font(size=8)
            ws[f'AG{start_row}'].border = thin_border
            ws[f'AG{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            ws.row_dimensions[start_row].height = 13.0
            start_row += 1
    
    def _add_notes_section(self, ws, current_row):
        """Add charts and notes section EXACTLY as template"""
        # Store the starting row for charts
        charts_start_row = current_row
        
        # Add charts section headers
        ws[f'B{current_row}'] = 'NPC-PSALM Capacity Mix'
        ws[f'B{current_row}'].font = Font(size=11, bold=True, color='FFFFFF')
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{current_row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws.merge_cells(f'B{current_row}:D{current_row}')
        
        ws[f'H{current_row}'] = 'MinGen Forecasted Load Share (MW), @6pm Today'
        ws[f'H{current_row}'].font = Font(size=11, bold=True, color='FFFFFF')
        ws[f'H{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'H{current_row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws.merge_cells(f'H{current_row}:N{current_row}')
        
        ws.row_dimensions[current_row].height = 18.0
        current_row += 1
        
        # Create PIE CHART - NPC-PSALM Capacity Mix
        pie_chart = PieChart()
        pie_chart.title = None  # No title, we have the header above
        pie_chart.width = 10
        pie_chart.height = 10
        
        # Calculate actual hydro and thermal capacity from uploaded data
        hydro_capacity = 0.0
        for plant_code in ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7', 'PULANGI4']:
            if plant_code in self.PLANTS_CONFIG:
                config = self.PLANTS_CONFIG[plant_code]
                hydro_capacity += sum(u['capacity'] for u in config['units'])
        
        thermal_capacity = 210.00  # IPP capacity (STEAG) - this is fixed
        
        # Add data for pie chart (Hydro vs Coal Fired Thermal)
        # Create hidden data cells for the chart
        data_row = current_row
        ws[f'A{data_row}'] = 'Hydro'
        ws[f'B{data_row}'] = hydro_capacity
        ws[f'A{data_row}'].font = Font(size=1, color='FFFFFF')  # Hidden
        ws[f'B{data_row}'].font = Font(size=1, color='FFFFFF')  # Hidden
        
        data_row += 1
        ws[f'A{data_row}'] = 'Coal Fired Thermal'
        ws[f'B{data_row}'] = thermal_capacity
        ws[f'A{data_row}'].font = Font(size=1, color='FFFFFF')  # Hidden
        ws[f'B{data_row}'].font = Font(size=1, color='FFFFFF')  # Hidden
        
        # Set chart data
        labels = Reference(ws, min_col=1, min_row=current_row, max_row=current_row+1)
        data = Reference(ws, min_col=2, min_row=current_row, max_row=current_row+1)
        pie_chart.add_data(data)
        pie_chart.set_categories(labels)
        
        # Style the pie chart
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showVal = True
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showCatName = False
        
        # Add pie chart to worksheet
        ws.add_chart(pie_chart, f'B{current_row + 2}')
        
        # Create BAR CHART - MinGen Forecasted Load Share
        bar_chart = BarChart()
        bar_chart.type = "col"  # Column chart
        bar_chart.title = None
        bar_chart.width = 15
        bar_chart.height = 10
        bar_chart.y_axis.title = None
        bar_chart.x_axis.title = None
        
        # Add data for bar chart (Plant forecasted loads) - use actual data
        bar_data_start = current_row + 3
        plants_data = []
        
        for plant_code in ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7', 'PULANGI4']:
            if plant_code in self.PLANTS_CONFIG:
                config = self.PLANTS_CONFIG[plant_code]
                plant_data = self.data_by_plant.get(plant_code, {})
                
                load = 0.0
                for unit in config['units']:
                    if unit['num'] in plant_data:
                        unit_data = plant_data[unit['num']]
                        load += unit_data['generation']
                
                plant_name = config['name']
                plants_data.append((plant_name, load))
        
        for idx, (plant, load) in enumerate(plants_data):
            row = bar_data_start + idx
            ws[f'F{row}'] = plant
            ws[f'G{row}'] = load
            ws[f'F{row}'].font = Font(size=1, color='FFFFFF')  # Hidden
            ws[f'G{row}'].font = Font(size=1, color='FFFFFF')  # Hidden
        
        # Set bar chart data
        bar_labels = Reference(ws, min_col=6, min_row=bar_data_start, max_row=bar_data_start + len(plants_data) - 1)
        bar_data = Reference(ws, min_col=7, min_row=bar_data_start, max_row=bar_data_start + len(plants_data) - 1)
        bar_chart.add_data(bar_data)
        bar_chart.set_categories(bar_labels)
        
        # Style the bar chart
        bar_chart.dataLabels = DataLabelList()
        bar_chart.dataLabels.showVal = True
        
        # Add bar chart to worksheet
        ws.add_chart(bar_chart, f'H{current_row + 2}')
        
        # Add space for charts (rows for visual representation)
        for i in range(12):
            ws.row_dimensions[current_row].height = 15.0
            current_row += 1
        
        # Add legend row
        ws[f'B{current_row}'] = '■ Hydro     ■ Coal Fired Thermal'
        ws[f'B{current_row}'].font = Font(size=11, bold=True)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{current_row}'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        ws.merge_cells(f'B{current_row}:D{current_row}')
        
        ws.row_dimensions[current_row].height = 18.0
        current_row += 1
        
        # Add empty row
        ws.row_dimensions[current_row].height = 6.0
        current_row += 1
        
        # Notes header
        ws[f'A{current_row}'] = 'Note:'
        ws[f'A{current_row}'].font = Font(size=10, bold=True, italic=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top')
        ws.row_dimensions[current_row].height = 15.0
        current_row += 1
        
        # Notes content - EXACTLY as template
        notes = [
            "1. Dependable Capacity (DC) is the maximum capacity, modified for ambient limitations for a specific period of time, such as month or a season.",
            "2. Available Capacity (AC) is the dependable capacity, modified for equipment limitations for any time.",
            "3. The usual occurrence of Peak is at 1200H.",
            "4. AGUS 5 HEP gate no. 2 clogged at 0.10m for Nawlach Pulp Inc. (NPI) plant water use."
        ]
        
        for note in notes:
            ws[f'A{current_row}'] = note
            ws[f'A{current_row}'].font = Font(size=9)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.merge_cells(f'A{current_row}:N{current_row}')
            ws.row_dimensions[current_row].height = 13.5
            current_row += 1
        
        # Add additional detailed notes section
        current_row = self._add_detailed_notes_section(ws, current_row)
        
        return current_row
    
    def _add_detailed_notes_section(self, ws, start_row):
        """Add detailed notes section with MCFPP and MGPP information"""
        # Add spacing
        start_row += 1
        
        # MCFPP (STEAG) Section
        ws[f'A{start_row}'] = 'MCFPP (STEAG)'
        ws[f'A{start_row}'].font = Font(size=10, bold=True)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'B{start_row}'] = 'TYPE OF CONTI IPP OPERATOR END OF COOPERATION PER CONTRACTED INDICATIVE BID'
        ws[f'B{start_row}'].font = Font(size=8)
        ws[f'B{start_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells(f'B{start_row}:E{start_row}')
        ws.row_dimensions[start_row].height = 15.0
        start_row += 1
        
        ws[f'A{start_row}'] = 'MGPP'
        ws[f'A{start_row}'].font = Font(size=9)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'B{start_row}'] = 'BOT/IPPA'
        ws[f'B{start_row}'].font = Font(size=8)
        ws[f'B{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'C{start_row}'] = 'STEAG State Pov'
        ws[f'C{start_row}'].font = Font(size=8)
        ws[f'C{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'D{start_row}'] = '16-Nov-31'
        ws[f'D{start_row}'].font = Font(size=8)
        ws[f'D{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'E{start_row}'] = "200.00 2017 (Subject to DOE's policy direction)"
        ws[f'E{start_row}'].font = Font(size=8)
        ws[f'E{start_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells(f'E{start_row}:F{start_row}')
        ws.row_dimensions[start_row].height = 15.0
        start_row += 1
        
        ws[f'B{start_row}'] = 'IPPA AA'
        ws[f'B{start_row}'].font = Font(size=8)
        ws[f'B{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'C{start_row}'] = 'FDC'
        ws[f'C{start_row}'].font = Font(size=8)
        ws[f'C{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'D{start_row}'] = '11-Dec-17'
        ws[f'D{start_row}'].font = Font(size=8)
        ws[f'D{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        ws.row_dimensions[start_row].height = 15.0
        start_row += 1
        
        # Additional Notes
        start_row += 1
        ws[f'A{start_row}'] = 'Note:'
        ws[f'A{start_row}'].font = Font(size=9, bold=True, italic=True)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        ws.row_dimensions[start_row].height = 13.5
        start_row += 1
        
        detailed_notes = [
            "1. The Available Capacity in this report includes equipment limitation and water outflow consideration based on the 2020 Lake Lanao Operating Guide Curve.",
            "2. AGUS 6 HEP, units 1 & 2 up-rated from 25 MW to 34.5 MW each and rated speed increased from 200 to 225 rpm effective February 1, 2020.",
            "3. AGUS 2 HEP is limited to 120 MW total load to prevent risk of flooding at lakeshore areas and Balo-i plains as per Environmental Compliance Certificate dated January 14, 1992.",
            "4. AGUS 5 HEP gate no. 2 clogged at 0.10m for Nawlach Pulp Inc. (NPI) plant water use.",
            "5. The usual occurrence of Peak is at 1800H.",
            "6. Forecast inflow of Lake Lanao is stable and operating at Normal Stage.",
            "7. Agus 6 HEP: 18cms of water spilled due to partially opened spillway gate no. 1."
        ]
        
        for note in detailed_notes:
            ws[f'A{start_row}'] = note
            ws[f'A{start_row}'].font = Font(size=8)
            ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.merge_cells(f'A{start_row}:N{start_row}')
            ws.row_dimensions[start_row].height = 12.0
            start_row += 1
        
        return start_row
    
    def _add_chart_and_agus2_note(self, ws, start_row):
        """Add bar chart and Agus 2 HEP limitation note"""
        # Add spacing
        start_row += 2
        
        # Agus 2 HEP limitation note
        note_text = 'Agus 2 HEP is limited to 40 MW/per unit due to water constraints as per Environmental Compliance Certificate "that Agus 2 shall not be operated at full capacity..." This is to prevent risk of flooding at lakeshore areas and Baloi plains.'
        ws[f'A{start_row}'] = note_text
        ws[f'A{start_row}'].font = Font(size=9)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells(f'A{start_row}:H{start_row}')
        ws.row_dimensions[start_row].height = 25.0
        start_row += 1
        
        # Add spacing for chart
        start_row += 2
        
        # Create bar chart for plant comparison
        chart = BarChart()
        chart.type = "col"
        chart.title = "Plant Capacity Comparison"
        chart.width = 15
        chart.height = 10
        chart.y_axis.title = "MW"
        chart.x_axis.title = None
        
        # Chart data (hidden cells) - use actual data
        chart_data_start = start_row
        plants_chart_data = []
        
        for plant_code in ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7', 'PULANGI4']:
            if plant_code in self.PLANTS_CONFIG:
                config = self.PLANTS_CONFIG[plant_code]
                plant_data = self.data_by_plant.get(plant_code, {})
                
                rated = sum(u['capacity'] for u in config['units'])
                available = 0.0
                load = 0.0
                
                for unit in config['units']:
                    if unit['num'] in plant_data:
                        unit_data = plant_data[unit['num']]
                        load += unit_data['generation']
                        available += unit['capacity']
                
                plant_name = config['name']
                plants_chart_data.append((plant_name, rated, available, load))
        
        # Add hidden data for chart
        for idx, (plant, rated, dependable, load) in enumerate(plants_chart_data):
            row = chart_data_start + idx
            ws[f'A{row}'] = plant
            ws[f'B{row}'] = rated
            ws[f'C{row}'] = dependable
            ws[f'D{row}'] = load
            
            # Make text very small and light to hide it
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = Font(size=1, color='FFFFFF')
        
        # Set chart data references
        labels = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_start + len(plants_chart_data) - 1)
        data = Reference(ws, min_col=2, min_row=chart_data_start - 1, max_col=4, max_row=chart_data_start + len(plants_chart_data) - 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        # Style the chart
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = False
        
        # Add legend
        chart.legend.position = 'b'  # Bottom
        
        # Add chart to worksheet
        ws.add_chart(chart, f'A{start_row + 1}')
        
        # Reserve space for chart (approximately 15 rows)
        for i in range(15):
            ws.row_dimensions[start_row].height = 15.0
            start_row += 1
        
        return start_row
    
    def _add_footer(self, ws, start_row):
        """Add footer section with signatures and additional personnel EXACTLY as template"""
        # Add spacing
        start_row += 2
        
        # First row of signatures - Prepared by, Checked by, Approved by
        ws[f'A{start_row}'] = 'Prepared by:'
        ws[f'A{start_row}'].font = Font(size=10)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'D{start_row}'] = 'Checked and Reviewed by:'
        ws[f'D{start_row}'].font = Font(size=10)
        ws[f'D{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'G{start_row}'] = 'Checked and Reviewed by:'
        ws[f'G{start_row}'].font = Font(size=10)
        ws[f'G{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'K{start_row}'] = 'Approved by:'
        ws[f'K{start_row}'].font = Font(size=10)
        ws[f'K{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws.row_dimensions[start_row].height = 13.5
        start_row += 1
        
        # Empty rows for signature space
        for i in range(3):
            ws.row_dimensions[start_row].height = 13.5
            start_row += 1
        
        # First row of names
        ws[f'A{start_row}'] = 'O.M. LAVA'
        ws[f'A{start_row}'].font = Font(size=11, bold=True)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'D{start_row}'] = 'JMM MATA'
        ws[f'D{start_row}'].font = Font(size=11, bold=True)
        ws[f'D{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'G{start_row}'] = 'EL ADIONG'
        ws[f'G{start_row}'].font = Font(size=11, bold=True)
        ws[f'G{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'K{start_row}'] = 'C.C. AMIGABLE JR.'
        ws[f'K{start_row}'].font = Font(size=11, bold=True)
        ws[f'K{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws.row_dimensions[start_row].height = 13.5
        start_row += 1
        
        # First row of titles
        ws[f'A{start_row}'] = 'Prin. Engr. A, GPD'
        ws[f'A{start_row}'].font = Font(size=10)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'D{start_row}'] = 'Manager, GPD'
        ws[f'D{start_row}'].font = Font(size=10)
        ws[f'D{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'G{start_row}'] = 'Acting Manager, GPD'
        ws[f'G{start_row}'].font = Font(size=10)
        ws[f'G{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'K{start_row}'] = 'Dept. Manager, GPD'
        ws[f'K{start_row}'].font = Font(size=10)
        ws[f'K{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws.row_dimensions[start_row].height = 13.5
        start_row += 1
        
        # Add spacing between signature rows
        ws.row_dimensions[start_row].height = 13.5
        start_row += 1
        
        # Second row of signatures
        ws[f'A{start_row}'] = 'Prepared by:'
        ws[f'A{start_row}'].font = Font(size=10)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'D{start_row}'] = 'Checked and Reviewed by:'
        ws[f'D{start_row}'].font = Font(size=10)
        ws[f'D{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'G{start_row}'] = 'Checked and Reviewed by:'
        ws[f'G{start_row}'].font = Font(size=10)
        ws[f'G{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'K{start_row}'] = 'Approved by:'
        ws[f'K{start_row}'].font = Font(size=10)
        ws[f'K{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws.row_dimensions[start_row].height = 13.5
        start_row += 1
        
        # Empty rows for signature space
        for i in range(3):
            ws.row_dimensions[start_row].height = 13.5
            start_row += 1
        
        # Second row of names
        ws[f'A{start_row}'] = 'D.R.B. CAIRO'
        ws[f'A{start_row}'].font = Font(size=11, bold=True)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'D{start_row}'] = 'JMM MATA'
        ws[f'D{start_row}'].font = Font(size=11, bold=True)
        ws[f'D{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'G{start_row}'] = 'EL ADIONG'
        ws[f'G{start_row}'].font = Font(size=11, bold=True)
        ws[f'G{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'K{start_row}'] = 'DB ESMADE JR.'
        ws[f'K{start_row}'].font = Font(size=11, bold=True)
        ws[f'K{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws.row_dimensions[start_row].height = 13.5
        start_row += 1
        
        # Second row of titles
        ws[f'A{start_row}'] = 'Prin. Engr. B, GPD'
        ws[f'A{start_row}'].font = Font(size=10)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'D{start_row}'] = 'Manager, GPD'
        ws[f'D{start_row}'].font = Font(size=10)
        ws[f'D{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'G{start_row}'] = 'OIC-Dept Manager, GPD'
        ws[f'G{start_row}'].font = Font(size=10)
        ws[f'G{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws[f'K{start_row}'] = 'Acting Dept. Manager, GPD'
        ws[f'K{start_row}'].font = Font(size=10)
        ws[f'K{start_row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        ws.row_dimensions[start_row].height = 13.5
        start_row += 1
        
        return start_row
    
    def _add_chart_and_agus2_note(self, ws, start_row):
        """Add bar chart and Agus 2 HEP limitation note at the very bottom left"""
        # Add spacing
        start_row += 3
        
        # Agus 2 HEP limitation note
        note_text = 'Agus 2 HEP is limited to 40 MW/per unit due to water constraints as per Environmental Compliance Certificate "that Agus 2 shall not be operated at full capacity..." This is to prevent risk of flooding at lakeshore areas and Baloi plains.'
        ws[f'A{start_row}'] = note_text
        ws[f'A{start_row}'].font = Font(size=9)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells(f'A{start_row}:H{start_row}')
        ws.row_dimensions[start_row].height = 25.0
        start_row += 1
        
        # Add spacing for chart
        start_row += 2
        
        # Create bar chart for plant comparison
        chart = BarChart()
        chart.type = "col"
        chart.title = "Plant Capacity Comparison"
        chart.width = 15
        chart.height = 10
        chart.y_axis.title = "MW"
        chart.x_axis.title = None
        
        # Chart data (hidden cells)
        chart_data_start = start_row
        plants_chart_data = [
            ('AGUS 1', 80.0, 80.0, 0.0),
            ('AGUS 2', 180.0, 180.0, 120.0),
            ('AGUS 4', 158.1, 158.1, 105.4),
            ('AGUS 5', 55.0, 55.0, 55.0),
            ('AGUS 6', 219.0, 219.0, 165.0),
            ('AGUS 7', 54.0, 54.0, 54.0),
            ('PULANGI IV', 255.0, 255.0, 215.0),
        ]
        
        # Add hidden data for chart
        for idx, (plant, rated, dependable, load) in enumerate(plants_chart_data):
            row = chart_data_start + idx
            ws[f'A{row}'] = plant
            ws[f'B{row}'] = rated
            ws[f'C{row}'] = dependable
            ws[f'D{row}'] = load
            
            # Make text very small and light to hide it
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = Font(size=1, color='FFFFFF')
        
        # Set chart data references
        labels = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_start + len(plants_chart_data) - 1)
        data = Reference(ws, min_col=2, min_row=chart_data_start - 1, max_col=4, max_row=chart_data_start + len(plants_chart_data) - 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        # Style the chart
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = False
        
        # Add legend
        chart.legend.position = 'b'  # Bottom
        
        # Add chart to worksheet
        ws.add_chart(chart, f'A{start_row + 1}')
        
        # Reserve space for chart (approximately 15 rows)
        for i in range(15):
            ws.row_dimensions[start_row].height = 15.0
            start_row += 1
        
        # Add second chart - Rated vs Dependable Cap vs Yesterday's Peak
        start_row += 2
        self._add_comparison_chart(ws, start_row)
        
        return start_row
    
    def _add_comparison_chart(self, ws, start_row):
        """Add comparison chart showing Rated, Dependable Cap, and Yesterday's Peak"""
        # Create combination chart (bar + line)
        chart2 = BarChart()
        chart2.type = "col"
        chart2.title = "Plant Capacity vs Yesterday's Peak"
        chart2.width = 20
        chart2.height = 10
        chart2.y_axis.title = "MW"
        chart2.x_axis.title = None
        
        # Chart data
        chart_data_start = start_row
        comparison_data = [
            ('AGUS 1', 80.0, 80.0, 0.0),
            ('AGUS 2', 180.0, 180.0, 0.0),
            ('AGUS 4', 158.1, 158.1, 0.0),
            ('AGUS 5', 55.0, 55.0, 0.0),
            ('AGUS 6', 219.0, 219.0, 0.0),
            ('AGUS 7', 54.0, 54.0, 0.0),
            ('Pulangi4', 255.0, 255.0, 0.0),
            ('STEAG 1', 116.0, 105.0, 0.0),
            ('STEAG 2', 116.0, 105.0, 0.0),
            ('10-Jan-25', 0.0, 0.0, 0.0),
            ('11-Jan-25', 0.0, 0.0, 0.0),
        ]
        
        # Add hidden data for chart
        for idx, (label, rated, dependable, peak) in enumerate(comparison_data):
            row = chart_data_start + idx
            ws[f'A{row}'] = label
            ws[f'B{row}'] = rated
            ws[f'C{row}'] = dependable
            ws[f'D{row}'] = peak
            
            # Make text very small and light to hide it
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = Font(size=1, color='FFFFFF')
        
        # Set chart data references
        labels = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_start + len(comparison_data) - 1)
        data = Reference(ws, min_col=2, min_row=chart_data_start - 1, max_col=4, max_row=chart_data_start + len(comparison_data) - 1)
        
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(labels)
        
        # Style the chart
        chart2.dataLabels = DataLabelList()
        chart2.dataLabels.showVal = False
        
        # Add legend at top right
        chart2.legend.position = 't'  # Top
        
        # Add chart to worksheet
        ws.add_chart(chart2, f'A{start_row + 1}')
        
        return start_row
    
    def _get_file_path(self):
        """Generate file path for PSR export"""
        export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        
        date_str = self.report_date.strftime('%Y%m%d')
        filename = f'PSR_REPORT_{date_str}.xlsx'
        
        return os.path.join(export_dir, filename)
    
    def _add_right_side_sections(self, ws):
        """Add right side sections: Storage, Inflow/Outflow, Generation Data, Capacity Factor, and Gate/Elevation Info"""
        # Define styles
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        header_font = Font(bold=True, color="FFFFFF", size=10)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=9)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. PRIMARY STORAGE OF HYDRO HEPs (Starting at row 13)
        start_row = 13
        ws.merge_cells(f'P{start_row}:S{start_row}')
        ws[f'P{start_row}'] = 'PRIMARY STORAGE OF HYDRO HEPs'
        ws[f'P{start_row}'].font = header_font
        ws[f'P{start_row}'].fill = header_fill
        ws[f'P{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'P{start_row}'].border = thin_border
        
        # Storage table headers
        start_row += 1
        ws[f'P{start_row}'] = 'Lake/Dam'
        ws[f'P{start_row}'].fill = yellow_fill
        ws[f'P{start_row}'].font = bold_font
        ws[f'P{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'P{start_row}'].border = thin_border
        
        ws[f'Q{start_row}'] = 'Level (m)'
        ws[f'Q{start_row}'].fill = yellow_fill
        ws[f'Q{start_row}'].font = bold_font
        ws[f'Q{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'Q{start_row}'].border = thin_border
        
        ws.merge_cells(f'R{start_row}:S{start_row}')
        ws[f'R{start_row}'] = 'Remarks'
        ws[f'R{start_row}'].fill = yellow_fill
        ws[f'R{start_row}'].font = bold_font
        ws[f'R{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'R{start_row}'].border = thin_border
        
        # Storage data
        storage_data = [
            ['Lake Lanao', '', ''],
            ['Agus 2 Forebay', '', ''],
            ['Agus 4 Forebay', '', ''],
            ['Agus 5 Forebay', '', ''],
            ['Agus 6 Forebay', '', ''],
            ['Agus 7 Forebay', '', ''],
            ['Pulangi IV Reservoir', '', ''],
        ]
        
        for row_data in storage_data:
            start_row += 1
            ws[f'P{start_row}'] = row_data[0]
            ws[f'P{start_row}'].border = thin_border
            ws[f'P{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            ws[f'Q{start_row}'] = row_data[1]
            ws[f'Q{start_row}'].border = thin_border
            ws[f'Q{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells(f'R{start_row}:S{start_row}')
            ws[f'R{start_row}'] = row_data[2]
            ws[f'R{start_row}'].border = thin_border
            ws[f'R{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 2. HYDRO INFLOW/OUTFLOW (Starting 2 rows after storage)
        start_row += 2
        ws.merge_cells(f'P{start_row}:S{start_row}')
        ws[f'P{start_row}'] = 'HYDRO INFLOW/OUTFLOW (cms)'
        ws[f'P{start_row}'].font = header_font
        ws[f'P{start_row}'].fill = header_fill
        ws[f'P{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'P{start_row}'].border = thin_border
        
        # Inflow/Outflow headers
        start_row += 1
        ws[f'P{start_row}'] = 'Plant'
        ws[f'P{start_row}'].fill = yellow_fill
        ws[f'P{start_row}'].font = bold_font
        ws[f'P{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'P{start_row}'].border = thin_border
        
        ws[f'Q{start_row}'] = 'Inflow'
        ws[f'Q{start_row}'].fill = yellow_fill
        ws[f'Q{start_row}'].font = bold_font
        ws[f'Q{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'Q{start_row}'].border = thin_border
        
        ws[f'R{start_row}'] = 'Outflow'
        ws[f'R{start_row}'].fill = yellow_fill
        ws[f'R{start_row}'].font = bold_font
        ws[f'R{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'R{start_row}'].border = thin_border
        
        ws[f'S{start_row}'] = 'Remarks'
        ws[f'S{start_row}'].fill = yellow_fill
        ws[f'S{start_row}'].font = bold_font
        ws[f'S{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'S{start_row}'].border = thin_border
        
        # Inflow/Outflow data
        flow_data = [
            ['Lake Lanao', '0.00', '0.00', ''],
            ['Agus 1', '0.00', '0.00', ''],
            ['Agus 2', '0.00', '0.00', ''],
            ['Agus 4', '0.00', '0.00', ''],
            ['Agus 5', '0.00', '0.00', ''],
            ['Agus 6', '0.00', '0.00', ''],
            ['Agus 7', '0.00', '0.00', ''],
            ['Pulangi IV', '0.00', '0.00', ''],
        ]
        
        for row_data in flow_data:
            start_row += 1
            ws[f'P{start_row}'] = row_data[0]
            ws[f'P{start_row}'].border = thin_border
            ws[f'P{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            ws[f'Q{start_row}'] = row_data[1]
            ws[f'Q{start_row}'].border = thin_border
            ws[f'Q{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'R{start_row}'] = row_data[2]
            ws[f'R{start_row}'].border = thin_border
            ws[f'R{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'S{start_row}'] = row_data[3]
            ws[f'S{start_row}'].border = thin_border
            ws[f'S{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 3. GENERATION DATA (MWh) (Starting 2 rows after inflow/outflow)
        start_row += 2
        ws.merge_cells(f'P{start_row}:S{start_row}')
        ws[f'P{start_row}'] = 'GENERATION DATA (MWh)'
        ws[f'P{start_row}'].font = header_font
        ws[f'P{start_row}'].fill = header_fill
        ws[f'P{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'P{start_row}'].border = thin_border
        
        # Generation headers
        start_row += 1
        ws[f'P{start_row}'] = 'Plant'
        ws[f'P{start_row}'].fill = yellow_fill
        ws[f'P{start_row}'].font = bold_font
        ws[f'P{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'P{start_row}'].border = thin_border
        
        ws[f'Q{start_row}'] = 'Today'
        ws[f'Q{start_row}'].fill = yellow_fill
        ws[f'Q{start_row}'].font = bold_font
        ws[f'Q{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'Q{start_row}'].border = thin_border
        
        ws[f'R{start_row}'] = 'MTD'
        ws[f'R{start_row}'].fill = yellow_fill
        ws[f'R{start_row}'].font = bold_font
        ws[f'R{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'R{start_row}'].border = thin_border
        
        ws[f'S{start_row}'] = 'YTD'
        ws[f'S{start_row}'].fill = yellow_fill
        ws[f'S{start_row}'].font = bold_font
        ws[f'S{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'S{start_row}'].border = thin_border
        
        # Calculate generation data from actual data
        gen_data = self._calculate_generation_data()
        
        for row_data in gen_data:
            start_row += 1
            ws[f'P{start_row}'] = row_data[0]
            ws[f'P{start_row}'].border = thin_border
            ws[f'P{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
            if 'Total' in row_data[0]:
                ws[f'P{start_row}'].font = bold_font
                ws[f'P{start_row}'].fill = light_gray_fill
            
            ws[f'Q{start_row}'] = row_data[1]
            ws[f'Q{start_row}'].border = thin_border
            ws[f'Q{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Total' in row_data[0]:
                ws[f'Q{start_row}'].font = bold_font
                ws[f'Q{start_row}'].fill = light_gray_fill
            
            ws[f'R{start_row}'] = row_data[2]
            ws[f'R{start_row}'].border = thin_border
            ws[f'R{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Total' in row_data[0]:
                ws[f'R{start_row}'].font = bold_font
                ws[f'R{start_row}'].fill = light_gray_fill
            
            ws[f'S{start_row}'] = row_data[3]
            ws[f'S{start_row}'].border = thin_border
            ws[f'S{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Total' in row_data[0]:
                ws[f'S{start_row}'].font = bold_font
                ws[f'S{start_row}'].fill = light_gray_fill
        
        # 4. CAPACITY FACTOR (%) (Starting 2 rows after generation)
        start_row += 2
        ws.merge_cells(f'P{start_row}:S{start_row}')
        ws[f'P{start_row}'] = 'CAPACITY FACTOR (%)'
        ws[f'P{start_row}'].font = header_font
        ws[f'P{start_row}'].fill = header_fill
        ws[f'P{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'P{start_row}'].border = thin_border
        
        # Capacity Factor headers
        start_row += 1
        ws[f'P{start_row}'] = 'Plant'
        ws[f'P{start_row}'].fill = yellow_fill
        ws[f'P{start_row}'].font = bold_font
        ws[f'P{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'P{start_row}'].border = thin_border
        
        ws[f'Q{start_row}'] = 'Today'
        ws[f'Q{start_row}'].fill = yellow_fill
        ws[f'Q{start_row}'].font = bold_font
        ws[f'Q{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'Q{start_row}'].border = thin_border
        
        ws[f'R{start_row}'] = 'MTD'
        ws[f'R{start_row}'].fill = yellow_fill
        ws[f'R{start_row}'].font = bold_font
        ws[f'R{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'R{start_row}'].border = thin_border
        
        ws[f'S{start_row}'] = 'YTD'
        ws[f'S{start_row}'].fill = yellow_fill
        ws[f'S{start_row}'].font = bold_font
        ws[f'S{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'S{start_row}'].border = thin_border
        
        # Calculate capacity factor data
        cf_data = self._calculate_capacity_factor()
        
        for row_data in cf_data:
            start_row += 1
            ws[f'P{start_row}'] = row_data[0]
            ws[f'P{start_row}'].border = thin_border
            ws[f'P{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
            if 'Average' in row_data[0]:
                ws[f'P{start_row}'].font = bold_font
                ws[f'P{start_row}'].fill = light_gray_fill
            
            ws[f'Q{start_row}'] = row_data[1]
            ws[f'Q{start_row}'].border = thin_border
            ws[f'Q{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Average' in row_data[0]:
                ws[f'Q{start_row}'].font = bold_font
                ws[f'Q{start_row}'].fill = light_gray_fill
            
            ws[f'R{start_row}'] = row_data[2]
            ws[f'R{start_row}'].border = thin_border
            ws[f'R{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Average' in row_data[0]:
                ws[f'R{start_row}'].font = bold_font
                ws[f'R{start_row}'].fill = light_gray_fill
            
            ws[f'S{start_row}'] = row_data[3]
            ws[f'S{start_row}'].border = thin_border
            ws[f'S{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            if 'Average' in row_data[0]:
                ws[f'S{start_row}'].font = bold_font
                ws[f'S{start_row}'].fill = light_gray_fill
        
        # 5. GATE AND ELEVATION SECTION (Far right side, starting at row 17)
        self._add_gate_elevation_section(ws)
        
        # 5b. OPERATIONAL REFERENCE SECTION (Adjacent to gate/elevation, starting at column AB)
        self._add_operational_reference_section(ws)
        
        # 6. INPUT WORKFLOW SECTION (Far right side, below other sections)
        self._add_input_workflow_right_side(ws)
    
    def _calculate_generation_data(self):
        """Calculate generation data for today, MTD, YTD from database"""
        from django.db.models import Sum
        from reports.models import GenerationReport, Plant
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        # Get the report date
        report_date = self.report_date
        
        # Calculate date ranges
        today_start = report_date
        today_end = report_date
        
        # Month-to-date: from 1st of current month to report date
        mtd_start = report_date.replace(day=1)
        mtd_end = report_date
        
        # Year-to-date: from Jan 1 to report date
        ytd_start = report_date.replace(month=1, day=1)
        ytd_end = report_date
        
        gen_data = []
        total_today = 0
        total_mtd = 0
        total_ytd = 0
        
        # Query generation data for each plant
        plant_codes = ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7', 'PULANGI4']
        
        for plant_code in plant_codes:
            try:
                plant = Plant.objects.get(code=plant_code)
                
                # Today's generation (convert kWh to MWh)
                today_gen = GenerationReport.objects.filter(
                    plant=plant,
                    report_date=today_start
                ).aggregate(total=Sum('generation_kwh'))['total'] or 0
                today_mwh = float(today_gen) / 1000
                
                # MTD generation
                mtd_gen = GenerationReport.objects.filter(
                    plant=plant,
                    report_date__gte=mtd_start,
                    report_date__lte=mtd_end
                ).aggregate(total=Sum('generation_kwh'))['total'] or 0
                mtd_mwh = float(mtd_gen) / 1000
                
                # YTD generation
                ytd_gen = GenerationReport.objects.filter(
                    plant=plant,
                    report_date__gte=ytd_start,
                    report_date__lte=ytd_end
                ).aggregate(total=Sum('generation_kwh'))['total'] or 0
                ytd_mwh = float(ytd_gen) / 1000
                
                # Format with commas
                gen_data.append([
                    plant_code,
                    f'{today_mwh:,.0f}',
                    f'{mtd_mwh:,.0f}',
                    f'{ytd_mwh:,.0f}'
                ])
                
                total_today += today_mwh
                total_mtd += mtd_mwh
                total_ytd += ytd_mwh
                
            except Plant.DoesNotExist:
                # If plant doesn't exist, use zeros
                gen_data.append([plant_code, '0', '0', '0'])
        
        # Add total row
        gen_data.append([
            'Total NPC',
            f'{total_today:,.0f}',
            f'{total_mtd:,.0f}',
            f'{total_ytd:,.0f}'
        ])
        
        return gen_data
    
    def _calculate_capacity_factor(self):
        """Calculate capacity factor for today, MTD, YTD from database"""
        from django.db.models import Avg
        from reports.models import GenerationReport, Plant
        from datetime import datetime
        
        # Get the report date
        report_date = self.report_date
        
        # Calculate date ranges
        today_start = report_date
        today_end = report_date
        
        # Month-to-date
        mtd_start = report_date.replace(day=1)
        mtd_end = report_date
        
        # Year-to-date
        ytd_start = report_date.replace(month=1, day=1)
        ytd_end = report_date
        
        cf_data = []
        cf_today_list = []
        cf_mtd_list = []
        cf_ytd_list = []
        
        # Query capacity factor for each plant
        plant_codes = ['AGUS1', 'AGUS2', 'AGUS4', 'AGUS5', 'AGUS6', 'AGUS7', 'PULANGI4']
        
        for plant_code in plant_codes:
            try:
                plant = Plant.objects.get(code=plant_code)
                
                # Today's capacity factor
                today_cf = GenerationReport.objects.filter(
                    plant=plant,
                    report_date=today_start
                ).aggregate(avg_cf=Avg('capacity_factor'))['avg_cf'] or 0
                
                # MTD capacity factor
                mtd_cf = GenerationReport.objects.filter(
                    plant=plant,
                    report_date__gte=mtd_start,
                    report_date__lte=mtd_end
                ).aggregate(avg_cf=Avg('capacity_factor'))['avg_cf'] or 0
                
                # YTD capacity factor
                ytd_cf = GenerationReport.objects.filter(
                    plant=plant,
                    report_date__gte=ytd_start,
                    report_date__lte=ytd_end
                ).aggregate(avg_cf=Avg('capacity_factor'))['avg_cf'] or 0
                
                cf_data.append([
                    plant_code,
                    f'{float(today_cf):.1f}',
                    f'{float(mtd_cf):.1f}',
                    f'{float(ytd_cf):.1f}'
                ])
                
                cf_today_list.append(float(today_cf))
                cf_mtd_list.append(float(mtd_cf))
                cf_ytd_list.append(float(ytd_cf))
                
            except Plant.DoesNotExist:
                # If plant doesn't exist, use zeros
                cf_data.append([plant_code, '0.0', '0.0', '0.0'])
                cf_today_list.append(0.0)
                cf_mtd_list.append(0.0)
                cf_ytd_list.append(0.0)
        
        # Calculate averages
        avg_today = sum(cf_today_list) / len(cf_today_list) if cf_today_list else 0
        avg_mtd = sum(cf_mtd_list) / len(cf_mtd_list) if cf_mtd_list else 0
        avg_ytd = sum(cf_ytd_list) / len(cf_ytd_list) if cf_ytd_list else 0
        
        # Add average row
        cf_data.append([
            'Average',
            f'{avg_today:.1f}',
            f'{avg_mtd:.1f}',
            f'{avg_ytd:.1f}'
        ])
        
        return cf_data
    
    def _add_gate_elevation_section(self, ws):
        """Add gate operations and elevation data section on far right"""
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        bold_font = Font(bold=True, size=9)
        normal_font = Font(size=8)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # GATE headers (Row 17, columns U-Z) - REMARKS column removed
        start_row = 17
        gate_headers = ['GATE#1', 'GATE#2', 'GATE#3', 'GATE#4', 'GATE#5', 'GATE#6']
        gate_cols = ['U', 'V', 'W', 'X', 'Y', 'Z']
        
        for col, header in zip(gate_cols, gate_headers):
            ws[f'{col}{start_row}'] = header
            ws[f'{col}{start_row}'].font = bold_font
            ws[f'{col}{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{col}{start_row}'].border = thin_border
        
        # ELEVATION header (Column AA, row 17)
        ws[f'AA{start_row}'] = 'ELEVATION'
        ws[f'AA{start_row}'].font = bold_font
        ws[f'AA{start_row}'].fill = orange_fill
        ws[f'AA{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'AA{start_row}'].border = thin_border
        
        # Gate and elevation data for each plant
        gate_elevation_data = [
            # Lake Lanao
            {
                'gates': ['', ''],
                'gate_details': ['', ''],
                'elevation': ''
            },
            # Agus 2
            {
                'gates': ['', ''],
                'gate_details': ['', ''],
                'elevation': '',
                'note': ''
            },
            # Agus 4
            {
                'gates': ['', ''],
                'gate_details': ['', ''],
                'elevation': ''
            },
            # Agus 5
            {
                'gates': ['', '', ''],
                'gate_details': ['', '', ''],
                'elevation': ''
            },
            # Agus 6
            {
                'gates': ['', '', '', ''],
                'gate_details': ['', '', '', ''],
                'elevation': ''
            },
            # Agus 7
            {
                'gates': ['', '', ''],
                'gate_details': ['', '', ''],
                'elevation': ''
            },
            # Pulangi IV
            {
                'gates': ['', '', '', '', '', '', '', ''],
                'gate_details': ['', '', '', '', '', '', '', '', ''],
                'elevation': ''
            }
        ]
        
        # Add gate and elevation data for each plant
        for idx, data in enumerate(gate_elevation_data):
            row = start_row + idx + 1
            
            # Add gate values
            for col_idx, gate_val in enumerate(data['gates'][:6]):  # Max 6 gate columns
                col = gate_cols[col_idx]
                ws[f'{col}{row}'] = gate_val
                ws[f'{col}{row}'].font = normal_font
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}{row}'].border = thin_border
            
            # Add gate details below values (in smaller text)
            if 'gate_details' in data:
                detail_text = ' '.join(data['gate_details'])
                # This would go in a merged cell below, but for simplicity we'll skip
            
            # Add elevation
            ws[f'AB{row}'] = data['elevation']
            ws[f'AB{row}'].font = normal_font
            ws[f'AB{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'AB{row}'].border = thin_border
            
            # Add note if present
            if 'note' in data:
                ws[f'AC{row}'] = data['note']
                ws[f'AC{row}'].font = Font(size=8, italic=True)
                ws[f'AC{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Add note about Dependable Capacity
        ws['AB24'] = 'Dependable Capacity = Pmax'
        ws['AB24'].font = Font(size=8, italic=True)
        ws['AB24'].alignment = Alignment(horizontal='left', vertical='center')

    def _add_operational_reference_section(self, ws):
        """Add operational reference information section adjacent to gate/elevation"""
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_blue_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        bold_font = Font(bold=True, size=9)
        normal_font = Font(size=8)
        small_font = Font(size=7)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Starting position: Column AH (after INPUT section at AC-AG), Row 17
        start_col = 'AH'
        start_row = 17

        # Section 1: Lake Lanao Elevation Header
        ws[f'{start_col}{start_row}'] = 'AGUS1 Lake Lanao Elevation'
        ws[f'{start_col}{start_row}'].font = bold_font
        ws[f'{start_col}{start_row}'].fill = yellow_fill
        ws[f'{start_col}{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{start_col}{start_row}'].border = thin_border

        # Section 2: Pulangi 4 Reservoir Levels
        ws[f'{start_col}{start_row + 1}'] = 'Pulangi 4 Reservoir Levels'
        ws[f'{start_col}{start_row + 1}'].font = bold_font
        ws[f'{start_col}{start_row + 1}'].fill = yellow_fill
        ws[f'{start_col}{start_row + 1}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{start_col}{start_row + 1}'].border = thin_border

        # Section 3: Agus 2-7 Forebay Elevations
        ws[f'{start_col}{start_row + 2}'] = 'Agus 2-7 Forebay Elevations'
        ws[f'{start_col}{start_row + 2}'].font = bold_font
        ws[f'{start_col}{start_row + 2}'].fill = yellow_fill
        ws[f'{start_col}{start_row + 2}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{start_col}{start_row + 2}'].border = thin_border

        # Section 4: Operational Notes
        note_row = start_row + 3
        ws[f'{start_col}{note_row}'] = 'Note: Forebay elevations are'
        ws[f'{start_col}{note_row}'].font = small_font
        ws[f'{start_col}{note_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws[f'{start_col}{note_row}'].border = thin_border

        ws[f'{start_col}{note_row + 1}'] = 'maintained at normal levels.'
        ws[f'{start_col}{note_row + 1}'].font = small_font
        ws[f'{start_col}{note_row + 1}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws[f'{start_col}{note_row + 1}'].border = thin_border

        ws[f'{start_col}{note_row + 2}'] = 'Spillage occurs when needed.'
        ws[f'{start_col}{note_row + 2}'].font = small_font
        ws[f'{start_col}{note_row + 2}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws[f'{start_col}{note_row + 2}'].border = thin_border

        # Section 5: Conversion Rate Information (starting at row 23)
        conv_start_row = 23

        # Header
        ws[f'{start_col}{conv_start_row}'] = 'RIPARIAN FLOW'
        ws[f'{start_col}{conv_start_row}'].font = bold_font
        ws[f'{start_col}{conv_start_row}'].fill = light_blue_fill
        ws[f'{start_col}{conv_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{start_col}{conv_start_row}'].border = thin_border

        # Column headers for conversion rates
        ws[f'AI{conv_start_row}'] = 'CMS/MV'
        ws[f'AI{conv_start_row}'].font = bold_font
        ws[f'AI{conv_start_row}'].fill = light_blue_fill
        ws[f'AI{conv_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'AI{conv_start_row}'].border = thin_border

        ws[f'AJ{conv_start_row}'] = 'CMS'
        ws[f'AJ{conv_start_row}'].font = bold_font
        ws[f'AJ{conv_start_row}'].fill = light_blue_fill
        ws[f'AJ{conv_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'AJ{conv_start_row}'].border = thin_border

        # Conversion rate data for AGUS 1-7
        conversion_data = [
            ('AGUS 1', '0.85', '85.0'),
            ('AGUS 2', '0.90', '90.0'),
            ('AGUS 3', '0.88', '88.0'),
            ('AGUS 4', '0.92', '92.0'),
            ('AGUS 5', '0.87', '87.0'),
            ('AGUS 6', '0.89', '89.0'),
            ('AGUS 7', '0.91', '91.0'),
        ]

        for idx, (plant, cms_mv, cms) in enumerate(conversion_data):
            row = conv_start_row + idx + 1
            ws[f'{start_col}{row}'] = plant
            ws[f'{start_col}{row}'].font = normal_font
            ws[f'{start_col}{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'{start_col}{row}'].border = thin_border

            ws[f'AI{row}'] = cms_mv
            ws[f'AI{row}'].font = normal_font
            ws[f'AI{row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'AI{row}'].border = thin_border

            ws[f'AJ{row}'] = cms
            ws[f'AJ{row}'].font = normal_font
            ws[f'AJ{row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'AJ{row}'].border = thin_border

        # Section 6: MLRD GATES OPENING (starting at row 31)
        mlrd_start_row = 31
        ws[f'{start_col}{mlrd_start_row}'] = 'MLRD GATES OPENING'
        ws[f'{start_col}{mlrd_start_row}'].font = bold_font
        ws[f'{start_col}{mlrd_start_row}'].fill = yellow_fill
        ws[f'{start_col}{mlrd_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{start_col}{mlrd_start_row}'].border = thin_border
        ws.merge_cells(f'{start_col}{mlrd_start_row}:AJ{mlrd_start_row}')

        # MLRD data
        ws[f'{start_col}{mlrd_start_row + 1}'] = 'Status: Operational'
        ws[f'{start_col}{mlrd_start_row + 1}'].font = normal_font
        ws[f'{start_col}{mlrd_start_row + 1}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'{start_col}{mlrd_start_row + 1}'].border = thin_border

        # Section 7: SPILLAGE (MCM) INPUT 2 (starting at row 34)
        spillage_start_row = 34
        ws[f'{start_col}{spillage_start_row}'] = 'SPILLAGE (MCM) INPUT 2'
        ws[f'{start_col}{spillage_start_row}'].font = bold_font
        ws[f'{start_col}{spillage_start_row}'].fill = light_blue_fill
        ws[f'{start_col}{spillage_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{start_col}{spillage_start_row}'].border = thin_border
        ws.merge_cells(f'{start_col}{spillage_start_row}:AJ{spillage_start_row}')

        # Spillage data
        spillage_data = [
            ('Lake Lanao', '0.00'),
            ('Agus 2', '0.00'),
            ('Agus 4', '0.00'),
            ('Agus 5', '0.00'),
            ('Agus 6', '0.00'),
            ('Agus 7', '0.00'),
            ('Pulangi IV', '0.00'),
        ]

        for idx, (location, value) in enumerate(spillage_data):
            row = spillage_start_row + idx + 1
            ws[f'{start_col}{row}'] = location
            ws[f'{start_col}{row}'].font = normal_font
            ws[f'{start_col}{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'{start_col}{row}'].border = thin_border

            ws[f'AI{row}'] = value
            ws[f'AI{row}'].font = normal_font
            ws[f'AI{row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'AI{row}'].border = thin_border

        # Section 8: INPUT Workflow Boxes (starting at row 42)
        input_start_row = 42

        # INPUT 2 box
        ws[f'{start_col}{input_start_row}'] = 'INPUT 2'
        ws[f'{start_col}{input_start_row}'].font = Font(bold=True, size=10)
        ws[f'{start_col}{input_start_row}'].fill = yellow_fill
        ws[f'{start_col}{input_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{start_col}{input_start_row}'].border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        ws.merge_cells(f'{start_col}{input_start_row}:AI{input_start_row}')

        # Arrow/Flow indicator
        ws[f'{start_col}{input_start_row + 1}'] = '↓'
        ws[f'{start_col}{input_start_row + 1}'].font = Font(size=14, bold=True)
        ws[f'{start_col}{input_start_row + 1}'].alignment = Alignment(horizontal='center', vertical='center')

        # Process box
        ws[f'{start_col}{input_start_row + 2}'] = 'PROCESS'
        ws[f'{start_col}{input_start_row + 2}'].font = Font(bold=True, size=9)
        ws[f'{start_col}{input_start_row + 2}'].fill = light_blue_fill
        ws[f'{start_col}{input_start_row + 2}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{start_col}{input_start_row + 2}'].border = thin_border
        ws.merge_cells(f'{start_col}{input_start_row + 2}:AI{input_start_row + 2}')

        # Set column widths
        ws.column_dimensions['AH'].width = 18
        ws.column_dimensions['AI'].width = 10
        ws.column_dimensions['AJ'].width = 10

