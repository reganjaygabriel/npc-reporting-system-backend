"""
Daily Plant Status Report Excel Generator
Generates formatted Excel report matching NPC's Daily Plant Status format
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
from decimal import Decimal
import os


class DailyStatusExporter:
    """Generate Daily Plant Status Report in Excel format"""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Daily Plant Status"
        
        # Define colors
        self.header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.total_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # Define fonts
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        self.bold_font = Font(name='Arial', size=10, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        
        # Define borders
        thin_border = Side(style='thin', color='000000')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def generate_report(self, report_date, plant_data, output_path):
        """
        Generate the complete Daily Plant Status Report
        
        Args:
            report_date: Date of the report (datetime object)
            plant_data: List of dictionaries with plant status data
            output_path: Path to save the Excel file
        
        Returns:
            Path to generated file
        """
        # Add header
        self._add_header(report_date)
        
        # Add plant status table
        self._add_plant_status_table(plant_data)
        
        # Add lake elevations
        self._add_lake_elevations()
        
        # Add chart
        self._add_capacity_chart(plant_data)
        
        # Add footer
        self._add_footer()
        
        # Adjust column widths
        self._adjust_column_widths()
        
        # Save file
        self.wb.save(output_path)
        return output_path
    
    def _add_header(self, report_date):
        """Add report header with title and date"""
        # Logo placeholder (row 1-3)
        self.ws.merge_cells('A1:A3')
        self.ws['A1'] = 'NPC'
        self.ws['A1'].font = Font(name='Arial', size=16, bold=True)
        self.ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Title
        self.ws.merge_cells('B1:F1')
        self.ws['B1'] = 'NATIONAL POWER CORPORATION'
        self.ws['B1'].font = self.title_font
        self.ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        
        self.ws.merge_cells('B2:F2')
        self.ws['B2'] = 'MINDANAO GENERATION'
        self.ws['B2'].font = self.title_font
        self.ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Date and time header
        self.ws.merge_cells('A5:F5')
        date_str = report_date.strftime('%A, %B %d, %Y')
        self.ws['A5'] = f'DAILY PLANT STATUS\nas of 12:00 NN    {date_str}'
        self.ws['A5'].font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        self.ws['A5'].fill = self.header_fill
        self.ws['A5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.ws.row_dimensions[5].height = 30
    
    def _add_plant_status_table(self, plant_data):
        """Add the main plant status table"""
        # Table headers (row 7)
        headers = ['PLANT\nUNIT NO.', 'Rated\nCapacity\n(MW)', 'Dependable\nCapacity\n(MW)', 
                   'Load at\n1200H\n(MW)', 'REMARKS']
        
        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=7, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        self.ws.row_dimensions[7].height = 45
        
        # Add plant data
        current_row = 8
        npc_total_capacity = 0
        npc_total_dependable = 0
        npc_total_load = 0
        
        for plant in plant_data:
            # Plant header row
            self.ws.cell(row=current_row, column=1).value = plant['name']
            self.ws.cell(row=current_row, column=1).font = self.bold_font
            self.ws.cell(row=current_row, column=1).fill = self.subheader_fill
            
            self.ws.cell(row=current_row, column=2).value = float(plant['total_capacity'])
            self.ws.cell(row=current_row, column=2).font = self.bold_font
            self.ws.cell(row=current_row, column=2).fill = self.subheader_fill
            self.ws.cell(row=current_row, column=2).number_format = '0.00'
            
            self.ws.cell(row=current_row, column=3).value = float(plant['dependable_capacity'])
            self.ws.cell(row=current_row, column=3).font = self.bold_font
            self.ws.cell(row=current_row, column=3).fill = self.subheader_fill
            self.ws.cell(row=current_row, column=3).number_format = '0.00'
            
            self.ws.cell(row=current_row, column=4).value = float(plant['total_load'])
            self.ws.cell(row=current_row, column=4).font = self.bold_font
            self.ws.cell(row=current_row, column=4).fill = self.subheader_fill
            self.ws.cell(row=current_row, column=4).number_format = '0.00'
            
            for col in range(1, 6):
                self.ws.cell(row=current_row, column=col).border = self.border
            
            current_row += 1
            
            # Unit rows
            for unit in plant['units']:
                self.ws.cell(row=current_row, column=1).value = f"  Unit {unit['number']}"
                self.ws.cell(row=current_row, column=1).alignment = Alignment(indent=2)
                
                self.ws.cell(row=current_row, column=2).value = float(unit['capacity'])
                self.ws.cell(row=current_row, column=2).number_format = '0.0'
                
                self.ws.cell(row=current_row, column=3).value = float(unit['dependable'])
                self.ws.cell(row=current_row, column=3).number_format = '0.0'
                
                self.ws.cell(row=current_row, column=4).value = float(unit['load'])
                self.ws.cell(row=current_row, column=4).number_format = '0.00'
                
                # Color code load (red if 0)
                if float(unit['load']) == 0:
                    self.ws.cell(row=current_row, column=4).font = Font(color="FF0000", bold=True)
                
                self.ws.cell(row=current_row, column=5).value = unit['remarks']
                self.ws.cell(row=current_row, column=5).alignment = Alignment(wrap_text=True)
                
                for col in range(1, 6):
                    self.ws.cell(row=current_row, column=col).border = self.border
                
                current_row += 1
            
            # Update totals
            npc_total_capacity += float(plant['total_capacity'])
            npc_total_dependable += float(plant['dependable_capacity'])
            npc_total_load += float(plant['total_load'])
        
        # NPC TOTAL row
        self.ws.cell(row=current_row, column=1).value = 'NPC TOTAL'
        self.ws.cell(row=current_row, column=1).font = self.bold_font
        self.ws.cell(row=current_row, column=1).fill = self.total_fill
        
        self.ws.cell(row=current_row, column=2).value = npc_total_capacity
        self.ws.cell(row=current_row, column=2).font = self.bold_font
        self.ws.cell(row=current_row, column=2).fill = self.total_fill
        self.ws.cell(row=current_row, column=2).number_format = '0.00'
        
        self.ws.cell(row=current_row, column=3).value = npc_total_dependable
        self.ws.cell(row=current_row, column=3).font = self.bold_font
        self.ws.cell(row=current_row, column=3).fill = self.total_fill
        self.ws.cell(row=current_row, column=3).number_format = '0.00'
        
        self.ws.cell(row=current_row, column=4).value = npc_total_load
        self.ws.cell(row=current_row, column=4).font = self.bold_font
        self.ws.cell(row=current_row, column=4).fill = self.total_fill
        self.ws.cell(row=current_row, column=4).number_format = '0.00'
        
        for col in range(1, 6):
            self.ws.cell(row=current_row, column=col).border = self.border
        
        return current_row + 2
    
    def _add_lake_elevations(self):
        """Add lake elevation data table"""
        start_row = self.ws.max_row + 2
        
        # Note
        self.ws.cell(row=start_row, column=1).value = 'Notes:'
        self.ws.cell(row=start_row, column=1).font = self.bold_font
        start_row += 1
        
        self.ws.cell(row=start_row, column=1).value = '- Forecast inflow of Lake Lanao is stable and operating at Normal Stage.'
        start_row += 2
        
        # Lake elevations table
        self.ws.cell(row=start_row, column=1).value = 'LAKE LANAO ELEVATION'
        self.ws.cell(row=start_row, column=1).font = self.bold_font
        self.ws.cell(row=start_row, column=2).value = '701.2 m.a.s.l.'
        
        elevations = [
            ('AGUS 2 Forebay Elevation', '637.3 m.a.s.l.'),
            ('AGUS 4 Forebay Elevation', '358.5 m.a.s.l.'),
            ('AGUS 5 Forebay Elevation', '242.8 m.a.s.l.'),
            ('AGUS 6 Forebay Elevation', '199.8 m.a.s.l.'),
            ('AGUS 7 Forebay Elevation', '34.6 m.a.s.l.'),
            ('PULANGI IV Reservoir Level', '283.5 m.a.s.l.'),
        ]
        
        for label, value in elevations:
            start_row += 1
            self.ws.cell(row=start_row, column=1).value = label
            self.ws.cell(row=start_row, column=2).value = value
    
    def _add_capacity_chart(self, plant_data):
        """Add capacity comparison chart"""
        # Chart will be added in the space to the right of the table
        # This is a simplified version - full chart implementation would be more complex
        pass
    
    def _add_footer(self):
        """Add report footer with signature"""
        footer_row = self.ws.max_row + 3
        
        self.ws.cell(row=footer_row, column=5).value = 'Noted by:'
        self.ws.cell(row=footer_row, column=5).alignment = Alignment(horizontal='right')
        
        footer_row += 2
        self.ws.cell(row=footer_row, column=5).value = 'D.B. ESMADE JR.'
        self.ws.cell(row=footer_row, column=5).font = self.bold_font
        self.ws.cell(row=footer_row, column=5).alignment = Alignment(horizontal='right')
        
        footer_row += 1
        self.ws.cell(row=footer_row, column=5).value = 'Dept. Manager, GPD'
        self.ws.cell(row=footer_row, column=5).alignment = Alignment(horizontal='right')
    
    def _adjust_column_widths(self):
        """Adjust column widths for better readability"""
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 12
        self.ws.column_dimensions['C'].width = 12
        self.ws.column_dimensions['D'].width = 12
        self.ws.column_dimensions['E'].width = 50


def generate_daily_status_report(report_date=None, output_filename=None):
    """
    Generate Daily Plant Status Report from database
    
    Args:
        report_date: Date for the report (defaults to today)
        output_filename: Output file path (auto-generated if None)
    
    Returns:
        Path to generated Excel file
    """
    from reports.models import Plant, GenerationReport, Unit
    from django.utils import timezone
    
    if report_date is None:
        report_date = timezone.now().date()
    
    if output_filename is None:
        date_str = report_date.strftime('%Y%m%d')
        output_filename = f'DAILY_PLANT_STATUS_{date_str}.xlsx'
    
    # Fetch data from database
    plants = Plant.objects.filter(is_active=True).order_by('code')
    plant_data = []
    
    for plant in plants:
        units = Unit.objects.filter(plant=plant, is_active=True).order_by('unit_number')
        
        unit_data = []
        total_capacity = 0
        total_dependable = 0
        total_load = 0
        
        for unit in units:
            # Get latest generation report for this unit
            latest_report = GenerationReport.objects.filter(
                unit=unit,
                report_date=report_date
            ).first()
            
            if latest_report:
                load = float(latest_report.generation_kwh) / 1000  # Convert to MW
                remarks = latest_report.remarks or 'OPERATIONAL. Maximized with respect to ave. outflow.'
            else:
                load = 0.0
                remarks = 'No data available'
            
            unit_data.append({
                'number': unit.unit_number,
                'capacity': float(unit.capacity_mw),
                'dependable': float(unit.capacity_mw) * 0.9,  # Assume 90% dependable
                'load': load,
                'remarks': remarks
            })
            
            total_capacity += float(unit.capacity_mw)
            total_dependable += float(unit.capacity_mw) * 0.9
            total_load += load
        
        plant_data.append({
            'name': plant.name,
            'code': plant.code,
            'total_capacity': total_capacity,
            'dependable_capacity': total_dependable,
            'total_load': total_load,
            'units': unit_data
        })
    
    # Generate report
    exporter = DailyStatusExporter()
    output_path = exporter.generate_report(report_date, plant_data, output_filename)
    
    return output_path
